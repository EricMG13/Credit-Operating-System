"""
Excel pricing sheet ingestion.
Parses spread/YTW/DM columns and persists to market_data_runs.
"""

from __future__ import annotations

import hashlib
import io
from datetime import date
from uuid import UUID

import anyio
import openpyxl
import structlog
from minio import Minio
from sqlalchemy.ext.asyncio import AsyncSession

from core.config import get_settings
from db.models import Document, MarketDataRun

logger = structlog.get_logger()
settings = get_settings()

# Column name aliases for common broker sheet formats
COLUMN_ALIASES = {
    "spread": ["spread", "spread_bps", "z_spread", "oas"],
    "ytw": ["ytw", "ytw_pct", "yield_to_worst", "yield"],
    "dm": ["dm", "dm_bps", "discount_margin"],
    "instrument": ["cusip", "isin", "instrument", "bond", "loan", "ticker", "description"],
}


def _normalize_col(name: str, col_type: str) -> bool:
    return name.lower().replace(" ", "_") in COLUMN_ALIASES.get(col_type, [])


def _get_minio() -> Minio:
    return Minio(
        settings.minio_endpoint,
        access_key=settings.minio_access_key,
        secret_key=settings.minio_secret_key,
        secure=settings.minio_secure,
    )


async def ingest_pricing_sheet(
    content: bytes,
    filename: str,
    issuer_id: UUID,
    run_date: str,  # YYYY-MM-DD
    db: AsyncSession,
) -> dict:
    """
    Parse a pricing sheet XLSX and persist market_data_run rows.
    Handles common broker sheet column naming variants.
    """
    content_hash = hashlib.sha256(content).hexdigest()
    minio_key = f"{issuer_id}/pricing/{content_hash[:16]}_{filename}"

    minio_client = _get_minio()
    await anyio.to_thread.run_sync(
        lambda: minio_client.put_object(
            bucket_name=settings.minio_bucket_docs,
            object_name=minio_key,
            data=io.BytesIO(content),
            length=len(content),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )

    # Store Document record
    doc = Document(
        issuer_id=issuer_id,
        doc_type="PricingSheet",
        file_name=filename,
        minio_key=minio_key,
        content_hash=content_hash,
        fiscal_period=run_date[:7],  # YYYY-MM
    )
    db.add(doc)
    await db.flush()
    await db.refresh(doc)

    # Parse XLSX (openpyxl is sync/CPU-bound → read into plain tuples off-loop)
    def _read_workbook() -> tuple[list[str], list[tuple]]:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        hdrs = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        return hdrs, data_rows

    headers, data_rows = await anyio.to_thread.run_sync(_read_workbook)

    col_map = {}
    for i, h in enumerate(headers):
        for col_type in COLUMN_ALIASES:
            if _normalize_col(h, col_type):
                col_map[col_type] = i

    parsed_date = date.fromisoformat(run_date)
    rows_created = 0

    def _to_float(value: object) -> float | None:
        """Coerce a numeric cell to float; ignore blanks and non-numeric text."""
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            logger.warning("Non-numeric pricing cell skipped", value=value)
            return None

    for row in data_rows:
        if not any(cell is not None for cell in row):
            continue
        instrument = row[col_map["instrument"]] if "instrument" in col_map else None
        spread = _to_float(row[col_map["spread"]]) if "spread" in col_map else None
        ytw = _to_float(row[col_map["ytw"]]) if "ytw" in col_map else None
        dm = _to_float(row[col_map["dm"]]) if "dm" in col_map else None

        if spread is None and ytw is None and dm is None:
            continue

        mdr = MarketDataRun(
            issuer_id=issuer_id,
            run_date=parsed_date,
            instrument=str(instrument) if instrument is not None else None,
            spread_bps=spread,
            ytw_pct=ytw,
            dm_bps=dm,
            source_file=minio_key,
        )
        db.add(mdr)
        rows_created += 1

    logger.info("Pricing sheet ingested", rows=rows_created, issuer_id=str(issuer_id))

    return {
        "document_id": doc.id,
        "issuer_id": issuer_id,
        "minio_key": minio_key,
        "chunks_created": 0,
        "message": f"Parsed {rows_created} pricing rows from {filename}.",
    }
