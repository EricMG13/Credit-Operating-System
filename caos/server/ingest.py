"""Document intake: content sniffing, vault storage, text chunking.

Files land in the vault directory (a mounted volume in production, local disk in
dev) and their extracted text is chunked into document_chunks rows for
downstream retrieval.
"""

from __future__ import annotations

import asyncio
import io
import logging
import multiprocessing
import re
import shlex
import subprocess
import tempfile
import uuid
from pathlib import Path
from typing import Callable, List, Optional, TypeVar

from fastapi import HTTPException, UploadFile

from config import get_settings
from xlsx_safety import PACKAGE_LIMIT_CODES, XlsxPackageError, validate_xlsx_package

logger = logging.getLogger("caos.ingest")

_PDF_MAGIC = b"%PDF-"
_OOXML_MAGIC = b"PK\x03\x04"

CHUNK_CHARS = 2400
CHUNK_OVERLAP = 240

# Surfaced (not raised) when a document parses to zero chunks. Upload stays
# intentionally lenient — a scanned/encrypted/empty file still vaults — but a
# silent ``chunks_created: 0`` reads as success and an analyst could build a view
# on a document the engine never actually read. This warning makes that explicit.
NO_CHUNKS_WARNING = (
    "0 chunks extracted — document may be scanned/encrypted/empty; "
    "it is vaulted but not searchable. Re-upload a text-based copy if it must be analyzed."
)

_READ_CHUNK = 1024 * 1024  # 1 MB
_MAX_XLSX_NONEMPTY_CELLS = 500_000
_MAX_XLSX_CELL_TEXT = 4_096
_MAX_XLSX_EXTRACTED_CHARS = 8_000_000
_T = TypeVar("_T")


def _parser_process_entry(send_conn, parser: Callable[..., _T], args: tuple) -> None:
    """Run one trusted parser in an isolated child and return its result/error."""
    try:
        send_conn.send((True, parser(*args)))
    except BaseException as exc:  # noqa: BLE001 — marshal parser failures to parent
        try:
            send_conn.send((False, exc))
        except Exception:  # pragma: no cover — only for an unpickleable exception
            send_conn.send((False, RuntimeError(f"{type(exc).__name__}: {exc}")))
    finally:
        send_conn.close()


def _run_parser_process(parser: Callable[..., _T], args: tuple, timeout: float) -> _T:
    """Run a parser in a killable spawned process with a hard deadline."""
    ctx = multiprocessing.get_context("spawn")
    recv_conn, send_conn = ctx.Pipe(duplex=False)
    process = ctx.Process(
        target=_parser_process_entry,
        args=(send_conn, parser, args),
        name="caos-upload-parser",
        daemon=True,
    )
    try:
        process.start()
        send_conn.close()
        if not recv_conn.poll(timeout):
            process.terminate()
            process.join(timeout=1)
            if process.is_alive():
                process.kill()
                process.join(timeout=1)
            raise TimeoutError("upload parser exceeded deadline")
        try:
            ok, value = recv_conn.recv()
        except EOFError as exc:
            raise RuntimeError(
                f"upload parser exited without a result (exit={process.exitcode})"
            ) from exc
        process.join(timeout=1)
        if not ok:
            raise value
        return value
    finally:
        recv_conn.close()
        send_conn.close()
        if process.is_alive():
            process.terminate()
            process.join(timeout=1)


async def parse_bounded(parser: Callable[..., _T], *args) -> _T:
    """Run a synchronous parser off-loop in a process that can be terminated."""
    try:
        return await asyncio.to_thread(
            _run_parser_process,
            parser,
            args,
            float(get_settings().upload_parse_timeout_s),
        )
    except TimeoutError as exc:
        raise HTTPException(
            422,
            "Upload parsing exceeded the configured time limit; split or simplify the file.",
        ) from exc


async def read_capped(file: UploadFile, *, max_bytes: int | None = None) -> bytes:
    """Read the upload incrementally, aborting as soon as it exceeds the cap.

    Reading the whole body before checking would let an oversized request
    occupy its full size in memory before the 413.
    """
    settings = get_settings()
    configured_limit = settings.max_upload_mb * 1024 * 1024
    limit = configured_limit if max_bytes is None else min(configured_limit, max_bytes)
    buf = bytearray()
    while chunk := await file.read(_READ_CHUNK):
        buf.extend(chunk)
        if len(buf) > limit:
            if limit % (1024 * 1024) == 0:
                label = f"{limit // (1024 * 1024)} MB"
            else:
                label = f"{limit} byte"
            raise HTTPException(413, f"File exceeds the {label} limit")
    if not buf:
        raise HTTPException(400, "Empty upload")
    return bytes(buf)


def sniff_pdf(content: bytes) -> None:
    if not content.startswith(_PDF_MAGIC):
        raise HTTPException(400, "Uploaded file is not a valid PDF.")


def sniff_xlsx(content: bytes) -> None:
    try:
        validate_xlsx_package(content)
    except XlsxPackageError as exc:
        if exc.code in PACKAGE_LIMIT_CODES:
            raise HTTPException(413, exc.message) from exc
        if content.startswith(_OOXML_MAGIC) and exc.code == "invalid_ooxml":
            raise HTTPException(
                400, "ZIP container is not an Excel workbook (missing OOXML parts)."
            ) from exc
        if exc.code == "invalid_ooxml":
            raise HTTPException(400, "Uploaded file is not a valid .xlsx workbook.") from exc
        raise HTTPException(400, exc.message) from exc


def store(content: bytes, file_name: str) -> str:
    """Write the raw file into the vault; returns the storage key."""
    settings = get_settings()
    safe = re.sub(r"[^A-Za-z0-9._-]", "_", Path(file_name).name) or "upload.bin"
    key = f"{uuid.uuid4().hex}/{safe}"
    path = Path(settings.caos_storage_dir) / key
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)
    return key


def remove_uncommitted(key: str) -> None:
    """Remove only a unique object created by a failed ingestion transaction."""
    settings = get_settings()
    parts = Path(key).parts
    if len(parts) != 2 or re.fullmatch(r"[0-9a-f]{32}", parts[0]) is None:
        raise ValueError("Refusing to remove an invalid ingestion vault key.")
    root = Path(settings.caos_storage_dir).resolve()
    path = (root / key).resolve()
    if not path.is_relative_to(root):
        raise ValueError("Ingestion storage key escaped the configured vault.")
    path.unlink(missing_ok=True)
    try:
        path.parent.rmdir()
    except OSError:
        pass


def _markitdown_text(content: bytes, filename: str) -> Optional[str]:
    """Structure-preserving extraction via an external markitdown CLI, when one
    is configured (CAOS_MARKITDOWN_CMD). markitdown needs Python 3.10+, so it
    runs out-of-process — the server keeps its own interpreter. Returns the
    Markdown, or None to fall back to the built-in extractors (also on any
    failure/timeout, so a misconfigured command never blocks an upload)."""
    settings = get_settings()
    cmd = settings.markitdown_cmd
    if not cmd:
        return None
    try:
        with tempfile.NamedTemporaryFile(suffix=Path(filename).suffix or ".bin") as tmp:
            tmp.write(content)
            tmp.flush()
            proc = subprocess.run(
                [*shlex.split(cmd), tmp.name],
                capture_output=True,
                timeout=settings.markitdown_timeout_s,
            )
        if proc.returncode == 0:
            text = proc.stdout.decode("utf-8", "replace").strip()
            if len(text) > settings.max_pdf_extracted_chars:
                raise HTTPException(413, "Extracted document text exceeds the configured limit.")
            return text or None
        logger.warning(
            "markitdown rc=%s — falling back. stderr=%s",
            proc.returncode,
            proc.stderr.decode("utf-8", "replace")[:200],
        )
    except (subprocess.SubprocessError, OSError) as exc:
        logger.warning("markitdown unavailable (%s) — falling back.", exc)
    return None


def _ocrmypdf_text(content: bytes) -> str:
    """Last-resort OCR for scanned/image PDFs via an external ocrmypdf CLI
    (CAOS_OCRMYPDF_CMD). ocrmypdf wraps Tesseract (a heavy native dep), so it
    runs out-of-process and stays out of the server image. Writes the recognized
    text to a sidecar file and returns it; returns "" on any failure/timeout/
    missing command so a scanned upload still vaults, just produces no chunks."""
    settings = get_settings()
    cmd = settings.ocrmypdf_cmd
    if not cmd:
        return ""
    try:
        with tempfile.TemporaryDirectory() as td:
            src = Path(td) / "in.pdf"
            out = Path(td) / "out.pdf"  # discarded; we only want the sidecar text
            sidecar = Path(td) / "text.txt"
            src.write_bytes(content)
            proc = subprocess.run(
                [*shlex.split(cmd), "--force-ocr", "--sidecar", str(sidecar),
                 str(src), str(out)],
                capture_output=True,
                timeout=settings.ocrmypdf_timeout_s,
            )
            if proc.returncode == 0 and sidecar.exists():
                text = sidecar.read_text("utf-8", "replace").strip()
                if len(text) > settings.max_pdf_extracted_chars:
                    raise HTTPException(413, "OCR text exceeds the configured limit.")
                return text
            logger.warning(
                "ocrmypdf rc=%s — no OCR text. stderr=%s",
                proc.returncode,
                proc.stderr.decode("utf-8", "replace")[:200],
            )
    except (subprocess.SubprocessError, OSError) as exc:
        logger.warning("ocrmypdf unavailable (%s) — no OCR text.", exc)
    return ""


def extract_pdf_text(content: bytes, filename: str = "upload.pdf") -> tuple[str, bool]:
    """Returns (text, used_ocr). ``used_ocr`` is True only when the OCR
    last-resort lane actually produced the returned text (D1) — callers use
    it to tag chunk provenance so a lower-fidelity OCR read is discountable,
    never silently indistinguishable from a native text-layer extraction."""
    md = _markitdown_text(content, filename)
    if md is not None:
        return md, False

    from pypdf import PdfReader

    try:
        reader = PdfReader(io.BytesIO(content))
        settings = get_settings()
        if len(reader.pages) > settings.max_pdf_pages:
            raise HTTPException(
                413, f"PDF exceeds the {settings.max_pdf_pages}-page extraction limit."
            )
        parts: list[str] = []
        extracted_chars = 0
        for page in reader.pages:
            page_text = page.extract_text() or ""
            extracted_chars += len(page_text)
            if extracted_chars > settings.max_pdf_extracted_chars:
                raise HTTPException(413, "Extracted PDF text exceeds the configured limit.")
            parts.append(page_text)
        text = "\n".join(parts)
    except HTTPException:
        raise
    except Exception:
        text = ""  # scanned / encrypted PDFs vault fine, just produce no chunks
    if text.strip():
        return text, False
    # No text layer → try OCR before giving up (scanned/image PDF).
    ocr_text = _ocrmypdf_text(content)
    return ocr_text, bool(ocr_text.strip())


def extract_xlsx_text(content: bytes, filename: str = "upload.xlsx") -> str:
    # Validate before either extraction engine sees attacker-controlled XML.
    # Routes call sniff_xlsx after AV scanning too; this local gate keeps direct
    # parser callers from bypassing the same package/resource policy.
    sniff_xlsx(content)
    md = _markitdown_text(content, filename)
    if md is not None:
        return md

    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return ""
    lines: List[str] = []
    nonempty_cells = 0
    extracted_chars = 0
    for ws in wb.worksheets:
        lines.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = []
            for cell in row:
                if cell is None:
                    continue
                nonempty_cells += 1
                if nonempty_cells > _MAX_XLSX_NONEMPTY_CELLS:
                    raise HTTPException(413, "Workbook contains too many non-empty cells.")
                value = str(cell)
                if len(value) > _MAX_XLSX_CELL_TEXT:
                    value = value[:_MAX_XLSX_CELL_TEXT]
                extracted_chars += len(value)
                if extracted_chars > _MAX_XLSX_EXTRACTED_CHARS:
                    raise HTTPException(413, "Workbook extracted text exceeds the safe limit.")
                cells.append(value)
            if cells:
                lines.append("\t".join(cells))
    wb.close()
    return "\n".join(lines)


def extract_text(content: bytes, filename: str = "download.bin") -> str:
    """Generic extraction for content of any supported type, used by the EDGAR
    retrieval lane (exhibits are usually .htm). Tries the configured markitdown
    CLI first (handles HTML/PDF/Office), then falls back: PDF → pypdf; HTML →
    crude tag strip; otherwise decoded text."""
    md = _markitdown_text(content, filename)
    if md is not None:
        return md

    if content[:5] == _PDF_MAGIC:
        from pypdf import PdfReader

        try:
            reader = PdfReader(io.BytesIO(content))
            return "\n".join((page.extract_text() or "") for page in reader.pages)
        except Exception:
            return ""

    text = content.decode("utf-8", "replace")
    if filename.lower().endswith((".htm", ".html")) or "<html" in text[:2000].lower():
        text = re.sub(r"(?is)<(script|style).*?</\1>", " ", text)
        text = re.sub(r"(?s)<[^>]+>", " ", text)
        text = re.sub(r"&nbsp;", " ", text).replace("&amp;", "&")
        text = re.sub(r"[ \t]+", " ", text)
        text = re.sub(r"\n\s*\n\s*\n+", "\n\n", text)
    return text.strip()


def chunk_text(text: str) -> List[str]:  # noqa: C901 - pre-existing tokenizer complexity
    text = text.strip()
    if not text:
        return []

    import tiktoken
    try:
        encoding = tiktoken.get_encoding("cl100k_base")
    except Exception:
        class MockEncoding:
            def encode(self, s):
                return s.split()
            def decode(self, ids):
                return " ".join(ids)
        encoding = MockEncoding()

    max_tokens = 512
    overlap_tokens = 64

    paragraphs = text.split("\n\n")
    blocks = []
    
    for p in paragraphs:
        p = p.strip()
        if not p:
            continue
        p_tokens = len(encoding.encode(p))
        if p_tokens <= max_tokens:
            blocks.append((p, p_tokens))
        else:
            lines = p.split("\n")
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                l_tokens = len(encoding.encode(line))
                if l_tokens <= max_tokens:
                    blocks.append((line, l_tokens))
                else:
                    words = line.split(" ")
                    current_word_chunk = []
                    current_word_tokens = 0
                    for word in words:
                        if not word:
                            continue
                        w_tokens = len(encoding.encode(" " + word))
                        if current_word_tokens + w_tokens > max_tokens:
                            if current_word_chunk:
                                word_text = " ".join(current_word_chunk)
                                blocks.append((word_text, current_word_tokens))
                            current_word_chunk = [word]
                            current_word_tokens = w_tokens
                        else:
                            current_word_chunk.append(word)
                            current_word_tokens += w_tokens
                    if current_word_chunk:
                        word_text = " ".join(current_word_chunk)
                        blocks.append((word_text, current_word_tokens))

    chunks = []
    current_chunk_blocks = []
    current_tokens = 0

    for block_text, block_tokens in blocks:
        if current_tokens + block_tokens > max_tokens:
            if current_chunk_blocks:
                chunks.append("\n\n".join(current_chunk_blocks))
                
                # Keep trailing blocks for overlap
                overlap_blocks = []
                overlap_tokens_sum = 0
                for b_txt in reversed(current_chunk_blocks):
                    b_tok = len(encoding.encode(b_txt))
                    if overlap_tokens_sum + b_tok <= overlap_tokens:
                        overlap_blocks.insert(0, b_txt)
                        overlap_tokens_sum += b_tok
                    else:
                        break
                current_chunk_blocks = overlap_blocks
                current_tokens = overlap_tokens_sum
            else:
                chunks.append(block_text)
                current_chunk_blocks = []
                current_tokens = 0
                continue
        
        current_chunk_blocks.append(block_text)
        current_tokens += block_tokens

    if current_chunk_blocks:
        chunks.append("\n\n".join(current_chunk_blocks))

    return [c.strip() for c in chunks if c.strip()]
