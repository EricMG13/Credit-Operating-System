"""Fail-closed Bloomberg-style XLSX preview for immutable market snapshots.

This module is deliberately pure: it never writes the vault or database.  The
commit route reuses the same parser over the same workbook bytes and mapping so
client-returned preview rows can never become an input authority.
"""

from __future__ import annotations

import hashlib
import io
import json
import math
import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, time, timezone
from pathlib import PurePosixPath
from typing import Any, Literal, Optional
from xml.etree import ElementTree

from pydantic import BaseModel, ConfigDict, Field, ValidationError, field_validator


MAX_PACKAGE_MEMBERS = 5_000
MAX_PACKAGE_UNCOMPRESSED_BYTES = 128 * 1024 * 1024
MAX_PACKAGE_MEMBER_BYTES = 64 * 1024 * 1024
MAX_COMPRESSION_RATIO = 100
MAX_SHEETS = 20
MAX_ROWS_PER_SHEET = 25_000
MAX_COLUMNS_PER_SHEET = 128
MAX_NONEMPTY_CELLS = 500_000
MAX_CELL_TEXT = 4_096
MAX_FORMULA_TEXT = 8_192
MAX_ISSUES = 500
MAX_PREVIEW_ROWS = 100
HEADER_SCAN_ROWS = 25

CANONICAL_REQUIRED = (
    "borrower",
    "instrument",
    "currency",
    "price",
    "discount_margin",
    "as_of",
)
IDENTITY_FIELDS = ("figi", "instrument_key")
OPTIONAL_FIELDS = (
    "bid",
    "ask",
    "benchmark",
    "floor",
    "spread",
    "maturity",
    "seniority",
    "rating",
    "sector",
    "sub_sector",
)
ALL_FIELDS = set((*IDENTITY_FIELDS, *CANONICAL_REQUIRED, *OPTIONAL_FIELDS))
CONSTANT_FIELDS = {"currency", "as_of"}
NUMERIC_FIELDS = {"price", "discount_margin", "bid", "ask", "floor", "spread"}


def _header_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


ALIASES: dict[str, set[str]] = {
    "figi": {_header_key(v) for v in ("FIGI", "Bloomberg ID", "Bloomberg FIGI")},
    "instrument_key": {_header_key(v) for v in ("Instrument Key", "Instrument ID", "Security ID")},
    "borrower": {_header_key(v) for v in ("Borrower", "Borrower Name", "Company")},
    "instrument": {_header_key(v) for v in ("Instrument", "Loan Name", "Security", "Facility")},
    "currency": {_header_key(v) for v in ("Currency", "CCY")},
    "price": {_header_key(v) for v in ("Price", "Mid Price", "Market Price")},
    "discount_margin": {_header_key(v) for v in ("Discount Margin", "DM", "Mid 3Y DM")},
    "as_of": {_header_key(v) for v in ("As Of", "As Of Date", "Market As Of")},
    "bid": {_header_key("Bid")},
    "ask": {_header_key("Ask")},
    "benchmark": {_header_key(v) for v in ("Benchmark", "Index")},
    "floor": {_header_key(v) for v in ("Floor", "Base Rate Floor")},
    "spread": {_header_key(v) for v in ("Spread", "Margin")},
    "maturity": {_header_key("Maturity")},
    "seniority": {_header_key(v) for v in ("Seniority", "Ranking")},
    "rating": {_header_key(v) for v in ("Rating", "Ratings")},
    "sector": {_header_key(v) for v in ("Sector", "Index Sector")},
    "sub_sector": {_header_key(v) for v in ("Sub Sector", "Sub-sector")},
}


class WorkbookMapping(BaseModel):
    """Analyst-confirmed workbook interpretation used by preview and commit."""

    model_config = ConfigDict(extra="forbid")

    sheet: Optional[str] = Field(default=None, min_length=1, max_length=31)
    header_row: Optional[int] = Field(default=None, ge=1, le=HEADER_SCAN_ROWS)
    columns: dict[str, str] = Field(default_factory=dict, max_length=len(ALL_FIELDS))
    constants: dict[str, Any] = Field(default_factory=dict, max_length=len(CONSTANT_FIELDS))

    @field_validator("columns")
    @classmethod
    def validate_columns(cls, value: dict[str, str]) -> dict[str, str]:
        unknown = sorted(set(value) - ALL_FIELDS)
        if unknown:
            raise ValueError(f"Unsupported mapped field(s): {', '.join(unknown)}")
        cleaned: dict[str, str] = {}
        for key, header in value.items():
            if not isinstance(header, str) or not header.strip() or len(header.strip()) > 255:
                raise ValueError(f"Column mapping for {key} must be a non-empty header")
            cleaned[key] = header.strip()
        return cleaned

    @field_validator("constants")
    @classmethod
    def validate_constants(cls, value: dict[str, Any]) -> dict[str, Any]:
        unknown = sorted(set(value) - CONSTANT_FIELDS)
        if unknown:
            raise ValueError(f"Only currency and as_of may be constants; got {', '.join(unknown)}")
        return value


class ImportIssue(BaseModel):
    severity: Literal["blocking", "warning"]
    code: str
    message: str
    row: Optional[int] = None
    column: Optional[str] = None
    field: Optional[str] = None


class WorkbookPreview(BaseModel):
    workbook_sha256: str
    selected_sheet: Optional[str]
    header_row: Optional[int]
    mapping: dict[str, Any]
    as_of: Optional[datetime]
    row_count: int
    accepted_count: int
    rejected_count: int
    formula_cell_count: int
    blocking_count: int
    warning_count: int
    preview_truncated: bool
    rows: list[dict[str, Any]]
    issues: list[ImportIssue]


class MarketWorkbookError(ValueError):
    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code
        self.message = message


def parse_mapping(raw: str | dict[str, Any] | WorkbookMapping | None) -> WorkbookMapping:
    if isinstance(raw, WorkbookMapping):
        return raw
    try:
        payload = json.loads(raw) if isinstance(raw, str) else (raw or {})
    except json.JSONDecodeError as exc:
        raise MarketWorkbookError("invalid_mapping_json", "Mapping must be valid JSON.") from exc
    try:
        return WorkbookMapping.model_validate(payload)
    except ValidationError as exc:
        raise MarketWorkbookError("invalid_mapping", str(exc)) from exc


def require_xlsx_filename(filename: str) -> None:
    """The v2 boundary is OOXML `.xlsx`, never legacy OLE or macro-enabled files."""
    if not filename or not filename.lower().endswith(".xlsx"):
        raise MarketWorkbookError("xlsx_required", "Market price imports accept .xlsx files only.")


def _validate_package(content: bytes) -> None:
    if not content.startswith(b"PK\x03\x04"):
        raise MarketWorkbookError("invalid_ooxml", "Uploaded file is not an OOXML .xlsx workbook.")
    try:
        package = zipfile.ZipFile(io.BytesIO(content))
    except (zipfile.BadZipFile, OSError) as exc:
        raise MarketWorkbookError("invalid_ooxml", "Uploaded file is not a valid OOXML package.") from exc
    with package:
        members = package.infolist()
        if len(members) > MAX_PACKAGE_MEMBERS:
            raise MarketWorkbookError("package_member_limit", "Workbook package contains too many members.")
        names = [member.filename for member in members]
        if len(set(names)) != len(names):
            raise MarketWorkbookError("duplicate_package_member", "Workbook package contains duplicate members.")
        if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
            raise MarketWorkbookError("invalid_ooxml", "Workbook package is missing required OOXML parts.")
        total_uncompressed = 0
        for member in members:
            path = PurePosixPath(member.filename)
            if path.is_absolute() or ".." in path.parts or "\\" in member.filename:
                raise MarketWorkbookError("unsafe_package_path", "Workbook package contains an unsafe member path.")
            if member.flag_bits & 0x1:
                raise MarketWorkbookError("encrypted_package", "Encrypted workbook packages are not supported.")
            if member.file_size > MAX_PACKAGE_MEMBER_BYTES:
                raise MarketWorkbookError("package_member_size_limit", "Workbook package member is too large.")
            total_uncompressed += member.file_size
            if total_uncompressed > MAX_PACKAGE_UNCOMPRESSED_BYTES:
                raise MarketWorkbookError("package_size_limit", "Workbook expands beyond the safe processing limit.")
            if member.file_size >= 1_000_000:
                ratio = member.file_size / max(1, member.compress_size)
                if ratio > MAX_COMPRESSION_RATIO:
                    raise MarketWorkbookError("compression_ratio_limit", "Workbook compression ratio exceeds the safe limit.")
            lowered = member.filename.lower()
            if (
                "vbaproject" in lowered
                or lowered.startswith("xl/externallinks/")
                or lowered.startswith("xl/embeddings/")
                or lowered.startswith("xl/querytables/")
                or lowered == "xl/connections.xml"
            ):
                raise MarketWorkbookError("active_or_external_content", "Macros, embedded objects, queries, and external links are not supported.")
        content_types = package.read("[Content_Types].xml").lower()
        if b"macroenabled" in content_types or b"vbaproject" in content_types:
            raise MarketWorkbookError("active_or_external_content", "Macro-enabled workbooks are not supported.")
        for member in members:
            if not member.filename.lower().endswith(".rels"):
                continue
            try:
                relationship_xml = package.read(member)
                lowered_xml = relationship_xml.lower()
                if b"<!doctype" in lowered_xml or b"<!entity" in lowered_xml:
                    raise MarketWorkbookError("invalid_relationships", "Workbook relationships contain a forbidden document type.")
                root = ElementTree.fromstring(relationship_xml)
            except MarketWorkbookError:
                raise
            except (ElementTree.ParseError, RuntimeError, KeyError) as exc:
                raise MarketWorkbookError("invalid_relationships", "Workbook relationships are malformed.") from exc
            for relationship in root.iter():
                if relationship.attrib.get("TargetMode", "").lower() == "external":
                    raise MarketWorkbookError("external_relationship", "External workbook relationships are not supported.")


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _finite_number(value: object) -> Optional[float]:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    number = float(value)
    return number if math.isfinite(number) else None


def _as_text(value: object, *, limit: int = MAX_CELL_TEXT) -> Optional[str]:
    if _is_blank(value):
        return None
    text = str(value).strip()
    return text if len(text) <= limit else None


def _as_utc(value: object) -> Optional[datetime]:
    if isinstance(value, datetime):
        result = value
    elif isinstance(value, date):
        result = datetime.combine(value, time.min)
    elif isinstance(value, str):
        candidate = value.strip()
        if not candidate:
            return None
        try:
            result = datetime.fromisoformat(candidate.replace("Z", "+00:00"))
        except ValueError:
            try:
                result = datetime.combine(date.fromisoformat(candidate), time.min)
            except ValueError:
                return None
    else:
        return None
    return result.replace(tzinfo=timezone.utc) if result.tzinfo is None else result.astimezone(timezone.utc)


def _formula_external(formula: str) -> bool:
    lowered = formula.lower()
    return (
        "[" in formula
        or "http://" in lowered
        or "https://" in lowered
        or "file://" in lowered
        or "webservice(" in lowered
        or "dde(" in lowered
    )


@dataclass
class _IssueLedger:
    issues: list[ImportIssue] = field(default_factory=list)
    overflowed: bool = False

    def add(
        self,
        severity: Literal["blocking", "warning"],
        code: str,
        message: str,
        *,
        row: int | None = None,
        column: str | None = None,
        field_name: str | None = None,
    ) -> None:
        if len(self.issues) < MAX_ISSUES:
            self.issues.append(ImportIssue(
                severity=severity,
                code=code,
                message=message,
                row=row,
                column=column,
                field=field_name,
            ))
        elif not self.overflowed:
            self.overflowed = True
            self.issues[-1] = ImportIssue(
                severity="blocking",
                code="issue_limit",
                message="Workbook produced too many validation issues; correct the source before retrying.",
            )


@dataclass
class _HeaderCandidate:
    sheet: str
    row_number: int
    headers: list[object]
    resolved: dict[str, int]


def _resolve_headers(headers: list[object], mapping: WorkbookMapping) -> tuple[dict[str, int], list[str]]:
    by_key: dict[str, list[int]] = {}
    for index, header in enumerate(headers):
        key = _header_key(header)
        if key:
            by_key.setdefault(key, []).append(index)
    resolved: dict[str, int] = {}
    ambiguous: list[str] = []
    for field_name in ALL_FIELDS:
        explicit = mapping.columns.get(field_name)
        keys = {_header_key(explicit)} if explicit else ALIASES[field_name]
        matches = sorted({index for key in keys for index in by_key.get(key, [])})
        if len(matches) == 1:
            resolved[field_name] = matches[0]
        elif len(matches) > 1:
            ambiguous.append(field_name)
    return resolved, ambiguous


def _candidate_complete(candidate: _HeaderCandidate, mapping: WorkbookMapping) -> bool:
    fields = set(candidate.resolved) | set(mapping.constants)
    return bool(fields.intersection(IDENTITY_FIELDS)) and all(field in fields for field in CANONICAL_REQUIRED)


def _find_candidate(formula_wb: Any, mapping: WorkbookMapping, ledger: _IssueLedger) -> Optional[_HeaderCandidate]:
    if len(formula_wb.sheetnames) > MAX_SHEETS:
        ledger.add("blocking", "sheet_limit", "Workbook contains too many sheets.")
        return None
    sheet_names = [mapping.sheet] if mapping.sheet else list(formula_wb.sheetnames)
    if mapping.sheet and mapping.sheet not in formula_wb.sheetnames:
        ledger.add("blocking", "sheet_not_found", "The mapped sheet does not exist.")
        return None
    candidates: list[_HeaderCandidate] = []
    for sheet_name in sheet_names:
        worksheet = formula_wb[sheet_name]
        max_row = worksheet.max_row
        max_column = worksheet.max_column
        if (
            (max_row is not None and max_row > MAX_ROWS_PER_SHEET)
            or (max_column is not None and max_column > MAX_COLUMNS_PER_SHEET)
        ):
            ledger.add("blocking", "worksheet_dimension_limit", f"Sheet {sheet_name!r} exceeds row or column limits.")
            continue
        rows = [mapping.header_row] if mapping.header_row else range(
            1, min(max_row or HEADER_SCAN_ROWS, HEADER_SCAN_ROWS) + 1
        )
        for row_number in rows:
            headers = [cell.value for cell in next(worksheet.iter_rows(min_row=row_number, max_row=row_number))]
            if len(headers) > MAX_COLUMNS_PER_SHEET and any(
                not _is_blank(value) for value in headers[MAX_COLUMNS_PER_SHEET:]
            ):
                ledger.add("blocking", "worksheet_dimension_limit", f"Sheet {sheet_name!r} exceeds the column limit.")
                continue
            headers = headers[:MAX_COLUMNS_PER_SHEET]
            resolved, ambiguous = _resolve_headers(headers, mapping)
            if ambiguous:
                continue
            candidates.append(_HeaderCandidate(sheet_name, row_number, headers, resolved))
    complete = [candidate for candidate in candidates if _candidate_complete(candidate, mapping)]
    if len(complete) == 1:
        return complete[0]
    if len(complete) > 1:
        ledger.add("blocking", "ambiguous_sheet_or_header", "More than one sheet/header row satisfies the mapping; select both explicitly.")
        return None
    best = sorted(candidates, key=lambda candidate: len(candidate.resolved), reverse=True)
    if best:
        fields = set(best[0].resolved) | set(mapping.constants)
        missing = [field for field in CANONICAL_REQUIRED if field not in fields]
        if not fields.intersection(IDENTITY_FIELDS):
            missing.insert(0, "figi_or_instrument_key")
        ledger.add("blocking", "missing_required_columns", f"Required mappings are missing: {', '.join(missing)}.")
    else:
        ledger.add("blocking", "header_not_found", "No usable workbook header was found in the first 25 rows.")
    return None


def preview_workbook(
    content: bytes,
    *,
    filename: str,
    mapping: WorkbookMapping | dict[str, Any] | str | None = None,
    now: datetime | None = None,
    row_limit: int | None = MAX_PREVIEW_ROWS,
) -> WorkbookPreview:
    """Validate and normalize a workbook without any external side effect."""
    require_xlsx_filename(filename)
    _validate_package(content)
    parsed_mapping = parse_mapping(mapping)
    workbook_sha256 = hashlib.sha256(content).hexdigest()
    ledger = _IssueLedger()
    try:
        from openpyxl import load_workbook

        formula_wb = load_workbook(io.BytesIO(content), read_only=True, data_only=False, keep_links=False)
        cached_wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True, keep_links=False)
    except Exception as exc:
        raise MarketWorkbookError("workbook_parse_failed", "Workbook could not be opened safely.") from exc
    candidate = _find_candidate(formula_wb, parsed_mapping, ledger)
    if candidate is None:
        return WorkbookPreview(
            workbook_sha256=workbook_sha256,
            selected_sheet=None,
            header_row=None,
            mapping=parsed_mapping.model_dump(mode="json"),
            as_of=None,
            row_count=0,
            accepted_count=0,
            rejected_count=0,
            formula_cell_count=0,
            blocking_count=sum(issue.severity == "blocking" for issue in ledger.issues),
            warning_count=sum(issue.severity == "warning" for issue in ledger.issues),
            preview_truncated=False,
            rows=[],
            issues=ledger.issues,
        )

    formula_ws = formula_wb[candidate.sheet]
    cached_ws = cached_wb[candidate.sheet]
    resolved = candidate.resolved
    now_utc = (now or datetime.now(timezone.utc)).astimezone(timezone.utc)
    normalized_rows: list[dict[str, Any]] = []
    rejected_rows = 0
    observed_as_of: Optional[datetime] = None
    seen_keys: set[str] = set()
    nonempty_cells = 0
    row_count = 0
    formula_cell_count = 0

    formula_rows = formula_ws.iter_rows(min_row=candidate.row_number + 1)
    cached_rows = cached_ws.iter_rows(min_row=candidate.row_number + 1)
    for row_number, (formula_row, cached_row) in enumerate(
        zip(formula_rows, cached_rows), start=candidate.row_number + 1
    ):
        if ledger.overflowed:
            break
        if row_number > MAX_ROWS_PER_SHEET:
            if any(not _is_blank(cell.value) for cell in formula_row):
                ledger.add("blocking", "worksheet_dimension_limit", "Worksheet exceeds the row limit.")
            break
        if len(formula_row) > MAX_COLUMNS_PER_SHEET and any(
            not _is_blank(cell.value) for cell in formula_row[MAX_COLUMNS_PER_SHEET:]
        ):
            ledger.add("blocking", "worksheet_dimension_limit", "Worksheet exceeds the column limit.", row=row_number)
            break
        nonempty_cells += sum(not _is_blank(cell.value) for cell in formula_row)
        if nonempty_cells > MAX_NONEMPTY_CELLS:
            ledger.add("blocking", "cell_limit", "Workbook contains too many non-empty cells.")
            break
        if not any(not _is_blank(formula_row[index].value) for index in resolved.values()):
            continue
        row_count += 1
        row_blocking_before = sum(issue.severity == "blocking" for issue in ledger.issues)
        row_rejected = False
        raw: dict[str, object] = dict(parsed_mapping.constants)
        formula_fields: list[str] = []
        for field_name, column_index in resolved.items():
            formula_cell = formula_row[column_index]
            cached_cell = cached_row[column_index]
            value = formula_cell.value
            if formula_cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
                formula_text = str(value)
                formula_fields.append(field_name)
                formula_cell_count += 1
                if len(formula_text) > MAX_FORMULA_TEXT:
                    ledger.add("blocking", "formula_length", "Formula exceeds the safe length limit.", row=row_number, column=formula_cell.coordinate, field_name=field_name)
                if _formula_external(formula_text):
                    ledger.add("blocking", "external_formula", "External-reference formulas are not supported.", row=row_number, column=formula_cell.coordinate, field_name=field_name)
                value = cached_cell.value
                if field_name in (*IDENTITY_FIELDS, *CANONICAL_REQUIRED):
                    cached_number = _finite_number(value)
                    if cached_number is None:
                        ledger.add("blocking", "required_formula_cache", "Required formula cell has no finite cached value.", row=row_number, column=formula_cell.coordinate, field_name=field_name)
            raw[field_name] = value

        figi = _as_text(raw.get("figi"), limit=32)
        explicit_key = _as_text(raw.get("instrument_key"), limit=160)
        if not figi and not explicit_key:
            row_rejected = True
            ledger.add("warning", "missing_instrument_identity", "Row rejected: FIGI or an explicit instrument key is required.", row=row_number)
        instrument_key = explicit_key or figi
        canonical_key = instrument_key.upper() if instrument_key else None
        if canonical_key and canonical_key in seen_keys:
            row_rejected = True
            ledger.add("blocking", "duplicate_instrument_key", "Instrument key is duplicated in this snapshot.", row=row_number, field_name="instrument_key")

        borrower = _as_text(raw.get("borrower"), limit=255)
        instrument = _as_text(raw.get("instrument"), limit=255)
        currency = _as_text(raw.get("currency"), limit=3)
        if not borrower:
            row_rejected = True
            ledger.add("warning", "invalid_borrower", "Row rejected: borrower is required and must fit the supported length.", row=row_number, field_name="borrower")
        if not instrument:
            row_rejected = True
            ledger.add("warning", "invalid_instrument", "Row rejected: instrument is required and must fit the supported length.", row=row_number, field_name="instrument")
        if not currency or not re.fullmatch(r"[A-Za-z]{3}", currency):
            row_rejected = True
            ledger.add("warning", "invalid_currency", "Row rejected: currency must be a three-letter code.", row=row_number, field_name="currency")
        else:
            currency = currency.upper()

        price = _finite_number(raw.get("price"))
        discount_margin = _finite_number(raw.get("discount_margin"))
        if price is None:
            row_rejected = True
            ledger.add("warning", "invalid_price", "Row rejected: price must be finite.", row=row_number, field_name="price")
        if discount_margin is None:
            row_rejected = True
            ledger.add("warning", "invalid_discount_margin", "Row rejected: discount margin must be finite.", row=row_number, field_name="discount_margin")

        row_as_of = _as_utc(raw.get("as_of"))
        if row_as_of is None:
            row_rejected = True
            ledger.add("warning", "invalid_as_of", "Row rejected: a real market as-of is required; upload time is never substituted.", row=row_number, field_name="as_of")
        elif row_as_of > now_utc:
            row_rejected = True
            ledger.add("blocking", "future_as_of", "Market as-of cannot be in the future.", row=row_number, field_name="as_of")
        elif not row_rejected and observed_as_of is None:
            observed_as_of = row_as_of
        elif not row_rejected and row_as_of != observed_as_of:
            row_rejected = True
            ledger.add("blocking", "inconsistent_as_of", "All rows in one snapshot must have the same market as-of.", row=row_number, field_name="as_of")

        optional: dict[str, Any] = {}
        for field_name in OPTIONAL_FIELDS:
            value = raw.get(field_name)
            if _is_blank(value):
                continue
            if field_name in NUMERIC_FIELDS:
                parsed = _finite_number(value)
                if parsed is None:
                    ledger.add("warning", "invalid_optional_number", f"Optional {field_name} value was ignored because it is not finite.", row=row_number, field_name=field_name)
                    continue
                optional[field_name] = parsed
            else:
                parsed = _as_text(value)
                if parsed is None:
                    ledger.add("warning", "invalid_optional_text", f"Optional {field_name} value was ignored because it is too long.", row=row_number, field_name=field_name)
                    continue
                optional[field_name] = parsed

        row_blocking_after = sum(issue.severity == "blocking" for issue in ledger.issues)
        if row_rejected or row_blocking_after > row_blocking_before:
            rejected_rows += 1
            continue
        if canonical_key:
            seen_keys.add(canonical_key)
        normalized_rows.append({
            "row": row_number,
            "instrument_key": instrument_key,
            "figi": figi.upper() if figi else None,
            "borrower": borrower,
            "instrument": instrument,
            "currency": currency,
            "price": price,
            "discount_margin": discount_margin,
            "as_of": row_as_of.isoformat() if row_as_of else None,
            "formula_fields": formula_fields,
            **optional,
        })

    if formula_cell_count:
        ledger.add(
            "warning",
            "cached_formula_value",
            f"{formula_cell_count} mapped formula cell(s) were not calculated; finite cached values were used where required.",
        )
    if not normalized_rows:
        ledger.add("blocking", "no_acceptable_rows", "Workbook contains no acceptable market rows.")
    blocking_count = sum(issue.severity == "blocking" for issue in ledger.issues)
    warning_count = sum(issue.severity == "warning" for issue in ledger.issues)
    preview_rows = normalized_rows if row_limit is None else normalized_rows[:max(0, row_limit)]
    return WorkbookPreview(
        workbook_sha256=workbook_sha256,
        selected_sheet=candidate.sheet,
        header_row=candidate.row_number,
        mapping={
            **parsed_mapping.model_dump(mode="json"),
            "columns": {
                field_name: str(candidate.headers[index])
                for field_name, index in sorted(resolved.items())
            },
        },
        as_of=observed_as_of,
        row_count=row_count,
        accepted_count=len(normalized_rows),
        rejected_count=rejected_rows,
        formula_cell_count=formula_cell_count,
        blocking_count=blocking_count,
        warning_count=warning_count,
        preview_truncated=row_limit is not None and len(normalized_rows) > max(0, row_limit),
        rows=preview_rows,
        issues=ledger.issues,
    )
