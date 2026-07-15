"""Pure, bounded CAOS Model Engine v2 workbook import/export.

The module intentionally owns no database, vault, authorization, or clock
state.  Routes may persist a workbook only after calling ``preview_workbook``
again over the committed bytes.  OOXML is screened with the market-import
package gate before openpyxl sees attacker-controlled XML, and formula and
cached-value views are always opened separately.  Formulas are evidence only;
CAOS never evaluates them.
"""

from __future__ import annotations

import hashlib
import io
import json
import math
import re
import zipfile
from calendar import monthrange
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any, Literal, Optional

from pydantic import BaseModel, ConfigDict, Field, ValidationError, field_validator, model_validator

from market_xlsx import (
    MAX_FORMULA_TEXT,
    MarketWorkbookError,
    _formula_external,
    _validate_package,
)
from model_engine_v2 import (
    ENGINE_VERSION,
    CellOverride,
    DebtInstrument,
    DebtPeriod,
    ModelAuthority,
    ModelCalculation,
    ModelDraftPayload,
    ModelPeriodInput,
    ModelUiPreferences,
    SUPPORTED_REPORTING_CURRENCIES,
    SUPPORTED_REPORTING_UNITS,
    calculate_model,
    is_finite_number,
)


SCHEMA_ID = "CAOS_MODEL_WORKBOOK_V1"
SHEET_NAMES = (
    "Cover",
    "Model",
    "Assumptions",
    "Debt Schedule",
    "Overrides",
    "Sources - Audit",
)
MAX_MODEL_SHEETS = 12
MAX_MODEL_FILE_BYTES = 32 * 1024 * 1024
MAX_ROWS_PER_SHEET = 10_000
MAX_COLUMNS_PER_SHEET = 96
MAX_DIMENSION_CELLS = 1_000_000
MAX_NONEMPTY_CELLS = 250_000
MAX_TEXT = 4_096
MAX_ISSUES = 500
MAX_FORMULAS = 10_000
MAX_FORMULA_REFERENCES = 50_000
MAX_LEGACY_HEADER_ROW = 25
_HASH = re.compile(r"^[0-9a-f]{64}$")
_KEY = re.compile(r"^[A-Za-z0-9_.:-]{1,300}$")
_CELL_REF = re.compile(
    r"(?:(?:'([^']+)'|([A-Za-z0-9 _.-]+))!)?(\$?[A-Z]{1,3}\$?\d+)",
    re.IGNORECASE,
)
_VOLATILE_FORMULA = re.compile(
    r"\b(?:NOW|TODAY|RAND|RANDBETWEEN|OFFSET|INDIRECT|CELL|INFO)\s*\(",
    re.IGNORECASE,
)
_ACTIVE_FORMULA = re.compile(
    r"\b(?:RTD|DDE|CALL|REGISTER\.ID|EXEC|SQL\.REQUEST)\s*\(",
    re.IGNORECASE,
)
_ACTIVE_PACKAGE_PREFIXES = (
    "xl/activex/",
    "xl/ctrlprops/",
    "customui/",
)

PERIOD_FIELDS = (
    "revenue",
    "reported_ebitda",
    "adjustments",
    "adjusted_ebitda",
    "cash",
    "total_debt",
    "net_debt",
    "cash_interest",
    "taxes",
    "capex",
    "working_capital_change",
    "other_cash_flow",
)
DEBT_PERIOD_FIELDS = (
    "opening_balance",
    "closing_balance",
    "draws",
    "repayments",
    "scheduled_amortization",
    "commitment",
    "benchmark_rate",
    "floor_rate",
    "spread_rate",
    "coupon_rate",
    "commitment_fee_rate",
    "pik_rate",
    "cash_fees",
    "hedge_effect",
    "fx_rate",
)
MODEL_FIELDS = (
    "revenue",
    "reported_ebitda",
    "adjustments",
    "adjusted_ebitda",
    "cash_interest",
    "total_debt",
    "cash",
    "net_debt",
    "gross_leverage",
    "net_leverage",
    "interest_coverage",
    "free_cash_flow",
)

COVER_HEADERS = ("key", "value")
MODEL_HEADERS = (
    "row_key",
    "period_key",
    "period_label",
    "metric",
    "value",
    "original_value",
    "formula",
    "overridden",
    "units",
)
ASSUMPTION_HEADERS = (
    "row_key",
    "period_key",
    "label",
    "kind",
    "months",
    *PERIOD_FIELDS,
    "authority_origin",
    "authority_method",
    "authority_source_ids",
    "authority_as_of",
    "units",
)
DEBT_HEADERS = (
    "row_key",
    "instrument_id",
    "name",
    "priority",
    "seniority",
    "currency",
    "rate_type",
    "maturity",
    "amortization",
    "benchmark_curve",
    "period_key",
    *DEBT_PERIOD_FIELDS,
    "source_ids",
    "authority_origin",
    "authority_method",
    "authority_source_ids",
    "authority_as_of",
    "units",
)
OVERRIDE_HEADERS = (
    "row_key",
    "node_id",
    "value_type",
    "value",
    "reason",
    "scope",
    "source",
    "expires_at",
    "original_value",
    "original_formula",
)
SOURCE_HEADERS = ("record_type", "key", "value")

ASSUMPTION_MAPPING_FIELDS = set(ASSUMPTION_HEADERS) - {"row_key", "units"}
DEBT_MAPPING_FIELDS = set(DEBT_HEADERS) - {"row_key", "units"}
OVERRIDE_MAPPING_FIELDS = set(OVERRIDE_HEADERS) - {
    "row_key",
    "original_value",
    "original_formula",
}


class StrictModel(BaseModel):
    model_config = ConfigDict(extra="forbid")


class WorkbookIssue(StrictModel):
    severity: Literal["blocking", "warning"]
    code: str = Field(min_length=1, max_length=96)
    message: str = Field(min_length=1, max_length=1_000)
    sheet: Optional[str] = Field(default=None, max_length=31)
    cell: Optional[str] = Field(default=None, max_length=32)
    field: Optional[str] = Field(default=None, max_length=96)


class MappingAmbiguity(StrictModel):
    table: Literal["assumptions", "debt_schedule", "overrides"]
    field: str = Field(min_length=1, max_length=96)
    selector: Literal["column", "row"] = "column"
    candidates: list[str] = Field(min_length=2, max_length=96)
    message: str = Field(min_length=1, max_length=500)


class FormulaAuditEntry(StrictModel):
    sheet: str = Field(min_length=1, max_length=31)
    cell: str = Field(min_length=1, max_length=32)
    formula: str = Field(min_length=1, max_length=MAX_FORMULA_TEXT)
    cached_value: Optional[float] = None
    disposition: Literal["input_candidate", "comparison_only", "unused"]
    blocking_codes: list[str] = Field(default_factory=list, max_length=10)


class EmbeddedHashes(StrictModel):
    engine_version: Optional[str] = None
    source_fingerprint: Optional[str] = None
    input_fingerprint: Optional[str] = None
    calculation_hash: Optional[str] = None


class WorkbookIdentity(StrictModel):
    issuer_id: str = Field(min_length=1, max_length=64)
    draft_revision: int = Field(ge=1)
    exported_by: str = Field(min_length=1, max_length=255)
    exported_at: datetime


class LegacyTableMapping(StrictModel):
    layout: Literal["records"] = "records"
    sheet: str = Field(min_length=1, max_length=31)
    header_row: int = Field(default=1, ge=1, le=MAX_LEGACY_HEADER_ROW)
    columns: dict[str, str] = Field(min_length=1, max_length=64)
    # Optional one-based physical column selections resolve duplicate normalized
    # headers only after an analyst has reviewed the preview ambiguity.
    column_indices: dict[str, int] = Field(default_factory=dict, max_length=64)

    @field_validator("columns")
    @classmethod
    def bounded_headers(cls, value: dict[str, str]) -> dict[str, str]:
        cleaned: dict[str, str] = {}
        for field_name, header in value.items():
            if len(field_name) > 96 or not isinstance(header, str):
                raise ValueError("mapped fields and headers must be bounded strings")
            header = header.strip()
            if not header or len(header) > 255:
                raise ValueError(f"mapped header for {field_name} must be non-empty and bounded")
            cleaned[field_name] = header
        return cleaned

    @model_validator(mode="after")
    def validate_column_indices(self) -> "LegacyTableMapping":
        unknown = sorted(set(self.column_indices) - set(self.columns))
        if unknown:
            raise ValueError(
                "column index selectors require a mapped field: "
                + ", ".join(unknown)
            )
        invalid = sorted(
            field_name
            for field_name, index in self.column_indices.items()
            if isinstance(index, bool) or index < 1 or index > MAX_COLUMNS_PER_SHEET
        )
        if invalid:
            raise ValueError(
                "column index selectors must be one-based and bounded: "
                + ", ".join(invalid)
            )
        return self


def _expanded_year(value: str) -> int:
    year = int(value)
    return 2000 + year if len(value) == 2 else year


def _normalize_period_key(value: object) -> Optional[str]:
    """Normalize a bounded set of reviewed spreadsheet period aliases.

    The importer deliberately does not guess at free-form dates. Two-digit
    years are the CAOS workbook window 2000-2099, while LTM aliases resolve to
    a stable calendar month-end key rather than upload time.
    """

    text = _read_text(value, limit=64)
    if text is None:
        return None
    normalized = re.sub(r"[\u2010-\u2015]", "-", text.strip()).upper()
    normalized = re.sub(r"\s+", " ", normalized)

    canonical = re.fullmatch(
        r"(?:FY\d{4}|Q[1-4]-\d{4}|YTD-Q[1-4]-\d{4}|LTM-\d{4}-\d{2}-\d{2}|\d{4}-\d{2}-\d{2})",
        normalized,
    )
    if canonical:
        return normalized

    match = re.fullmatch(r"FY\s*['-]?(\d{2}|\d{4})\s*(?:A|E|F)?", normalized)
    if match:
        return f"FY{_expanded_year(match.group(1)):04d}"
    match = re.fullmatch(r"Q([1-4])\s*[- /]?\s*['-]?(\d{2}|\d{4})\s*(?:A|E|F)?", normalized)
    if match:
        return f"Q{match.group(1)}-{_expanded_year(match.group(2)):04d}"
    match = re.fullmatch(
        r"YTD\s*[- ]?\s*Q([1-4])\s*[- /]?\s*['-]?(\d{2}|\d{4})\s*(?:A|E|F)?",
        normalized,
    )
    if match:
        return f"YTD-Q{match.group(1)}-{_expanded_year(match.group(2)):04d}"

    months = {
        "JAN": 1, "JANUARY": 1, "FEB": 2, "FEBRUARY": 2,
        "MAR": 3, "MARCH": 3, "APR": 4, "APRIL": 4,
        "MAY": 5, "JUN": 6, "JUNE": 6, "JUL": 7, "JULY": 7,
        "AUG": 8, "AUGUST": 8, "SEP": 9, "SEPT": 9, "SEPTEMBER": 9,
        "OCT": 10, "OCTOBER": 10, "NOV": 11, "NOVEMBER": 11,
        "DEC": 12, "DECEMBER": 12,
    }
    match = re.fullmatch(
        r"LTM\s*[- ]?\s*([A-Z]+)\s*[- /]?\s*['-]?(\d{2}|\d{4})",
        normalized,
    )
    if match and match.group(1) in months:
        year = _expanded_year(match.group(2))
        month = months[match.group(1)]
        return f"LTM-{year:04d}-{month:02d}-{monthrange(year, month)[1]:02d}"
    return None


def _default_period_months(period_key: str) -> Optional[int]:
    if period_key.startswith("Q"):
        return 3
    if period_key.startswith("YTD-Q"):
        return int(period_key[5]) * 3
    if period_key.startswith(("FY", "LTM-")):
        return 12
    # A bare ISO date has no unambiguous start date. Analysts must bind months.
    return None


class LegacyMatrixMapping(StrictModel):
    """Explicit account-row/period-column mapping for close-format models."""

    layout: Literal["account_period_matrix"]
    sheet: str = Field(min_length=1, max_length=31)
    header_row: int = Field(default=1, ge=1, le=MAX_LEGACY_HEADER_ROW)
    account_column: str = Field(min_length=1, max_length=255)
    account_column_index: Optional[int] = Field(
        default=None, ge=1, le=MAX_COLUMNS_PER_SHEET
    )
    account_rows: dict[str, str] = Field(min_length=1, max_length=len(PERIOD_FIELDS))
    account_row_indices: dict[str, int] = Field(default_factory=dict, max_length=len(PERIOD_FIELDS))
    period_columns: dict[str, str] = Field(min_length=1, max_length=120)
    period_column_indices: dict[str, int] = Field(default_factory=dict, max_length=120)
    period_labels: dict[str, str] = Field(default_factory=dict, max_length=120)
    period_kinds: dict[str, Literal["actual", "forecast", "ltm", "pro_forma"]] = Field(
        min_length=1, max_length=120
    )
    period_months: dict[str, int] = Field(default_factory=dict, max_length=120)

    @field_validator("account_rows")
    @classmethod
    def validate_account_rows(cls, value: dict[str, str]) -> dict[str, str]:
        unknown = sorted(set(value) - set(PERIOD_FIELDS))
        if unknown:
            raise ValueError(f"unsupported matrix account field(s): {', '.join(unknown)}")
        cleaned: dict[str, str] = {}
        for field_name, label in value.items():
            label = label.strip()
            if not label or len(label) > 255:
                raise ValueError(f"matrix account label for {field_name} must be bounded")
            cleaned[field_name] = label
        return cleaned

    @field_validator(
        "period_columns",
        "period_column_indices",
        "period_labels",
        "period_kinds",
        "period_months",
    )
    @classmethod
    def normalize_period_map(cls, value: dict[str, Any]) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for raw_key, mapped in value.items():
            period_key = _normalize_period_key(raw_key)
            if period_key is None:
                raise ValueError(f"unsupported matrix period alias: {raw_key}")
            if period_key in result:
                raise ValueError(f"matrix period aliases collide at {period_key}")
            result[period_key] = mapped
        return result

    @model_validator(mode="after")
    def validate_matrix(self) -> "LegacyMatrixMapping":
        for field_name, label in self.period_columns.items():
            if not isinstance(label, str) or not label.strip() or len(label.strip()) > 255:
                raise ValueError(f"matrix period header for {field_name} must be bounded")
        period_keys = set(self.period_columns)
        if set(self.period_kinds) != period_keys:
            missing = sorted(period_keys - set(self.period_kinds))
            extra = sorted(set(self.period_kinds) - period_keys)
            raise ValueError(
                "matrix period_kinds must exactly bind period_columns"
                + (f"; missing: {', '.join(missing)}" if missing else "")
                + (f"; extra: {', '.join(extra)}" if extra else "")
            )
        for name, keys in (
            ("period_labels", set(self.period_labels)),
            ("period_months", set(self.period_months)),
            ("period_column_indices", set(self.period_column_indices)),
        ):
            unknown = sorted(keys - period_keys)
            if unknown:
                raise ValueError(f"{name} references unmapped period(s): {', '.join(unknown)}")
        unknown_rows = sorted(set(self.account_row_indices) - set(self.account_rows))
        if unknown_rows:
            raise ValueError(
                "account_row_indices references unmapped account(s): "
                + ", ".join(unknown_rows)
            )
        for field_name, row in self.account_row_indices.items():
            if isinstance(row, bool) or row <= self.header_row or row > MAX_ROWS_PER_SHEET:
                raise ValueError(f"matrix row selector for {field_name} is out of range")
        for period_key, column in self.period_column_indices.items():
            if isinstance(column, bool) or column < 1 or column > MAX_COLUMNS_PER_SHEET:
                raise ValueError(f"matrix column selector for {period_key} is out of range")
        for period_key, months in self.period_months.items():
            if isinstance(months, bool) or months < 1 or months > 24:
                raise ValueError(f"matrix months for {period_key} must be between 1 and 24")
        missing_months = sorted(
            period_key
            for period_key in period_keys
            if _default_period_months(period_key) is None
            and period_key not in self.period_months
        )
        if missing_months:
            raise ValueError(
                "matrix ISO-date periods require explicit period_months: "
                + ", ".join(missing_months)
            )
        return self


class LegacyWorkbookMapping(StrictModel):
    mode: Literal["mapped_legacy"] = "mapped_legacy"
    assumptions: LegacyTableMapping | LegacyMatrixMapping
    debt_schedule: Optional[LegacyTableMapping] = None
    overrides: Optional[LegacyTableMapping] = None
    reporting_currency: str = Field(min_length=3, max_length=3)
    reporting_unit: str = Field(min_length=1, max_length=32)
    source_ids: list[str] = Field(default_factory=list, max_length=500)
    authority_as_of: Optional[datetime] = None

    @field_validator("reporting_currency")
    @classmethod
    def normalize_currency(cls, value: str) -> str:
        value = value.strip().upper()
        if value not in SUPPORTED_REPORTING_CURRENCIES:
            raise ValueError("reporting_currency is not supported")
        return value

    @field_validator("reporting_unit")
    @classmethod
    def normalize_reporting_unit(cls, value: str) -> str:
        value = value.strip().lower()
        if value not in SUPPORTED_REPORTING_UNITS:
            raise ValueError("reporting_unit is not supported")
        return value

    @model_validator(mode="after")
    def validate_fields(self) -> "LegacyWorkbookMapping":
        if isinstance(self.assumptions, LegacyMatrixMapping):
            if self.debt_schedule is not None or self.overrides is not None:
                raise ValueError(
                    "matrix assumptions mapping cannot be combined with debt or override tables"
                )
            return self
        tables = (
            ("assumptions", self.assumptions, ASSUMPTION_MAPPING_FIELDS),
            ("debt_schedule", self.debt_schedule, DEBT_MAPPING_FIELDS),
            ("overrides", self.overrides, OVERRIDE_MAPPING_FIELDS),
        )
        for name, table, allowed in tables:
            if table is None:
                continue
            unknown = sorted(set(table.columns) - allowed)
            if unknown:
                raise ValueError(f"unsupported {name} mapped field(s): {', '.join(unknown)}")
        required_assumptions = {"period_key", "label", "kind"}
        missing = sorted(required_assumptions - set(self.assumptions.columns))
        if missing:
            raise ValueError(f"assumptions mapping is missing: {', '.join(missing)}")
        if not set(PERIOD_FIELDS).intersection(self.assumptions.columns):
            raise ValueError("assumptions mapping must include at least one model value")
        if self.debt_schedule is not None:
            required_debt = {
                "instrument_id", "name", "priority", "seniority", "currency", "period_key"
            }
            missing = sorted(required_debt - set(self.debt_schedule.columns))
            if missing:
                raise ValueError(f"debt schedule mapping is missing: {', '.join(missing)}")
        if self.overrides is not None:
            required_overrides = {"node_id", "value_type", "value"}
            missing = sorted(required_overrides - set(self.overrides.columns))
            if missing:
                raise ValueError(f"overrides mapping is missing: {', '.join(missing)}")
        return self


class ModelWorkbookPreview(StrictModel):
    mode: Literal["strict_v1", "mapped_legacy"]
    workbook_sha256: str
    sheet_names: list[str]
    mapping: Optional[LegacyWorkbookMapping]
    draft_payload: Optional[ModelDraftPayload]
    calculation: Optional[ModelCalculation]
    identity: Optional[WorkbookIdentity] = None
    embedded_hashes: EmbeddedHashes
    formula_audit: list[FormulaAuditEntry]
    ambiguities: list[MappingAmbiguity]
    issues: list[WorkbookIssue]
    blocking_count: int
    warning_count: int


class ModelWorkbookError(ValueError):
    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code
        self.message = message


class _Ledger:
    def __init__(self) -> None:
        self.issues: list[WorkbookIssue] = []
        self.overflowed = False

    def add(
        self,
        severity: Literal["blocking", "warning"],
        code: str,
        message: str,
        *,
        sheet: str | None = None,
        cell: str | None = None,
        field: str | None = None,
    ) -> None:
        if len(self.issues) < MAX_ISSUES:
            self.issues.append(WorkbookIssue(
                severity=severity,
                code=code,
                message=message,
                sheet=sheet,
                cell=cell,
                field=field,
            ))
        elif not self.overflowed:
            self.overflowed = True
            self.issues[-1] = WorkbookIssue(
                severity="blocking",
                code="issue_limit",
                message="Workbook produced too many validation issues.",
            )


def _safe_excel_text(value: str) -> str:
    """Force attacker-controlled text to remain text when opened in Excel."""

    if value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _read_text(value: object, *, limit: int = MAX_TEXT) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if text.startswith("'") and text[1:].lstrip().startswith(("=", "+", "-", "@")):
        text = text[1:]
    if not text or len(text) > limit:
        return None
    return text


def _json_list(value: object, *, limit: int = 500) -> Optional[list[str]]:
    if value is None or value == "":
        return []
    if isinstance(value, list):
        parsed = value
    elif isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return None
    else:
        return None
    if not isinstance(parsed, list) or len(parsed) > limit:
        return None
    result: list[str] = []
    for item in parsed:
        text = _read_text(item, limit=240)
        if text is None:
            return None
        result.append(text)
    return result


def _ui_preferences(value: object) -> Optional[ModelUiPreferences]:
    if value in (None, ""):
        return ModelUiPreferences()
    if not isinstance(value, str) or len(value) > MAX_TEXT:
        return None
    try:
        parsed = json.loads(value)
        return ModelUiPreferences.model_validate(parsed)
    except (json.JSONDecodeError, ValidationError):
        return None


def _iso(value: Optional[datetime]) -> Optional[str]:
    if value is None:
        return None
    return value.astimezone(timezone.utc).isoformat()


def _datetime(value: object) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value.strip().replace("Z", "+00:00"))
        except ValueError:
            return None
    else:
        return None
    return parsed.replace(tzinfo=timezone.utc) if parsed.tzinfo is None else parsed.astimezone(timezone.utc)


def _finite(value: object) -> Optional[float]:
    if not is_finite_number(value):
        return None
    return float(value)


def _integer(value: object) -> Optional[int]:
    number = _finite(value)
    if number is None or not number.is_integer():
        return None
    return int(number)


def _header_key(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _write_row(worksheet: Any, values: list[object] | tuple[object, ...]) -> None:
    worksheet.append([
        _safe_excel_text(value) if isinstance(value, str) else value for value in values
    ])


def _formula_for_metric(period: Any, metric: str) -> tuple[Optional[float], Optional[str], bool]:
    preferred = (f"calc:{period.period_key}:{metric}", f"input:{period.period_key}:{metric}")
    nodes = {node.node_id: node for node in period.nodes}
    node = next((nodes[node_id] for node_id in preferred if node_id in nodes), None)
    if node is None:
        return None, None, False
    return node.original_value, node.formula, node.overridden


def _validate_generated_workbook_bounds(workbook: Any) -> None:
    """Ensure an exported workbook can pass the same bounded parser contract."""

    dimension_cells = 0
    for sheet in workbook.worksheets:
        max_row = sheet.max_row or 0
        max_column = sheet.max_column or 0
        if max_row > MAX_ROWS_PER_SHEET or max_column > MAX_COLUMNS_PER_SHEET:
            raise ModelWorkbookError(
                "workbook_export_limit",
                f"Generated {sheet.title} sheet exceeds model-workbook dimensions.",
            )
        dimension_cells += max_row * max_column
    if dimension_cells > MAX_DIMENSION_CELLS:
        raise ModelWorkbookError(
            "workbook_export_limit",
            "Generated workbook exceeds the aggregate model-workbook dimension limit.",
        )
    nonempty_cells = sum(
        cell.value not in (None, "")
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    if nonempty_cells > MAX_NONEMPTY_CELLS:
        raise ModelWorkbookError(
            "workbook_export_limit",
            "Generated workbook exceeds the non-empty-cell processing limit.",
        )


def render_model_workbook(
    payload: ModelDraftPayload,
    *,
    issuer_id: str,
    draft_revision: int,
    exported_by: str,
    exported_at: Optional[datetime] = None,
) -> bytes:
    """Render, reopen, and structurally verify the canonical six-sheet workbook."""

    if not issuer_id.strip() or len(issuer_id) > 64:
        raise ValueError("issuer_id must be a bounded non-empty identifier")
    if draft_revision < 1:
        raise ValueError("draft_revision must be positive")
    if not exported_by.strip() or len(exported_by) > 255:
        raise ValueError("exported_by must be a bounded non-empty identifier")
    timestamp = (exported_at or datetime.now(timezone.utc)).astimezone(timezone.utc)
    calculation = calculate_model(payload, evaluated_at=timestamp)

    from openpyxl import Workbook, load_workbook

    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    model = workbook.create_sheet("Model")
    assumptions = workbook.create_sheet("Assumptions")
    debt = workbook.create_sheet("Debt Schedule")
    overrides = workbook.create_sheet("Overrides")
    sources = workbook.create_sheet("Sources - Audit")
    metadata: list[tuple[str, object]] = [
        ("schema_id", SCHEMA_ID),
        ("schema_version", 1),
        ("engine_version", calculation.engine_version),
        ("issuer_id", issuer_id),
        ("draft_revision", draft_revision),
        ("reporting_currency", payload.reporting_currency),
        ("reporting_unit", payload.reporting_unit),
        ("ui_preferences", json.dumps(payload.ui_preferences.model_dump(mode="json"), sort_keys=True, separators=(",", ":"))),
        ("source_fingerprint", calculation.source_fingerprint),
        ("input_fingerprint", calculation.input_fingerprint),
        ("calculation_hash", calculation.calculation_hash),
        ("exported_by", exported_by),
        ("exported_at", timestamp.isoformat()),
    ]

    _write_row(cover, COVER_HEADERS)
    for key, value in metadata:
        _write_row(cover, [key, value])

    _write_row(model, MODEL_HEADERS)
    for period in calculation.periods:
        for metric in MODEL_FIELDS:
            original, formula, overridden = _formula_for_metric(period, metric)
            units = "multiple" if metric in {"gross_leverage", "net_leverage", "interest_coverage"} else payload.reporting_unit
            _write_row(model, [
                f"model:{metric}",
                period.period_key,
                period.label,
                metric,
                getattr(period, metric),
                original,
                formula,
                overridden,
                units,
            ])

    _write_row(assumptions, ASSUMPTION_HEADERS)
    for period in payload.periods:
        _write_row(assumptions, [
            f"period:{period.period_key}",
            period.period_key,
            period.label,
            period.kind,
            period.months,
            *(getattr(period, field) for field in PERIOD_FIELDS),
            period.authority.origin,
            period.authority.method,
            json.dumps(period.authority.source_ids, separators=(",", ":")),
            _iso(period.authority.as_of),
            payload.reporting_unit,
        ])

    _write_row(debt, DEBT_HEADERS)
    # Workbook row order preserves canonical payload order; display sorting is
    # a UI concern and reordering would change the engine input fingerprint.
    for instrument in payload.debt_instruments:
        for point in instrument.periods:
            _write_row(debt, [
                f"debt:{instrument.instrument_id}:{point.period_key}",
                instrument.instrument_id,
                instrument.name,
                instrument.priority,
                instrument.seniority,
                instrument.currency,
                instrument.rate_type,
                instrument.maturity,
                instrument.amortization,
                instrument.benchmark_curve,
                point.period_key,
                *(getattr(point, field) for field in DEBT_PERIOD_FIELDS),
                json.dumps(instrument.sources, separators=(",", ":")),
                instrument.authority.origin,
                instrument.authority.method,
                json.dumps(instrument.authority.source_ids, separators=(",", ":")),
                _iso(instrument.authority.as_of),
                payload.reporting_unit,
            ])

    calculation_nodes = {
        node.node_id: node
        for period in calculation.periods
        for node in period.nodes
    }
    _write_row(overrides, OVERRIDE_HEADERS)
    for override in payload.overrides:
        displaced = calculation_nodes.get(override.node_id)
        _write_row(overrides, [
            override.node_id,
            override.node_id,
            override.value_type,
            override.value,
            override.reason,
            override.scope,
            override.source,
            _iso(override.expires_at),
            displaced.original_value if displaced else None,
            displaced.formula if displaced else None,
        ])

    _write_row(sources, SOURCE_HEADERS)
    for key, value in metadata:
        _write_row(sources, ["metadata", key, value])
    # Preserve the exact top-level source list because it participates in the
    # engine input fingerprint.  The audit union below is separately sorted.
    for source_id in payload.source_ids:
        _write_row(sources, ["payload_source", "source_id", source_id])
    all_sources = set(payload.source_ids)
    for period in payload.periods:
        all_sources.update(period.authority.source_ids)
    for instrument in payload.debt_instruments:
        all_sources.update(instrument.sources)
        all_sources.update(instrument.authority.source_ids)
    for source_id in sorted(all_sources):
        _write_row(sources, ["authority_source", "source_id", source_id])

    _validate_generated_workbook_bounds(workbook)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    workbook.properties.title = "CAOS Model Workbook v1"
    workbook.properties.subject = calculation.calculation_hash
    workbook.properties.creator = "CAOS Model Engine v2"
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.calculation.calcMode = "manual"
    output = io.BytesIO()
    workbook.save(output)
    content = output.getvalue()

    reopened = load_workbook(io.BytesIO(content), read_only=True, data_only=False, keep_links=False)
    try:
        if tuple(reopened.sheetnames) != SHEET_NAMES:
            raise ValueError("generated workbook failed six-sheet structural verification")
        if any(cell.data_type == "f" for sheet in reopened.worksheets for row in sheet.iter_rows() for cell in row):
            raise ValueError("generated workbook contains executable formulas")
    finally:
        reopened.close()
    return content


def parse_mapping(raw: str | dict[str, Any] | LegacyWorkbookMapping | None) -> Optional[LegacyWorkbookMapping]:
    if raw is None or raw == "":
        return None
    if isinstance(raw, LegacyWorkbookMapping):
        return raw
    try:
        payload = json.loads(raw) if isinstance(raw, str) else raw
    except json.JSONDecodeError as exc:
        raise ModelWorkbookError("invalid_mapping_json", "Mapping must be valid JSON.") from exc
    try:
        return LegacyWorkbookMapping.model_validate(payload)
    except ValidationError as exc:
        raise ModelWorkbookError("invalid_mapping", str(exc)) from exc


def _validate_model_package(content: bytes) -> None:
    """Apply model-workbook limits and reject OOXML active-control parts."""

    if len(content) > MAX_MODEL_FILE_BYTES:
        raise ModelWorkbookError(
            "model_workbook_size_limit",
            "Model workbook exceeds the 32 MB processing limit.",
        )
    try:
        _validate_package(content)
    except MarketWorkbookError as exc:
        raise ModelWorkbookError(exc.code, exc.message) from exc
    # The shared OOXML gate rejects VBA, embeddings, queries, connections, and
    # external relationships. ActiveX/control/custom-UI parts are additional
    # executable surfaces that a nominal .xlsx package can still contain.
    with zipfile.ZipFile(io.BytesIO(content)) as package:
        for member in package.infolist():
            lowered = member.filename.lower()
            if lowered.startswith(_ACTIVE_PACKAGE_PREFIXES):
                raise ModelWorkbookError(
                    "active_or_external_content",
                    "ActiveX controls and custom Office UI parts are not supported.",
                )


def _workbook_value(formula_cell: Any, cached_cell: Any) -> object:
    if formula_cell.data_type == "f" or (
        isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")
    ):
        return cached_cell.value
    return formula_cell.value


def _scan_formulas(formula_wb: Any, cached_wb: Any, ledger: _Ledger) -> list[FormulaAuditEntry]:
    entries: list[FormulaAuditEntry] = []
    dependencies: dict[tuple[str, str], set[tuple[str, str]]] = defaultdict(set)
    formula_locations: set[tuple[str, str]] = set()
    canonical_sheets = {sheet.title.casefold(): sheet.title for sheet in formula_wb.worksheets}
    total_cells = 0
    total_formula_references = 0
    dimension_cells = 0
    dimension_blocked = False
    # Check every declared dimension before iterating any rows. A tiny XML part
    # can otherwise declare several large blank grids and make openpyxl
    # materialize millions of empty cells despite the non-empty-cell cap.
    for sheet in formula_wb.worksheets:
        max_row = sheet.max_row or 0
        max_column = sheet.max_column or 0
        if max_row > MAX_ROWS_PER_SHEET or max_column > MAX_COLUMNS_PER_SHEET:
            ledger.add("blocking", "worksheet_dimension_limit", "Worksheet exceeds model-workbook bounds.", sheet=sheet.title)
            dimension_blocked = True
        dimension_cells += max_row * max_column
    if dimension_cells > MAX_DIMENSION_CELLS:
        ledger.add(
            "blocking",
            "workbook_dimension_limit",
            "Workbook declared dimensions exceed the aggregate processing bound.",
        )
        dimension_blocked = True
    if dimension_blocked:
        return entries

    for sheet in formula_wb.worksheets:
        cached_sheet = cached_wb[sheet.title]
        for formula_row, cached_row in zip(sheet.iter_rows(), cached_sheet.iter_rows()):
            total_cells += sum(cell.value not in (None, "") for cell in formula_row)
            if total_cells > MAX_NONEMPTY_CELLS:
                ledger.add("blocking", "cell_limit", "Workbook contains too many non-empty cells.")
                return entries
            for formula_cell, cached_cell in zip(formula_row, cached_row):
                raw = formula_cell.value
                if not (formula_cell.data_type == "f" or (isinstance(raw, str) and raw.startswith("="))):
                    continue
                if len(entries) >= MAX_FORMULAS:
                    ledger.add("blocking", "formula_limit", "Workbook contains too many formulas.")
                    return entries
                formula = str(raw)
                location = (sheet.title, formula_cell.coordinate.upper())
                formula_locations.add(location)
                blocking: list[str] = []
                cached = _finite(cached_cell.value)
                if len(formula) > MAX_FORMULA_TEXT:
                    blocking.append("formula_length")
                if _formula_external(formula):
                    blocking.append("external_formula")
                if _VOLATILE_FORMULA.search(formula):
                    blocking.append("volatile_formula")
                if _ACTIVE_FORMULA.search(formula) or ("|" in formula and "!" in formula):
                    blocking.append("active_formula")
                if cached is None:
                    blocking.append("formula_cache_required")
                for code in blocking:
                    ledger.add(
                        "blocking",
                        code,
                        {
                            "formula_length": "Formula exceeds the safe length limit.",
                            "external_formula": "External-reference formulas are not supported.",
                            "volatile_formula": "Volatile formulas are not supported.",
                            "active_formula": "Active-data and command formulas are not supported.",
                            "formula_cache_required": "Formula requires a finite cached value; CAOS never calculates it.",
                        }[code],
                        sheet=sheet.title,
                        cell=formula_cell.coordinate,
                    )
                for match in _CELL_REF.finditer(formula):
                    total_formula_references += 1
                    if total_formula_references > MAX_FORMULA_REFERENCES:
                        ledger.add(
                            "blocking",
                            "formula_dependency_limit",
                            "Workbook formulas contain too many cell references.",
                        )
                        return entries
                    raw_target_sheet = (match.group(1) or match.group(2) or sheet.title).strip()
                    target_sheet = canonical_sheets.get(
                        raw_target_sheet.casefold(),
                        raw_target_sheet,
                    )
                    target_cell = match.group(3).replace("$", "").upper()
                    dependencies[location].add((target_sheet, target_cell))
                entries.append(FormulaAuditEntry(
                    sheet=sheet.title,
                    cell=formula_cell.coordinate,
                    formula=formula[:MAX_FORMULA_TEXT],
                    cached_value=cached,
                    disposition="unused",
                    blocking_codes=blocking,
                ))

    graph = {
        source: {target for target in targets if target in formula_locations}
        for source, targets in dependencies.items()
    }
    # Iterative DFS avoids a RecursionError on a validly bounded but deeply
    # chained workbook (formula count is allowed to exceed Python's stack).
    state: dict[tuple[str, str], Literal["visiting", "visited"]] = {}
    circular: set[tuple[str, str]] = set()
    for root in formula_locations:
        if state.get(root) == "visited":
            continue
        path: list[tuple[str, str]] = [root]
        positions = {root: 0}
        state[root] = "visiting"
        stack: list[tuple[tuple[str, str], Any]] = [
            (root, iter(graph.get(root, set())))
        ]
        while stack:
            node, targets = stack[-1]
            try:
                target = next(targets)
            except StopIteration:
                stack.pop()
                state[node] = "visited"
                positions.pop(node, None)
                path.pop()
                continue
            target_state = state.get(target)
            if target_state == "visiting":
                circular.update(path[positions[target]:])
            elif target_state != "visited":
                state[target] = "visiting"
                positions[target] = len(path)
                path.append(target)
                stack.append((target, iter(graph.get(target, set()))))
    if circular:
        by_location = {(entry.sheet, entry.cell.upper()): entry for entry in entries}
        for sheet_name, cell in sorted(circular):
            ledger.add("blocking", "circular_formula", "Circular formula references are not supported.", sheet=sheet_name, cell=cell)
            entry = by_location[(sheet_name, cell)]
            entry.blocking_codes.append("circular_formula")
    return entries


def _require_headers(worksheet: Any, expected: tuple[str, ...], ledger: _Ledger) -> Optional[dict[str, int]]:
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    observed = tuple(str(value).strip() if value is not None else "" for value in headers)
    if observed != expected:
        ledger.add(
            "blocking",
            "invalid_headers",
            f"Expected exact headers: {', '.join(expected)}.",
            sheet=worksheet.title,
        )
        return None
    return {header: index for index, header in enumerate(expected)}


def _metadata(
    formula_wb: Any,
    cached_wb: Any,
    ledger: _Ledger,
) -> tuple[dict[str, object], list[str], list[str]]:
    cover = formula_wb["Cover"]
    cached_cover = cached_wb["Cover"]
    if _require_headers(cover, COVER_HEADERS, ledger) is None:
        return {}, [], []
    values: dict[str, object] = {}
    for formula_row, cached_row in zip(
        cover.iter_rows(min_row=2), cached_cover.iter_rows(min_row=2)
    ):
        key = _read_text(_workbook_value(formula_row[0], cached_row[0]), limit=96)
        if not key:
            continue
        if key in values:
            ledger.add("blocking", "duplicate_metadata", "Metadata key is duplicated.", sheet="Cover", field=key)
            continue
        values[key] = _workbook_value(formula_row[1], cached_row[1])

    sources = formula_wb["Sources - Audit"]
    cached_sources = cached_wb["Sources - Audit"]
    if _require_headers(sources, SOURCE_HEADERS, ledger) is None:
        return values, [], []
    source_metadata: dict[str, object] = {}
    source_ids: list[str] = []
    authority_source_ids: list[str] = []
    for formula_row, cached_row in zip(
        sources.iter_rows(min_row=2), cached_sources.iter_rows(min_row=2)
    ):
        record_type = _read_text(_workbook_value(formula_row[0], cached_row[0]), limit=32)
        key = _read_text(_workbook_value(formula_row[1], cached_row[1]), limit=96)
        value = _workbook_value(formula_row[2], cached_row[2])
        if record_type == "metadata" and key:
            if key in source_metadata:
                ledger.add("blocking", "duplicate_audit_metadata", "Audit metadata key is duplicated.", sheet="Sources - Audit", field=key)
            source_metadata[key] = value
        elif record_type == "payload_source" and key == "source_id":
            source_id = _read_text(value, limit=240)
            if source_id:
                source_ids.append(source_id)
            else:
                ledger.add(
                    "blocking",
                    "invalid_source_audit_row",
                    "Payload source rows require a bounded source ID.",
                    sheet="Sources - Audit",
                )
        elif record_type == "authority_source" and key == "source_id":
            source_id = _read_text(value, limit=240)
            if source_id:
                authority_source_ids.append(source_id)
            else:
                ledger.add(
                    "blocking",
                    "invalid_source_audit_row",
                    "Authority source rows require a bounded source ID.",
                    sheet="Sources - Audit",
                )
        elif any(item not in (None, "") for item in (record_type, key, value)):
            ledger.add(
                "blocking",
                "invalid_source_audit_row",
                "Sources - Audit contains an unsupported record type or key.",
                sheet="Sources - Audit",
            )
    if values != source_metadata:
        ledger.add("blocking", "metadata_mismatch", "Cover and Sources - Audit metadata do not match.")
    return values, source_ids, authority_source_ids


def _resolve_table(
    formula_wb: Any,
    table_name: Literal["assumptions", "debt_schedule", "overrides"],
    table: LegacyTableMapping,
    ledger: _Ledger,
    ambiguities: list[MappingAmbiguity],
) -> Optional[tuple[Any, dict[str, int]]]:
    if table.sheet not in formula_wb.sheetnames:
        ledger.add("blocking", "mapped_sheet_not_found", "Mapped sheet does not exist.", sheet=table.sheet)
        return None
    worksheet = formula_wb[table.sheet]
    headers = [cell.value for cell in next(worksheet.iter_rows(min_row=table.header_row, max_row=table.header_row))]
    by_key: dict[str, list[int]] = defaultdict(list)
    for index, header in enumerate(headers):
        key = _header_key(header)
        if key:
            by_key[key].append(index)
    resolved: dict[str, int] = {}
    for field_name, mapped_header in table.columns.items():
        candidates = by_key.get(_header_key(mapped_header), [])
        selected_column = table.column_indices.get(field_name)
        if selected_column is not None:
            selected_index = selected_column - 1
            if selected_index >= len(headers):
                ledger.add(
                    "blocking",
                    "mapped_column_selector_out_of_range",
                    "Selected mapped column is outside the worksheet header range.",
                    sheet=table.sheet,
                    field=field_name,
                )
            elif _header_key(headers[selected_index]) != _header_key(mapped_header):
                ledger.add(
                    "blocking",
                    "mapped_column_selector_mismatch",
                    "Selected mapped column does not match the reviewed header.",
                    sheet=table.sheet,
                    field=field_name,
                )
            else:
                resolved[field_name] = selected_index
        elif len(candidates) == 1:
            resolved[field_name] = candidates[0]
        elif len(candidates) > 1:
            labels = [f"{mapped_header} ({index + 1})" for index in candidates]
            ambiguities.append(MappingAmbiguity(
                table=table_name,
                field=field_name,
                candidates=labels,
                message="Mapped header appears more than once; select a unique source column.",
            ))
            ledger.add("blocking", "ambiguous_mapping", "Mapped header appears more than once.", sheet=table.sheet, field=field_name)
        else:
            ledger.add("blocking", "mapped_column_not_found", "Mapped header was not found.", sheet=table.sheet, field=field_name)
    if len(resolved) != len(table.columns):
        return None
    return worksheet, resolved


def _resolve_matrix(
    formula_wb: Any,
    matrix: LegacyMatrixMapping,
    ledger: _Ledger,
    ambiguities: list[MappingAmbiguity],
) -> Optional[tuple[Any, int, dict[str, int], dict[str, int]]]:
    """Resolve explicit matrix selectors to physical one-based rows/zero-based columns."""

    if matrix.sheet not in formula_wb.sheetnames:
        ledger.add(
            "blocking",
            "mapped_sheet_not_found",
            "Mapped matrix sheet does not exist.",
            sheet=matrix.sheet,
        )
        return None
    worksheet = formula_wb[matrix.sheet]
    headers = [
        cell.value
        for cell in next(
            worksheet.iter_rows(min_row=matrix.header_row, max_row=matrix.header_row)
        )
    ]
    by_header: dict[str, list[int]] = defaultdict(list)
    for index, header in enumerate(headers):
        key = _header_key(header)
        if key:
            by_header[key].append(index)

    account_candidates = by_header.get(_header_key(matrix.account_column), [])
    account_index: Optional[int] = None
    if matrix.account_column_index is not None:
        selected = matrix.account_column_index - 1
        if selected >= len(headers):
            ledger.add(
                "blocking",
                "matrix_account_selector_out_of_range",
                "Selected account column is outside the worksheet header range.",
                sheet=matrix.sheet,
                field="account_column",
            )
        elif _header_key(headers[selected]) != _header_key(matrix.account_column):
            ledger.add(
                "blocking",
                "matrix_account_selector_mismatch",
                "Selected account column does not match the reviewed header.",
                sheet=matrix.sheet,
                field="account_column",
            )
        else:
            account_index = selected
    elif len(account_candidates) == 1:
        account_index = account_candidates[0]
    elif len(account_candidates) > 1:
        ambiguities.append(MappingAmbiguity(
            table="assumptions",
            field="account_column",
            selector="column",
            candidates=[f"{matrix.account_column} ({index + 1})" for index in account_candidates],
            message="Account header appears more than once; select its physical column.",
        ))
        ledger.add(
            "blocking",
            "ambiguous_mapping",
            "Matrix account header appears more than once.",
            sheet=matrix.sheet,
            field="account_column",
        )
    else:
        ledger.add(
            "blocking",
            "mapped_column_not_found",
            "Matrix account header was not found.",
            sheet=matrix.sheet,
            field="account_column",
        )

    period_columns: dict[str, int] = {}
    for period_key, mapped_header in matrix.period_columns.items():
        candidates = by_header.get(_header_key(mapped_header), [])
        selected_column = matrix.period_column_indices.get(period_key)
        if selected_column is not None:
            selected = selected_column - 1
            if selected >= len(headers):
                ledger.add(
                    "blocking",
                    "matrix_period_selector_out_of_range",
                    "Selected period column is outside the worksheet header range.",
                    sheet=matrix.sheet,
                    field=period_key,
                )
            elif _header_key(headers[selected]) != _header_key(mapped_header):
                ledger.add(
                    "blocking",
                    "matrix_period_selector_mismatch",
                    "Selected period column does not match the reviewed header.",
                    sheet=matrix.sheet,
                    field=period_key,
                )
            else:
                period_columns[period_key] = selected
        elif len(candidates) == 1:
            period_columns[period_key] = candidates[0]
        elif len(candidates) > 1:
            ambiguities.append(MappingAmbiguity(
                table="assumptions",
                field=period_key,
                selector="column",
                candidates=[f"{mapped_header} ({index + 1})" for index in candidates],
                message="Mapped period header appears more than once; select its physical column.",
            ))
            ledger.add(
                "blocking",
                "ambiguous_mapping",
                "Mapped matrix period header appears more than once.",
                sheet=matrix.sheet,
                field=period_key,
            )
        else:
            ledger.add(
                "blocking",
                "mapped_column_not_found",
                "Mapped matrix period header was not found.",
                sheet=matrix.sheet,
                field=period_key,
            )

    account_rows: dict[str, int] = {}
    if account_index is not None:
        row_candidates: dict[str, list[int]] = defaultdict(list)
        for row_number in range(matrix.header_row + 1, (worksheet.max_row or 0) + 1):
            key = _header_key(worksheet.cell(row_number, account_index + 1).value)
            if key:
                row_candidates[key].append(row_number)
        for field_name, mapped_label in matrix.account_rows.items():
            candidates = row_candidates.get(_header_key(mapped_label), [])
            selected_row = matrix.account_row_indices.get(field_name)
            if selected_row is not None:
                if selected_row > (worksheet.max_row or 0):
                    ledger.add(
                        "blocking",
                        "matrix_account_row_selector_out_of_range",
                        "Selected account row is outside the worksheet range.",
                        sheet=matrix.sheet,
                        field=field_name,
                    )
                elif _header_key(
                    worksheet.cell(selected_row, account_index + 1).value
                ) != _header_key(mapped_label):
                    ledger.add(
                        "blocking",
                        "matrix_account_row_selector_mismatch",
                        "Selected account row does not match the reviewed label.",
                        sheet=matrix.sheet,
                        field=field_name,
                    )
                else:
                    account_rows[field_name] = selected_row
            elif len(candidates) == 1:
                account_rows[field_name] = candidates[0]
            elif len(candidates) > 1:
                ambiguities.append(MappingAmbiguity(
                    table="assumptions",
                    field=field_name,
                    selector="row",
                    candidates=[f"{mapped_label} ({row})" for row in candidates],
                    message="Mapped account label appears more than once; select its physical row.",
                ))
                ledger.add(
                    "blocking",
                    "ambiguous_mapping",
                    "Mapped matrix account label appears more than once.",
                    sheet=matrix.sheet,
                    field=field_name,
                )
            else:
                ledger.add(
                    "blocking",
                    "mapped_account_not_found",
                    "Mapped matrix account label was not found.",
                    sheet=matrix.sheet,
                    field=field_name,
                )

    if len(set(account_rows.values())) != len(account_rows):
        ledger.add(
            "blocking",
            "matrix_account_collision",
            "Multiple model accounts resolve to the same physical row.",
            sheet=matrix.sheet,
        )
    if len(set(period_columns.values())) != len(period_columns):
        ledger.add(
            "blocking",
            "matrix_period_collision",
            "Multiple stable periods resolve to the same physical column.",
            sheet=matrix.sheet,
        )
    if (
        account_index is None
        or len(account_rows) != len(matrix.account_rows)
        or len(period_columns) != len(matrix.period_columns)
        or any(
            issue.code in {"matrix_account_collision", "matrix_period_collision"}
            for issue in ledger.issues
        )
    ):
        return None
    return worksheet, account_index, account_rows, period_columns


def _iter_table_rows(formula_ws: Any, cached_ws: Any, start_row: int, resolved: dict[str, int]):
    for row_number, (formula_row, cached_row) in enumerate(
        zip(formula_ws.iter_rows(min_row=start_row), cached_ws.iter_rows(min_row=start_row)),
        start=start_row,
    ):
        raw = {
            field: _workbook_value(formula_row[index], cached_row[index])
            for field, index in resolved.items()
        }
        if any(value not in (None, "") for value in raw.values()):
            yield row_number, raw


def _authority(raw: dict[str, object], default_sources: list[str], default_as_of: Optional[datetime]) -> Optional[ModelAuthority]:
    source_ids = _json_list(raw.get("authority_source_ids"))
    if source_ids is None:
        return None
    try:
        return ModelAuthority(
            origin=_read_text(raw.get("authority_origin"), limit=24) or "imported",
            method=_read_text(raw.get("authority_method"), limit=64) or "model_workbook",
            source_ids=source_ids if "authority_source_ids" in raw else default_sources,
            as_of=_datetime(raw.get("authority_as_of")) or default_as_of,
        )
    except ValidationError:
        return None


def _build_matrix_payload(
    formula_wb: Any,
    cached_wb: Any,
    *,
    matrix: LegacyMatrixMapping,
    resolved: tuple[Any, int, dict[str, int], dict[str, int]],
    reporting_currency: str,
    reporting_unit: str,
    source_ids: list[str],
    authority_as_of: Optional[datetime],
    ledger: _Ledger,
) -> tuple[Optional[ModelDraftPayload], set[tuple[str, str]]]:
    """Build period inputs from a reviewed matrix without executing formulas."""

    from openpyxl.utils import get_column_letter

    worksheet, _account_index, account_rows, period_columns = resolved
    cached = cached_wb[worksheet.title]
    formula_cells: set[tuple[str, str]] = set()
    periods: list[ModelPeriodInput] = []
    has_formula_input = False
    authority = ModelAuthority(
        origin="imported",
        method="model_workbook",
        source_ids=source_ids,
        as_of=authority_as_of,
    )
    for period_key, column_index in period_columns.items():
        values: dict[str, Optional[float]] = {}
        invalid = False
        for field_name, row_number in account_rows.items():
            formula_cell = worksheet.cell(row_number, column_index + 1)
            cached_cell = cached.cell(row_number, column_index + 1)
            cell_ref = f"{get_column_letter(column_index + 1)}{row_number}"
            if formula_cell.data_type == "f" or (
                isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")
            ):
                formula_cells.add((worksheet.title, cell_ref))
                has_formula_input = True
                invalid = True
                ledger.add(
                    "blocking",
                    "matrix_formula_not_authoritative",
                    "Mapped matrix formulas are comparison evidence only; bind a finite value cell.",
                    sheet=worksheet.title,
                    cell=cell_ref,
                    field=field_name,
                )
                values[field_name] = None
                continue
            value = _workbook_value(formula_cell, cached_cell)
            if value in (None, ""):
                values[field_name] = None
                continue
            parsed = _finite(value)
            if parsed is None:
                invalid = True
                ledger.add(
                    "blocking",
                    "invalid_number",
                    "Mapped matrix input must be a finite numeric cell.",
                    sheet=worksheet.title,
                    cell=cell_ref,
                    field=field_name,
                )
            values[field_name] = parsed
        if invalid:
            continue
        label = matrix.period_labels.get(period_key) or _read_text(
            worksheet.cell(matrix.header_row, column_index + 1).value,
            limit=64,
        )
        try:
            periods.append(ModelPeriodInput(
                period_key=period_key,
                label=label or period_key,
                kind=matrix.period_kinds[period_key],
                months=(
                    matrix.period_months.get(period_key)
                    or _default_period_months(period_key)
                ),
                authority=authority,
                **values,
            ))
        except ValidationError as exc:
            ledger.add(
                "blocking",
                "invalid_period",
                str(exc),
                sheet=worksheet.title,
                field=period_key,
            )
    if has_formula_input:
        return None, formula_cells
    if not periods:
        ledger.add(
            "blocking",
            "no_model_periods",
            "Mapped matrix contains no valid model periods.",
            sheet=worksheet.title,
        )
        return None, formula_cells
    try:
        return ModelDraftPayload(
            reporting_currency=reporting_currency,
            reporting_unit=reporting_unit,
            periods=periods,
            debt_instruments=[],
            overrides=[],
            ui_preferences=ModelUiPreferences(),
            source_ids=source_ids,
        ), formula_cells
    except ValidationError as exc:
        ledger.add("blocking", "invalid_model_payload", str(exc), sheet=worksheet.title)
        return None, formula_cells


def _build_payload(
    formula_wb: Any,
    cached_wb: Any,
    *,
    assumption_sheet: str,
    assumption_start: int,
    assumption_columns: dict[str, int],
    debt_table: Optional[tuple[str, int, dict[str, int]]],
    override_table: Optional[tuple[str, int, dict[str, int]]],
    reporting_currency: str,
    reporting_unit: str,
    source_ids: list[str],
    authority_as_of: Optional[datetime],
    ui_preferences: Optional[ModelUiPreferences],
    strict: bool,
    ledger: _Ledger,
) -> Optional[ModelDraftPayload]:
    periods: list[ModelPeriodInput] = []
    formula_ws = formula_wb[assumption_sheet]
    cached_ws = cached_wb[assumption_sheet]
    for row_number, raw in _iter_table_rows(formula_ws, cached_ws, assumption_start, assumption_columns):
        raw_period_key = _read_text(raw.get("period_key"), limit=64)
        period_key = _normalize_period_key(raw_period_key)
        if raw_period_key and period_key is None:
            ledger.add(
                "blocking",
                "unsupported_period_alias",
                "Period must be a supported fiscal, quarter, YTD, LTM, or ISO-date alias.",
                sheet=assumption_sheet,
                cell=f"period:{row_number}",
                field="period_key",
            )
        row_key = _read_text(raw.get("row_key"), limit=300)
        if strict and period_key and row_key != f"period:{period_key}":
            ledger.add("blocking", "unstable_row_key", "Assumption row_key does not match its stable period key.", sheet=assumption_sheet, cell=f"A{row_number}")
        authority = _authority(raw, source_ids, authority_as_of)
        values: dict[str, Optional[float]] = {}
        invalid = False
        for field_name in PERIOD_FIELDS:
            value = raw.get(field_name)
            if value in (None, ""):
                values[field_name] = None
            else:
                parsed = _finite(value)
                if parsed is None:
                    invalid = True
                    ledger.add("blocking", "invalid_number", "Model input must be a finite numeric cell.", sheet=assumption_sheet, field=field_name)
                values[field_name] = parsed
        months_raw = raw.get("months")
        months = (
            _default_period_months(period_key)
            if months_raw in (None, "") and period_key is not None
            else _integer(months_raw)
        )
        if months is None:
            invalid = True
            ledger.add(
                "blocking",
                "invalid_months",
                "months must be an integer; ISO-date periods require an explicit value.",
                sheet=assumption_sheet,
            )
        units = _read_text(raw.get("units"), limit=32)
        if strict and units != reporting_unit:
            invalid = True
            ledger.add("blocking", "unit_mismatch", "Assumption units do not match workbook reporting_unit.", sheet=assumption_sheet)
        if not period_key or authority is None:
            invalid = True
            ledger.add("blocking", "invalid_period_authority", "Period key and valid authority are required.", sheet=assumption_sheet)
        if invalid:
            continue
        try:
            periods.append(ModelPeriodInput(
                period_key=period_key,
                label=_read_text(raw.get("label"), limit=64) or "",
                kind=_read_text(raw.get("kind"), limit=24) or "",
                months=months,
                authority=authority,
                **values,
            ))
        except ValidationError as exc:
            ledger.add("blocking", "invalid_period", str(exc), sheet=assumption_sheet)

    instruments: list[DebtInstrument] = []
    if debt_table is not None:
        sheet_name, start_row, columns = debt_table
        grouped: dict[str, dict[str, Any]] = {}
        for row_number, raw in _iter_table_rows(formula_wb[sheet_name], cached_wb[sheet_name], start_row, columns):
            instrument_id = _read_text(raw.get("instrument_id"), limit=96)
            raw_period_key = _read_text(raw.get("period_key"), limit=64)
            period_key = _normalize_period_key(raw_period_key)
            if raw_period_key and period_key is None:
                ledger.add(
                    "blocking",
                    "unsupported_period_alias",
                    "Debt period must be a supported fiscal, quarter, YTD, LTM, or ISO-date alias.",
                    sheet=sheet_name,
                    cell=f"period:{row_number}",
                    field="period_key",
                )
            row_key = _read_text(raw.get("row_key"), limit=300)
            if strict and instrument_id and period_key and row_key != f"debt:{instrument_id}:{period_key}":
                ledger.add("blocking", "unstable_row_key", "Debt row_key does not match instrument and period keys.", sheet=sheet_name, cell=f"A{row_number}")
            priority = _integer(raw.get("priority"))
            authority = _authority(raw, source_ids, authority_as_of)
            instrument_sources = _json_list(raw.get("source_ids"))
            point_values: dict[str, object] = {"period_key": period_key}
            invalid = not instrument_id or not period_key or priority is None or authority is None or instrument_sources is None
            for field_name in DEBT_PERIOD_FIELDS:
                value = raw.get(field_name)
                if value in (None, ""):
                    continue
                parsed = _finite(value)
                if parsed is None:
                    invalid = True
                    ledger.add("blocking", "invalid_number", "Debt input must be a finite numeric cell.", sheet=sheet_name, field=field_name)
                else:
                    point_values[field_name] = parsed
            units = _read_text(raw.get("units"), limit=32)
            if strict and units != reporting_unit:
                invalid = True
                ledger.add("blocking", "unit_mismatch", "Debt units do not match workbook reporting_unit.", sheet=sheet_name)
            if invalid:
                ledger.add("blocking", "invalid_debt_row", "Debt row is missing a required stable key, authority, or typed value.", sheet=sheet_name)
                continue
            metadata = {
                "instrument_id": instrument_id,
                "name": _read_text(raw.get("name"), limit=160) or "",
                "priority": priority,
                "seniority": _read_text(raw.get("seniority"), limit=64) or "",
                "currency": _read_text(raw.get("currency"), limit=3) or "",
                "rate_type": _read_text(raw.get("rate_type"), limit=16),
                "maturity": _read_text(raw.get("maturity"), limit=24),
                "amortization": _read_text(raw.get("amortization"), limit=240),
                "benchmark_curve": _read_text(raw.get("benchmark_curve"), limit=120),
                "sources": instrument_sources,
                "authority": authority,
            }
            existing = grouped.get(instrument_id)
            if existing is not None and existing["metadata"] != metadata:
                ledger.add("blocking", "inconsistent_instrument", "Repeated debt instrument metadata is inconsistent.", sheet=sheet_name, field=instrument_id)
                continue
            try:
                point = DebtPeriod.model_validate(point_values)
            except ValidationError as exc:
                ledger.add("blocking", "invalid_debt_period", str(exc), sheet=sheet_name)
                continue
            grouped.setdefault(instrument_id, {"metadata": metadata, "periods": []})["periods"].append(point)
        for item in grouped.values():
            try:
                instruments.append(DebtInstrument(**item["metadata"], periods=item["periods"]))
            except ValidationError as exc:
                ledger.add("blocking", "invalid_debt_instrument", str(exc))

    overrides: list[CellOverride] = []
    invalid_override_expiry = False
    if override_table is not None:
        sheet_name, start_row, columns = override_table
        for row_number, raw in _iter_table_rows(formula_wb[sheet_name], cached_wb[sheet_name], start_row, columns):
            node_id = _read_text(raw.get("node_id"), limit=300)
            row_key = _read_text(raw.get("row_key"), limit=300)
            if strict and node_id and row_key != node_id:
                ledger.add("blocking", "unstable_row_key", "Override row_key must equal node_id.", sheet=sheet_name, cell=f"A{row_number}")
            value_type = _read_text(raw.get("value_type"), limit=16)
            value = raw.get("value")
            parsed_value = None if value in (None, "") else _finite(value)
            if value not in (None, "") and parsed_value is None:
                ledger.add("blocking", "invalid_override_value", "Override value must be finite or blank for null.", sheet=sheet_name)
                continue
            raw_expires_at = raw.get("expires_at")
            expires_at = _datetime(raw_expires_at)
            has_expiry = raw_expires_at is not None and not (
                isinstance(raw_expires_at, str) and not raw_expires_at.strip()
            )
            if has_expiry and expires_at is None:
                ledger.add(
                    "blocking",
                    "invalid_override_expiry",
                    "Override expires_at must be a valid ISO timestamp or blank.",
                    sheet=sheet_name,
                    field="expires_at",
                )
                invalid_override_expiry = True
                continue
            try:
                overrides.append(CellOverride(
                    node_id=node_id or "",
                    value_type=value_type or "",
                    value=parsed_value,
                    reason=_read_text(raw.get("reason"), limit=1_000),
                    scope=_read_text(raw.get("scope"), limit=64) or "draft",
                    source=_read_text(raw.get("source"), limit=240),
                    expires_at=expires_at,
                ))
            except ValidationError as exc:
                ledger.add("blocking", "invalid_override", str(exc), sheet=sheet_name)

    if invalid_override_expiry:
        return None
    if not periods:
        ledger.add("blocking", "no_model_periods", "Workbook contains no valid model periods.")
        return None
    try:
        return ModelDraftPayload(
            reporting_currency=reporting_currency,
            reporting_unit=reporting_unit,
            periods=periods,
            debt_instruments=instruments,
            overrides=overrides,
            ui_preferences=ui_preferences or ModelUiPreferences(),
            source_ids=source_ids,
        )
    except ValidationError as exc:
        ledger.add("blocking", "invalid_model_payload", str(exc))
        return None


def _mark_formula_dispositions(
    audit: list[FormulaAuditEntry],
    input_cells: set[tuple[str, str]],
    comparison_cells: set[tuple[str, str]],
) -> None:
    for entry in audit:
        location = (entry.sheet, entry.cell.upper())
        if location in input_cells:
            entry.disposition = "input_candidate"
        elif location in comparison_cells:
            entry.disposition = "comparison_only"


def _mapped_formula_cells(sheet: str, start_row: int, columns: dict[str, int], authoritative_fields: set[str], worksheet: Any) -> set[tuple[str, str]]:
    from openpyxl.utils import get_column_letter

    result: set[tuple[str, str]] = set()
    for row_number, row in enumerate(worksheet.iter_rows(min_row=start_row), start=start_row):
        for field_name, index in columns.items():
            if field_name in authoritative_fields and index < len(row):
                result.add((sheet, f"{get_column_letter(index + 1)}{row_number}"))
    return result


def _tie_out_model(
    formula_wb: Any,
    cached_wb: Any,
    calculation: ModelCalculation,
    reporting_unit: str,
    ledger: _Ledger,
) -> None:
    worksheet = formula_wb["Model"]
    cached = cached_wb["Model"]
    columns = _require_headers(worksheet, MODEL_HEADERS, ledger)
    if columns is None:
        return
    expected: dict[tuple[str, str], dict[str, object]] = {}
    for period in calculation.periods:
        for metric in MODEL_FIELDS:
            original, formula, overridden = _formula_for_metric(period, metric)
            expected[(f"model:{metric}", period.period_key)] = {
                "period_label": period.label,
                "metric": metric,
                "value": getattr(period, metric),
                "original_value": original,
                "formula": formula,
                "overridden": overridden,
                "units": (
                    "multiple"
                    if metric in {"gross_leverage", "net_leverage", "interest_coverage"}
                    else reporting_unit
                ),
            }

    observed: dict[tuple[str, str], dict[str, object]] = {}
    for row_number, raw in _iter_table_rows(worksheet, cached, 2, columns):
        row_key = _read_text(raw.get("row_key"), limit=300)
        period_key = _read_text(raw.get("period_key"), limit=24)
        metric = _read_text(raw.get("metric"), limit=96)
        if row_key != f"model:{metric}":
            ledger.add("blocking", "unstable_model_row_key", "Model row_key does not match metric.", sheet="Model")
            continue
        key = (row_key or "", period_key or "")
        if key in observed:
            ledger.add("blocking", "duplicate_model_row", "Model row identity is duplicated.", sheet="Model")
            continue
        parsed_numbers: dict[str, Optional[float]] = {}
        for field_name in ("value", "original_value"):
            value = raw.get(field_name)
            parsed = None if value in (None, "") else _finite(value)
            if value not in (None, "") and parsed is None:
                ledger.add(
                    "blocking",
                    "invalid_model_number",
                    "Model presentation numbers must be finite or blank.",
                    sheet="Model",
                    cell=f"{field_name}:{row_number}",
                )
            parsed_numbers[field_name] = parsed
        overridden = raw.get("overridden")
        if type(overridden) is not bool:
            ledger.add(
                "blocking",
                "invalid_model_override_flag",
                "Model overridden flag must be a Boolean cell.",
                sheet="Model",
                cell=f"H{row_number}",
            )
        observed[key] = {
            "period_label": _read_text(raw.get("period_label"), limit=64),
            "metric": metric,
            **parsed_numbers,
            "formula": _read_text(raw.get("formula"), limit=MAX_FORMULA_TEXT),
            "overridden": overridden if type(overridden) is bool else None,
            "units": _read_text(raw.get("units"), limit=32),
        }
    if set(observed) != set(expected):
        ledger.add("blocking", "model_row_set_mismatch", "Model presentation rows do not match the canonical calculation row set.", sheet="Model")
        return
    for key, expected_row in expected.items():
        actual_row = observed[key]
        for field_name, expected_value in expected_row.items():
            actual = actual_row[field_name]
            if field_name in {"value", "original_value"}:
                ties = (
                    expected_value is None and actual is None
                ) or (
                    expected_value is not None
                    and actual is not None
                    and math.isclose(
                        float(actual),
                        float(expected_value),
                        rel_tol=1e-9,
                        abs_tol=1e-8,
                    )
                )
            else:
                ties = actual == expected_value
            if not ties:
                ledger.add(
                    "blocking",
                    "model_tieout_mismatch",
                    "Model presentation and audit fields must tie to Model Engine v2.",
                    sheet="Model",
                    field=f"{key[0]}:{key[1]}:{field_name}",
                )


def preview_workbook(
    content: bytes,
    *,
    filename: str,
    mapping: str | dict[str, Any] | LegacyWorkbookMapping | None = None,
    evaluated_at: Optional[datetime] = None,
) -> ModelWorkbookPreview:
    """Safely derive a payload and engine result without side effects."""

    if not filename or not filename.lower().endswith(".xlsx"):
        raise ModelWorkbookError("xlsx_required", "Model imports accept .xlsx files only.")
    _validate_model_package(content)
    parsed_mapping = parse_mapping(mapping)
    mode: Literal["strict_v1", "mapped_legacy"] = "mapped_legacy" if parsed_mapping else "strict_v1"
    ledger = _Ledger()
    ambiguities: list[MappingAmbiguity] = []

    try:
        from openpyxl import load_workbook

        formula_wb = load_workbook(io.BytesIO(content), read_only=True, data_only=False, keep_links=False)
        cached_wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True, keep_links=False)
    except Exception as exc:
        raise ModelWorkbookError("workbook_parse_failed", "Workbook could not be opened safely.") from exc

    try:
        if len(formula_wb.sheetnames) > MAX_MODEL_SHEETS:
            ledger.add("blocking", "sheet_limit", "Workbook contains too many sheets.")
        for sheet in formula_wb.worksheets:
            if sheet.sheet_state != "visible":
                ledger.add("blocking", "hidden_sheet", "Hidden sheets are not accepted.", sheet=sheet.title)
        audit = _scan_formulas(formula_wb, cached_wb, ledger)
        if any(
            issue.code in {
                "worksheet_dimension_limit",
                "workbook_dimension_limit",
                "cell_limit",
                "formula_limit",
                "formula_dependency_limit",
            }
            for issue in ledger.issues
        ):
            return ModelWorkbookPreview(
                mode=mode,
                workbook_sha256=hashlib.sha256(content).hexdigest(),
                sheet_names=list(formula_wb.sheetnames),
                mapping=parsed_mapping,
                draft_payload=None,
                calculation=None,
                embedded_hashes=EmbeddedHashes(),
                formula_audit=audit,
                ambiguities=ambiguities,
                issues=ledger.issues,
                blocking_count=sum(issue.severity == "blocking" for issue in ledger.issues),
                warning_count=sum(issue.severity == "warning" for issue in ledger.issues),
            )
        embedded = EmbeddedHashes()
        identity: Optional[WorkbookIdentity] = None
        payload: Optional[ModelDraftPayload] = None
        calculation: Optional[ModelCalculation] = None
        audit_authority_source_ids: list[str] = []
        input_cells: set[tuple[str, str]] = set()
        comparison_cells: set[tuple[str, str]] = set()
        calculation_time = evaluated_at

        if mode == "strict_v1":
            if tuple(formula_wb.sheetnames) != SHEET_NAMES:
                ledger.add("blocking", "sheet_contract", "Strict CAOS v1 requires exactly the six canonical sheets in order.")
            if all(name in formula_wb.sheetnames for name in SHEET_NAMES):
                metadata, source_ids, audit_authority_source_ids = _metadata(
                    formula_wb,
                    cached_wb,
                    ledger,
                )
                required_metadata = {
                    "schema_id", "schema_version", "engine_version", "issuer_id",
                    "draft_revision", "reporting_currency", "reporting_unit", "ui_preferences",
                    "source_fingerprint", "input_fingerprint", "calculation_hash",
                    "exported_by", "exported_at",
                }
                missing = sorted(required_metadata - set(metadata))
                if missing:
                    ledger.add("blocking", "missing_metadata", f"Missing metadata: {', '.join(missing)}.")
                unexpected = sorted(set(metadata) - required_metadata)
                if unexpected:
                    ledger.add(
                        "blocking",
                        "unexpected_metadata",
                        f"Unsupported metadata: {', '.join(unexpected)}.",
                    )
                engine_version = _read_text(metadata.get("engine_version"), limit=32)
                source_hash = _read_text(metadata.get("source_fingerprint"), limit=64)
                input_hash = _read_text(metadata.get("input_fingerprint"), limit=64)
                calculation_hash = _read_text(metadata.get("calculation_hash"), limit=64)
                embedded = EmbeddedHashes(
                    engine_version=engine_version,
                    source_fingerprint=source_hash,
                    input_fingerprint=input_hash,
                    calculation_hash=calculation_hash,
                )
                if metadata.get("schema_id") != SCHEMA_ID or _integer(metadata.get("schema_version")) != 1:
                    ledger.add("blocking", "schema_version", "Workbook is not the supported CAOS model-workbook v1 schema.")
                if engine_version != ENGINE_VERSION:
                    ledger.add("blocking", "engine_version", "Workbook engine version is not supported.")
                for name, value in (("source_fingerprint", source_hash), ("input_fingerprint", input_hash), ("calculation_hash", calculation_hash)):
                    if value is None or not _HASH.fullmatch(value):
                        ledger.add("blocking", "invalid_hash", "Embedded hash must be lowercase SHA-256.", field=name)
                issuer_id = _read_text(metadata.get("issuer_id"), limit=64)
                draft_revision = _integer(metadata.get("draft_revision"))
                exported_by = _read_text(metadata.get("exported_by"), limit=255)
                calculation_time = _datetime(metadata.get("exported_at"))
                if issuer_id is None:
                    ledger.add("blocking", "invalid_issuer_id", "issuer_id must be a bounded non-empty identifier.")
                if draft_revision is None or draft_revision < 1:
                    ledger.add("blocking", "invalid_draft_revision", "draft_revision must be a positive integer.")
                if exported_by is None:
                    ledger.add("blocking", "invalid_exported_by", "exported_by must be a bounded non-empty identifier.")
                if calculation_time is None:
                    ledger.add("blocking", "invalid_exported_at", "exported_at must be an ISO timestamp.")
                if (
                    issuer_id is not None
                    and draft_revision is not None
                    and draft_revision >= 1
                    and exported_by is not None
                    and calculation_time is not None
                ):
                    identity = WorkbookIdentity(
                        issuer_id=issuer_id,
                        draft_revision=draft_revision,
                        exported_by=exported_by,
                        exported_at=calculation_time,
                    )
                preferences = _ui_preferences(metadata.get("ui_preferences"))
                if preferences is None:
                    ledger.add("blocking", "invalid_ui_preferences", "ui_preferences must be valid strict JSON.")

                assumption_columns = _require_headers(formula_wb["Assumptions"], ASSUMPTION_HEADERS, ledger)
                debt_columns = _require_headers(formula_wb["Debt Schedule"], DEBT_HEADERS, ledger)
                override_columns = _require_headers(formula_wb["Overrides"], OVERRIDE_HEADERS, ledger)
                if assumption_columns is not None and debt_columns is not None and override_columns is not None:
                    payload = _build_payload(
                        formula_wb,
                        cached_wb,
                        assumption_sheet="Assumptions",
                        assumption_start=2,
                        assumption_columns=assumption_columns,
                        debt_table=("Debt Schedule", 2, debt_columns),
                        override_table=("Overrides", 2, override_columns),
                        reporting_currency=_read_text(metadata.get("reporting_currency"), limit=3) or "",
                        reporting_unit=_read_text(metadata.get("reporting_unit"), limit=32) or "",
                        source_ids=source_ids,
                        authority_as_of=None,
                        ui_preferences=preferences,
                        strict=True,
                        ledger=ledger,
                    )
                    input_cells |= _mapped_formula_cells("Assumptions", 2, assumption_columns, set(PERIOD_FIELDS) | {"months"}, formula_wb["Assumptions"])
                    input_cells |= _mapped_formula_cells("Debt Schedule", 2, debt_columns, set(DEBT_PERIOD_FIELDS) | {"priority"}, formula_wb["Debt Schedule"])
                    input_cells |= _mapped_formula_cells("Overrides", 2, override_columns, {"value"}, formula_wb["Overrides"])
                    comparison_cells |= _mapped_formula_cells(
                        "Model",
                        2,
                        {
                            "value": MODEL_HEADERS.index("value"),
                            "original_value": MODEL_HEADERS.index("original_value"),
                        },
                        {"value", "original_value"},
                        formula_wb["Model"],
                    )
            else:
                metadata = {}
        else:
            assert parsed_mapping is not None
            if isinstance(parsed_mapping.assumptions, LegacyMatrixMapping):
                resolved_matrix = _resolve_matrix(
                    formula_wb,
                    parsed_mapping.assumptions,
                    ledger,
                    ambiguities,
                )
                if resolved_matrix is not None:
                    payload, matrix_formula_cells = _build_matrix_payload(
                        formula_wb,
                        cached_wb,
                        matrix=parsed_mapping.assumptions,
                        resolved=resolved_matrix,
                        reporting_currency=parsed_mapping.reporting_currency,
                        reporting_unit=parsed_mapping.reporting_unit,
                        source_ids=parsed_mapping.source_ids,
                        authority_as_of=parsed_mapping.authority_as_of,
                        ledger=ledger,
                    )
                    # Imported formulas are retained in the formula audit only;
                    # they never become authoritative matrix inputs.
                    comparison_cells |= matrix_formula_cells
            else:
                resolved_assumptions = _resolve_table(
                    formula_wb,
                    "assumptions",
                    parsed_mapping.assumptions,
                    ledger,
                    ambiguities,
                )
                resolved_debt = _resolve_table(formula_wb, "debt_schedule", parsed_mapping.debt_schedule, ledger, ambiguities) if parsed_mapping.debt_schedule else None
                resolved_overrides = _resolve_table(formula_wb, "overrides", parsed_mapping.overrides, ledger, ambiguities) if parsed_mapping.overrides else None
                if resolved_assumptions is not None:
                    assumption_ws, assumption_columns = resolved_assumptions
                    debt_tuple = None
                    if resolved_debt is not None and parsed_mapping.debt_schedule is not None:
                        debt_tuple = (resolved_debt[0].title, parsed_mapping.debt_schedule.header_row + 1, resolved_debt[1])
                    override_tuple = None
                    if resolved_overrides is not None and parsed_mapping.overrides is not None:
                        override_tuple = (resolved_overrides[0].title, parsed_mapping.overrides.header_row + 1, resolved_overrides[1])
                    payload = _build_payload(
                        formula_wb,
                        cached_wb,
                        assumption_sheet=assumption_ws.title,
                        assumption_start=parsed_mapping.assumptions.header_row + 1,
                        assumption_columns=assumption_columns,
                        debt_table=debt_tuple,
                        override_table=override_tuple,
                        reporting_currency=parsed_mapping.reporting_currency,
                        reporting_unit=parsed_mapping.reporting_unit,
                        source_ids=parsed_mapping.source_ids,
                        authority_as_of=parsed_mapping.authority_as_of,
                        ui_preferences=ModelUiPreferences(),
                        strict=False,
                        ledger=ledger,
                    )
                    input_cells |= _mapped_formula_cells(
                        assumption_ws.title,
                        parsed_mapping.assumptions.header_row + 1,
                        assumption_columns,
                        set(assumption_columns),
                        assumption_ws,
                    )
                    if resolved_debt is not None and parsed_mapping.debt_schedule is not None:
                        input_cells |= _mapped_formula_cells(
                            resolved_debt[0].title,
                            parsed_mapping.debt_schedule.header_row + 1,
                            resolved_debt[1],
                            set(resolved_debt[1]),
                            resolved_debt[0],
                        )
                    if resolved_overrides is not None and parsed_mapping.overrides is not None:
                        input_cells |= _mapped_formula_cells(
                            resolved_overrides[0].title,
                            parsed_mapping.overrides.header_row + 1,
                            resolved_overrides[1],
                            set(resolved_overrides[1]),
                            resolved_overrides[0],
                        )
            ledger.add("warning", "legacy_mapping", "Mapped legacy workbook has no embedded CAOS authority hashes; commit must bind the calculated fingerprints.")

        _mark_formula_dispositions(audit, input_cells, comparison_cells)
        formula_inputs = [
            entry for entry in audit if entry.disposition == "input_candidate"
        ]
        for entry in formula_inputs:
            entry.disposition = "comparison_only"
            ledger.add(
                "blocking",
                "imported_formula_not_authoritative",
                "Imported formulas are comparison evidence only; bind a finite value cell.",
                sheet=entry.sheet,
                cell=entry.cell,
            )
        if formula_inputs:
            payload = None
        # Unsafe or stale formula evidence is never promoted into a draft,
        # even though its cached value remains visible in the audit ledger.
        if any(entry.blocking_codes for entry in audit):
            payload = None
        if payload is not None:
            calculation = calculate_model(payload, evaluated_at=calculation_time)
            if mode == "strict_v1":
                expected_authority_sources = set(payload.source_ids)
                for period in payload.periods:
                    expected_authority_sources.update(period.authority.source_ids)
                for instrument in payload.debt_instruments:
                    expected_authority_sources.update(instrument.sources)
                    expected_authority_sources.update(instrument.authority.source_ids)
                if audit_authority_source_ids != sorted(expected_authority_sources):
                    ledger.add(
                        "blocking",
                        "source_audit_mismatch",
                        "Sources - Audit authority-source rows do not match the canonical payload.",
                        sheet="Sources - Audit",
                    )
                for field_name, actual in (
                    ("source_fingerprint", calculation.source_fingerprint),
                    ("input_fingerprint", calculation.input_fingerprint),
                    ("calculation_hash", calculation.calculation_hash),
                ):
                    if getattr(embedded, field_name) != actual:
                        ledger.add("blocking", "fingerprint_mismatch", "Embedded workbook authority does not match recalculated Model Engine v2 output.", field=field_name)
                if "Model" in formula_wb.sheetnames:
                    _tie_out_model(
                        formula_wb,
                        cached_wb,
                        calculation,
                        payload.reporting_unit,
                        ledger,
                    )

        blocking_count = sum(issue.severity == "blocking" for issue in ledger.issues)
        warning_count = sum(issue.severity == "warning" for issue in ledger.issues)
        return ModelWorkbookPreview(
            mode=mode,
            workbook_sha256=hashlib.sha256(content).hexdigest(),
            sheet_names=list(formula_wb.sheetnames),
            mapping=parsed_mapping,
            draft_payload=payload,
            calculation=calculation,
            identity=identity,
            embedded_hashes=embedded,
            formula_audit=audit,
            ambiguities=ambiguities,
            issues=ledger.issues,
            blocking_count=blocking_count,
            warning_count=warning_count,
        )
    finally:
        formula_wb.close()
        cached_wb.close()
