"""Parse uploaded portfolio files → position / constraint / mandate dicts.

Deterministic, stdlib + openpyxl only (no LLM). Three parsers:

  * ``parse_holdings_xlsx`` — the Holdings sheet → CLO positions (rows where the
    ``Holdings`` par column > 0; ``Size ($Mn)`` is the loan facility size, kept as
    reference). Ratings via ``ratings.parse_rating_cell``.
  * ``parse_constraints_csv`` — a constraints / compliance-monitor CSV → constraint
    *definitions* (limit_value/op parsed off the limit text; current/status are
    computed later in engine/portfolio.py).
  * ``parse_mandate_csv`` — a Section/Field/Value mandate file → a nested dict.

Column detection is header-keyword based so it survives the exact attachment shapes
without hard-coding column order.
"""

from __future__ import annotations

import csv
import io
import math
import re
from typing import Any, Dict, List, Optional, Tuple

import ratings
from xlsx_safety import validate_xlsx_package

# ── Holdings ─────────────────────────────────────────────────────────────────
_HOLD_COLS = {
    "borrower_name": ("borrower name", "borrower", "company", "issuer name", "name"),
    "ticker": ("ticker",),
    "figi": ("figi",),
    "loan_name": ("loan name",),
    "sector": ("index sector", "sector"),
    "sub_sector": ("sub sector", "sub-sector", "subsector"),
    "ranking": ("ranking",),
    "ratings": ("ratings", "rating"),
    "par": ("holdings",),          # the CLO's par position ($) — the source of truth
    "facility": ("size ($mn)", "size"),
    "margin": ("margin",),
    "maturity": ("maturity",),
    "bid": ("bid",),
    "ask": ("ask",),
    "ytm": ("mid ytm", "ytm"),
    "dm": ("mid 3y dm", "dm"),
}


def _hdr_map(header: Tuple, spec: Dict[str, Tuple[str, ...]]) -> Dict[str, int]:
    """First column whose lowered header matches a key's alias list."""
    lower = [(i, str(h).strip().lower()) for i, h in enumerate(header) if h is not None]
    out: Dict[str, int] = {}
    for key, aliases in spec.items():
        for i, h in lower:
            if key in out:
                break
            # exact alias, or "sector" keyword — but 'sub sector' must not steal 'sector'
            if h in aliases or (key == "sector" and h == "index sector"):
                out[key] = i
                break
    return out


def _num(v: Any) -> Optional[float]:
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return f if math.isfinite(f) else None


def _empty_ledger() -> Dict[str, Any]:
    return {"rows_read": 0, "truncated": False, "dropped_unparseable_par": 0}


def parse_holdings_xlsx(content: bytes, max_rows: int = 20000) -> Dict[str, Any]:
    """Holdings sheet → {"positions": [...], "ledger": {...}}.

    ``positions`` is the same list this function used to return bare. ``ledger``
    is a row-accounting dict so callers can tell a complete parse from a silent
    partial one — otherwise a truncation or a batch of unparseable rows reads
    identically to "every row was a legitimate reference row" (n_positions looks
    complete either way):
      * ``rows_read`` — count of all non-header rows actually iterated, before
        any ``max_rows`` break.
      * ``truncated`` — True if the ``max_rows`` break fired (rows beyond it
        were never read at all).
      * ``dropped_unparseable_par`` — rows where the par cell held SOME value
        (non-empty, non-None) that numeric coercion couldn't turn into a finite
        float — e.g. text-formatted "2,500,000", or a formula whose cached
        value is None under ``data_only=True``. This is deliberately distinct
        from a legitimate ``par <= 0`` market-universe reference row (empty/None
        par cell, or a genuine non-positive number) — only a present-but-broken
        value counts as dropped.
    """
    validate_xlsx_package(content)
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return {"positions": [], "ledger": _empty_ledger()}
    # Prefer a sheet literally named Holdings; else the first sheet.
    ws = next((s for s in wb.worksheets if s.title.strip().lower() == "holdings"), wb.worksheets[0])
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return {"positions": [], "ledger": _empty_ledger()}
    cols = _hdr_map(header, _HOLD_COLS)
    if "par" not in cols:
        return {"positions": [], "ledger": _empty_ledger()}  # no CLO-holdings column → nothing to ingest

    # data_only=True (above) gives cached values but collapses an un-recalculated
    # formula (workbook never reopened in Excel since the formula was entered) to
    # the SAME None a genuinely blank cell produces — undercounting
    # dropped_unparseable_par. Cross-reference a data_only=False pass, Cell
    # objects only, so a None-valued par cell can be told apart from a real
    # formula that just lacks a cache. Best-effort: any failure here degrades to
    # the pre-existing (blank-cell) behavior rather than breaking the parse.
    par_formula_rows: Any = iter(())
    try:
        wb_formulas = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
        ws_formulas = next(
            (s for s in wb_formulas.worksheets if s.title.strip().lower() == "holdings"),
            wb_formulas.worksheets[0],
        )
        formula_rows = ws_formulas.iter_rows()
        next(formula_rows, None)  # skip header, mirroring the values-only pass
        par_formula_rows = formula_rows
    except Exception:
        pass

    def cell(row, key):
        i = cols.get(key)
        return row[i] if i is not None and i < len(row) else None

    par_col = cols.get("par")
    out: List[Dict[str, Any]] = []
    rows_read = 0
    truncated = False
    dropped_unparseable_par = 0
    for n, row in enumerate(rows):
        if n >= max_rows:
            truncated = True
            break
        rows_read += 1
        frow = next(par_formula_rows, None)
        raw_par = cell(row, "par")
        par = _num(raw_par)
        if not par or par <= 0:
            # A genuinely empty/None par cell is a legitimate market-universe
            # reference row, not a defect. A cell that HAD a value but failed
            # numeric coercion (text, broken formula cache, ...) is a real
            # parse failure — count it so callers can see the loss.
            par_present = not (raw_par is None or (isinstance(raw_par, str) and not raw_par.strip()))
            if not par_present and raw_par is None and frow is not None and par_col is not None:
                if par_col < len(frow) and frow[par_col].data_type == "f":
                    par_present = True  # uncached formula, not actually blank
            if par is None and par_present:
                dropped_unparseable_par += 1
            continue  # not a CLO position (market-universe reference row)
        moody, sp = ratings.parse_rating_cell(cell(row, "ratings"))
        bid, ask = _num(cell(row, "bid")), _num(cell(row, "ask"))
        # `bid if bid is not None else ask`, NOT `bid or ask`: a fully distressed
        # zero-BID quote is a real 0.0 mark, and truthiness dropped it to None
        # (dash) while the mirror zero-ask case survived (triage 2026-07-16 P3).
        price = (bid + ask) / 2 if bid is not None and ask is not None else (
            bid if bid is not None else ask)
        name = cell(row, "borrower_name")
        out.append({
            "borrower_name": str(name).strip() if name else "—",
            "ticker": _s(cell(row, "ticker")),
            "figi": _s(cell(row, "figi")),
            "loan_name": _s(cell(row, "loan_name")),
            "sector": _s(cell(row, "sector")),
            "sub_sector": _s(cell(row, "sub_sector")),
            "ranking": _s(cell(row, "ranking")),
            "rating_moody": moody,
            "rating_sp": sp,
            "par_usd": par,
            "facility_musd": _num(cell(row, "facility")),
            "margin_bps": _num(cell(row, "margin")),
            "maturity": _s(cell(row, "maturity")),
            "price": price,
            "ytm": _num(cell(row, "ytm")),
            "dm": _num(cell(row, "dm")),
        })
    return {"positions": out, "ledger": {
        "rows_read": rows_read,
        "truncated": truncated,
        "dropped_unparseable_par": dropped_unparseable_par,
    }}


def _s(v: Any) -> Optional[str]:
    return str(v).strip() if v not in (None, "") else None


# ── Constraints ──────────────────────────────────────────────────────────────
_OP = {"≤": "<=", "<=": "<=", "<": "<", "≥": ">=", ">=": ">=", ">": ">"}
_LIMIT_RE = re.compile(r"(≤|>=|<=|≥|<|>)?\s*([0-9]+(?:\.[0-9]+)?)\s*(%|x|yrs|years|notches)?", re.I)
# Word-form direction, for limits written without a glyph: "Min 90% 1st lien" is a
# FLOOR — defaulting it to a max inverted every word-form minimum, so a breached
# hard floor rendered Pass/Watch with positive headroom, and the deeper the breach
# the greener the status (triage 2026-07-16 P1).
_WORD_MIN = re.compile(r"\b(?:min(?:imum)?|at\s+least|no\s+(?:less|lower)\s+than|floor)\b", re.I)
_WORD_MAX = re.compile(r"\b(?:max(?:imum)?|at\s+most|no\s+(?:more|greater|higher)\s+than|cap(?:ped)?|up\s+to)\b", re.I)
# A cp1252/Excel-legacy save degrades the ≥/≤ glyphs to "?": the direction is
# gone, and guessing one would silently invert half of them. Un-inferable → the
# row stays "Info" (value dropped), never a guessed ceiling.
# A second mangling hits the same failure mode from a different byte: a legacy
# Symbol-font ">=" (0xB3) decoded via _read_csv's "utf-8-sig" + errors="replace"
# lands as U+FFFD (the Unicode replacement character), not a literal "?". Left
# unhandled, U+FFFD matches neither this pattern nor _WORD_MIN/_WORD_MAX, so it
# fell through to the bare-number default (op="<=") and silently inverted a
# hard floor into a ceiling — e.g. "≥ 90.0% NAV" mangled to "� 90.0% NAV"
# would flip a 99.4% first-lien book from compliant to a rendered breach.
# Treat a leading U+FFFD exactly like a leading "?": un-inferable, stays Info.
_DEGLYPHED = re.compile(r"^[?\ufffd]\s*[0-9]")

_CONS_COLS = {
    "code": ("id",),
    "category": ("constraint category", "category"),
    "parameter": ("parameter",),
    "limit_text": ("limit", "limit / guideline", "limit/guideline", "guideline"),
    "breach_type": ("breach type",),
    "source_document": ("source document", "source"),
}


def _parse_limit(text: str) -> Tuple[Optional[float], Optional[str], Optional[str]]:
    """'≤ 2.5% NAV' → (2.5, '%', '<='). 'Info only' / '—' → (None, None, None).
    Direction precedence: an explicit glyph wins; else a word form ('Min 90%' →
    '>=', 'Max 40%' / 'up to' → '<='); a de-glyphed '? 90%' (cp1252 save) or a
    replacement-character '� 90%' (legacy Symbol-font glyph mangled through
    utf-8-sig/replace decoding) is un-inferable and stays Info; only a genuinely
    bare number defaults to a max.
    Rating-floor style text with a name before the number (e.g. '≥ Caa1
    (numeric ≤ 17)') yields the first numeric it finds, but such rows aren't
    category-mapped so they stay 'Info'."""
    t = (text or "").strip()
    if not t or t.lower() in ("info only", "—", "-", "n/a"):
        return (None, None, None)
    if _DEGLYPHED.match(t):
        return (None, None, None)  # direction lost in transit — never guess one
    m = _LIMIT_RE.search(t)
    if not m:
        return (None, None, None)
    if m.group(1):
        op = _OP.get(m.group(1))
    elif _WORD_MIN.search(t):
        op = ">="
    elif _WORD_MAX.search(t):
        op = "<="
    else:
        op = "<="  # bare number → treat as a max
    value = float(m.group(2))
    unit = m.group(3)
    unit = {"yrs": "years", "years": "years"}.get((unit or "").lower(), unit) if unit else (
        "%" if "%" in t else None)
    return (value, unit, op)


def parse_constraints_csv(content: bytes) -> List[Dict[str, Any]]:
    """A constraints / compliance-monitor CSV → constraint definition dicts."""
    rows = _read_csv(content)
    if not rows:
        return []
    cols = _csv_hdr_map(rows[0], _CONS_COLS)
    if "category" not in cols or "limit_text" not in cols:
        return []
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        cat = _csv_cell(r, cols, "category")
        if not cat:
            continue
        limit_text = _csv_cell(r, cols, "limit_text")
        value, unit, op = _parse_limit(limit_text or "")
        out.append({
            "code": _csv_cell(r, cols, "code"),
            "category": cat,
            "parameter": _csv_cell(r, cols, "parameter"),
            "limit_text": limit_text,
            "limit_value": value,
            "limit_unit": unit,
            "limit_op": op,
            "breach_type": _csv_cell(r, cols, "breach_type"),
            "source_document": _csv_cell(r, cols, "source_document"),
        })
    return out


# ── Mandate ──────────────────────────────────────────────────────────────────
def parse_mandate_csv(content: bytes) -> Dict[str, Any]:
    """A Section/Field/Value(/Notes) mandate file → {section: [{field, value, notes}]}."""
    rows = _read_csv(content)
    if not rows:
        return {}
    cols = _csv_hdr_map(rows[0], {
        "section": ("section",), "field": ("field",), "value": ("value",), "notes": ("notes",)})
    if "field" not in cols or "value" not in cols:
        return {}
    out: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows[1:]:
        field = _csv_cell(r, cols, "field")
        if not field:
            continue
        section = _csv_cell(r, cols, "section") or "General"
        out.setdefault(section, []).append({
            "field": field,
            "value": _csv_cell(r, cols, "value"),
            "notes": _csv_cell(r, cols, "notes"),
        })
    return out


# ── CSV helpers ──────────────────────────────────────────────────────────────
def _read_csv(content: bytes) -> List[List[str]]:
    try:
        text = content.decode("utf-8-sig", "replace")
    except Exception:
        return []
    return [row for row in csv.reader(io.StringIO(text))]


def _csv_hdr_map(header: List[str], spec: Dict[str, Tuple[str, ...]]) -> Dict[str, int]:
    lower = [(i, str(h).strip().lower()) for i, h in enumerate(header)]
    out: Dict[str, int] = {}
    for key, aliases in spec.items():
        for i, h in lower:
            if key not in out and h in aliases:
                out[key] = i
                break
    return out


def _csv_cell(row: List[str], cols: Dict[str, int], key: str) -> Optional[str]:
    i = cols.get(key)
    if i is None or i >= len(row):
        return None
    v = str(row[i]).strip()
    return v or None


if __name__ == "__main__":
    # ponytail: one runnable self-check — parse each shape end to end.
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Holdings"
    ws.append(["Ticker", "Borrower Name", "Index Sector", "FIGI", "Ranking", "Ratings",
               "Size ($Mn)", "Margin", "Bid", "Ask", "Mid 3Y DM", "Holdings"])
    ws.append(["ACRISU", "Acrisure LLC", "Insurance", "BBG01", "1L Gtd Sr. Secd", "B2 / B",
               1500, 325, 90.5, 91.5, 517, 1_000_000])
    ws.append(["NOISE", "Market Ref Co", "Software", "BBG02", "1L Sr. Secd", "B1 / B+",
               800, 400, 99, 100, 300, None])  # no Holdings → market row, skipped
    ws.append(["BROKEN", "Bad Par Co", "Software", "BBG03", "1L Sr. Secd", "B1 / B+",
               800, 400, 99, 100, 300, "2,500,000"])  # text-formatted par → coercion fails, dropped
    buf = io.BytesIO()
    wb.save(buf)
    result = parse_holdings_xlsx(buf.getvalue())
    pos, ledger = result["positions"], result["ledger"]
    assert len(pos) == 1, pos
    p = pos[0]
    assert p["par_usd"] == 1_000_000 and p["facility_musd"] == 1500
    assert (p["rating_moody"], p["rating_sp"]) == ("B2", "B")
    assert p["price"] == 91.0 and p["sector"] == "Insurance" and p["ranking"].startswith("1L")
    assert ledger == {"rows_read": 3, "truncated": False, "dropped_unparseable_par": 1}, ledger

    cons_csv = (
        "ID,Constraint Category,Parameter,Limit,Breach Type,Source Document,Current Value,Status\r\n"
        "C-01,Single Name,Max single issuer,≤ 2.5% NAV,Hard,Indenture §7.11,2.37%,Watch\r\n"
        "C-09,Instrument,Min 1st Lien,≥ 90.0% NAV,Hard,Indenture §7.11,99.39%,Pass\r\n"
        "C-16,Covenant,Cov-Lite %,Info only,Info,,~100%,Pass\r\n"
    ).encode()
    cons = {c["code"]: c for c in parse_constraints_csv(cons_csv)}
    assert cons["C-01"]["limit_value"] == 2.5 and cons["C-01"]["limit_op"] == "<="
    assert cons["C-09"]["limit_value"] == 90.0 and cons["C-09"]["limit_op"] == ">="
    assert cons["C-16"]["limit_value"] is None  # "Info only" → not a numeric limit
    assert cons["C-01"]["source_document"] == "Indenture §7.11"

    mand_csv = (
        "Section,Field,Value,Notes\r\n"
        "Strategy,Legal Final Maturity,15-Mar-2038,\r\n"
        "Key Parties,Trustee,Bank of New York Mellon,\r\n"
    ).encode()
    m = parse_mandate_csv(mand_csv)
    assert m["Strategy"][0]["field"] == "Legal Final Maturity"
    assert m["Key Parties"][0]["value"] == "Bank of New York Mellon"
    print("portfolio_ingest.py self-check OK")
