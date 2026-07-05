"""Agency-rating scale + deterministic extraction from structured ingest.

Two responsibilities, one home so the Moody's scale has a single source of truth:

  * the rating scale (Moody's senior→junior, S&P/Fitch index-aligned, idealized
    rating factors) + parse/index/bucket helpers — shared by routes/digest.py
    (WARF / CCC watch) and engine/querygraph.py (the rating-distribution walk).
  * ``extract_ratings_from_workbook`` — read the Ratings column an uploaded
    holdings / market-data / pricing sheet already carries (e.g. "B2 / B" =
    Moody's / S&P) and return per-row {figi, ticker, name, moody, sp}.

No LLM, no paid feed: ratings are collected off the columns the analyst already
uploads (migration 0014 made them analyst-typed; this replaces the typing).
routes/ingestion.py calls the extractor on xlsx upload and writes matches onto
existing issuers.
"""

from __future__ import annotations

import io
from typing import Dict, List, Optional, Tuple

# Moody's scale senior→junior; the S&P/Fitch scale is index-aligned so a missing
# Moody's rating translates positionally. Idealized rating factors drive WARF.
MOODY = ("aaa", "aa1", "aa2", "aa3", "a1", "a2", "a3", "baa1", "baa2", "baa3",
         "ba1", "ba2", "ba3", "b1", "b2", "b3", "caa1", "caa2", "caa3", "ca", "c")
SP = ("aaa", "aa+", "aa", "aa-", "a+", "a", "a-", "bbb+", "bbb", "bbb-",
      "bb+", "bb", "bb-", "b+", "b", "b-", "ccc+", "ccc", "ccc-", "cc", "c")
FACTORS = (1, 10, 20, 40, 70, 120, 180, 260, 360, 610,
           940, 1350, 1766, 2220, 2720, 3490, 4770, 6500, 8070, 10000, 10000)
MOODY_IDX = {r: i for i, r in enumerate(MOODY)}
SP_IDX = {r: i for i, r in enumerate(SP)}
B3_IDX = MOODY.index("b3")  # CCC-cliff watch: B3/B- and below

# Header keywords that name the issuer/borrower column across sheet shapes
# (Holdings: "Borrower Name"; Market Data: "Company").
_NAME_HEADERS = ("borrower name", "borrower", "company", "issuer name", "issuer", "name")


def _first_token(raw: object) -> Optional[str]:
    """First whitespace-delimited token with a trailing '*' watch marker stripped,
    original case. '(negative)' outlooks and 'BB- *' watch flags never fail the
    parse. None for blank cells."""
    s = str(raw).strip() if raw is not None else ""
    if not s:
        return None
    tok = s.split()[0].rstrip("*").strip()
    return tok or None


def rating_index(moody: Optional[str] = None, sp: Optional[str] = None,
                 fitch: Optional[str] = None) -> Optional[int]:
    """Scale index for the best available rating (Moody's preferred, then S&P,
    then Fitch). None when nothing parses onto a known scale."""
    for raw, idx in ((moody, MOODY_IDX), (sp, SP_IDX), (fitch, SP_IDX)):
        tok = _first_token(raw)
        if tok and tok.lower() in idx:
            return idx[tok.lower()]
    return None


def rating_bucket(idx: Optional[int]) -> str:
    """Coarse bucket for the cross-issuer distribution, mirroring the CP-6A
    exposure report's rating table: IG / BB / B / CCC / Unrated."""
    if idx is None:
        return "Unrated"
    if idx <= MOODY_IDX["baa3"]:
        return "IG"
    if idx <= MOODY_IDX["ba3"]:
        return "BB"
    if idx <= MOODY_IDX["b3"]:
        return "B"
    return "CCC"


def parse_rating_cell(cell: object) -> Tuple[Optional[str], Optional[str]]:
    """A 'Ratings' cell → (moody, sp). Handles 'B2 / B' (Moody's / S&P), a lone
    value, extra whitespace, and '*'/outlook annotations. Only tokens that land
    on the respective scale are returned — a mis-detected column yields
    (None, None) rather than polluting the rating columns. Either side may be
    None."""
    if cell is None:
        return (None, None)
    parts = str(cell).split("/")
    left = _first_token(parts[0]) if parts else None
    right = _first_token(parts[1]) if len(parts) > 1 else None
    moody = left if (left and left.lower() in MOODY_IDX) else None
    sp = right if (right and right.lower() in SP_IDX) else None
    # A lone S&P-style grade in the first slot ("BB+") with no Moody's counterpart.
    if moody is None and sp is None and left and left.lower() in SP_IDX:
        sp = left
    return (moody, sp)


def _column_map(header: Tuple) -> Dict[str, int]:
    """Header row → {rating,figi,ticker,name: col_index}. Keyword match, first
    hit wins, so it survives both the Holdings and Market Data header shapes."""
    cols: Dict[str, int] = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        t = str(h).strip().lower()
        if not t:
            continue
        if "rating" in t:
            cols.setdefault("rating", i)
        elif t == "figi":
            cols.setdefault("figi", i)
        elif t == "ticker":
            cols.setdefault("ticker", i)
        elif t in _NAME_HEADERS:
            cols.setdefault("name", i)
    return cols


def _cell(row: Tuple, cols: Dict[str, int], key: str) -> Optional[str]:
    i = cols.get(key)
    if i is None or i >= len(row):
        return None
    v = row[i]
    return str(v).strip() if v not in (None, "") else None


def extract_ratings_from_workbook(content: bytes, max_rows: int = 20000) -> List[dict]:
    """Every sheet with a detectable Ratings column → [{figi,ticker,name,moody,sp}]
    for each row carrying a parseable rating. Returns [] on any parse failure so a
    malformed workbook never breaks the upload. ``max_rows`` bounds a single sheet
    (the upload is already size-capped upstream)."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return []
    out: List[dict] = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            continue
        cols = _column_map(header)
        if "rating" not in cols:
            continue
        ri = cols["rating"]
        for n, row in enumerate(rows):
            if n >= max_rows:
                break
            cell = row[ri] if ri < len(row) else None
            moody, sp = parse_rating_cell(cell)
            if not moody and not sp:
                continue
            out.append({
                "figi": _cell(row, cols, "figi"),
                "ticker": _cell(row, cols, "ticker"),
                "name": _cell(row, cols, "name"),
                "moody": moody,
                "sp": sp,
            })
    return out


if __name__ == "__main__":
    # ponytail: one runnable self-check — parse, index alignment, bucketing, and
    # end-to-end workbook extraction (build an xlsx in memory, read it back).
    assert parse_rating_cell("B2 / B") == ("B2", "B")
    assert parse_rating_cell("Ba1 / BB+") == ("Ba1", "BB+")
    assert parse_rating_cell("B1 / BB- *") == ("B1", "BB-")   # watch marker dropped
    assert parse_rating_cell("Caa1") == ("Caa1", None)
    assert parse_rating_cell("B+") == (None, "B+")            # lone S&P grade
    assert parse_rating_cell("not-a-rating") == (None, None)
    assert parse_rating_cell(None) == (None, None)

    assert rating_index(moody="B2") == rating_index(sp="B")  # scales index-aligned
    assert rating_index(moody="B2 (negative)") == rating_index(moody="B2")
    assert rating_index() is None
    assert rating_bucket(rating_index(moody="Caa1")) == "CCC"
    assert rating_bucket(rating_index(sp="BBB")) == "IG"
    assert rating_bucket(rating_index(moody="Ba2")) == "BB"
    assert rating_bucket(None) == "Unrated"

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Ticker", "Borrower Name", "FIGI", "Ratings"])
    ws.append(["ACRISU", "Acrisure LLC", "BBG01B6UZZ33", "B2 / B"])
    ws.append(["NOISE", "No Rating Co", "BBG000", None])  # skipped: no rating
    buf = io.BytesIO()
    wb.save(buf)
    recs = extract_ratings_from_workbook(buf.getvalue())
    assert len(recs) == 1, recs
    assert recs[0] == {"figi": "BBG01B6UZZ33", "ticker": "ACRISU",
                       "name": "Acrisure LLC", "moody": "B2", "sp": "B"}, recs[0]
    print("ratings.py self-check OK")
