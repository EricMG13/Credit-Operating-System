"""Binary renderers for immutable Report Studio versions.

The renderers accept only the frozen ``ReportVersion`` payload/authority. They do
not query mutable domain state, recalculate credit metrics, or interpret prose.
"""

from __future__ import annotations

import asyncio
import html
import json
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Iterable, Literal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle

from config import get_settings


_INK = "1D2433"
_NAVY = "172238"
_BLUE = "2F64B7"
_MUTED = "687386"
_CREAM = "F7F5EE"
_MAX_RENDERED_OVERRIDES = 500
_export_sem: "asyncio.Semaphore | None" = None


def _export_semaphore() -> asyncio.Semaphore:
    global _export_sem
    if _export_sem is None:
        _export_sem = asyncio.Semaphore(
            max(1, get_settings().caos_report_export_concurrency)
        )
    return _export_sem


def _display(value: Any, *, limit: int = 2_000) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        text = str(value)
    else:
        text = json.dumps(value, sort_keys=True, default=str, ensure_ascii=False)
    return text if len(text) <= limit else f"{text[:limit - 3]}..."


def _xlsx_text(value: Any, *, limit: int = 2_000) -> str:
    """Keep analyst/source text inert when a workbook is opened in Excel."""
    text = _display(value, limit=limit)
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def _xlsx_scalar(value: Any) -> Any:
    if isinstance(value, (int, float, bool)):
        return value
    return _xlsx_text(value)


def _model_reporting_metadata(model: dict) -> tuple[str, str, str]:
    """Read currency/scale only from the immutable model input snapshot.

    Legacy checkpoints may not carry these fields.  Those exports must remain
    explicit about the gap rather than silently inheriting a renderer default.
    """

    payload = model.get("payload") if isinstance(model.get("payload"), dict) else {}
    raw_currency = payload.get("reporting_currency")
    raw_unit = payload.get("reporting_unit")
    currency = (
        raw_currency.strip().upper()
        if isinstance(raw_currency, str) and raw_currency.strip()
        else "Unavailable"
    )
    unit = (
        raw_unit.strip()
        if isinstance(raw_unit, str) and raw_unit.strip()
        else "Unavailable"
    )
    return currency, unit, f"{currency} {unit}"


def _instrument_currencies(model: dict) -> dict[str, str]:
    payload = model.get("payload") if isinstance(model.get("payload"), dict) else {}
    instruments = (
        payload.get("debt_instruments")
        if isinstance(payload.get("debt_instruments"), list)
        else []
    )
    currencies: dict[str, str] = {}
    for instrument in instruments:
        if not isinstance(instrument, dict):
            continue
        instrument_id = instrument.get("instrument_id")
        currency = instrument.get("currency")
        if (
            isinstance(instrument_id, str)
            and instrument_id
            and isinstance(currency, str)
            and currency.strip()
        ):
            currencies[instrument_id] = currency.strip().upper()
    return currencies


def _aware_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, str) and value.strip():
        try:
            parsed = datetime.fromisoformat(value.strip().replace("Z", "+00:00"))
        except ValueError:
            return None
    else:
        return None
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        return None
    return parsed.astimezone(timezone.utc)


def _model_override_rows(
    model: dict,
    *,
    report_event_at: Any,
) -> tuple[list[dict[str, Any]], int]:
    payload = model.get("payload") if isinstance(model.get("payload"), dict) else {}
    overrides = (
        payload.get("overrides") if isinstance(payload.get("overrides"), list) else []
    )
    calculation = (
        model.get("calculation") if isinstance(model.get("calculation"), dict) else {}
    )
    nodes: dict[str, dict] = {}
    for period in calculation.get("periods") or []:
        if not isinstance(period, dict):
            continue
        for node in period.get("nodes") or []:
            if isinstance(node, dict) and isinstance(node.get("node_id"), str):
                nodes[node["node_id"]] = node
    event_at = _aware_datetime(report_event_at)
    rows: list[dict[str, Any]] = []
    valid_overrides = [item for item in overrides if isinstance(item, dict)]
    valid_overrides.sort(key=lambda item: _display(item.get("node_id")))
    for override in valid_overrides[:_MAX_RENDERED_OVERRIDES]:
        expiry = _aware_datetime(override.get("expires_at"))
        if override.get("expires_at") is None:
            override_status = "ACTIVE AT REPORT EVENT"
        elif expiry is None or event_at is None:
            override_status = "STATUS UNKNOWN"
        elif expiry > event_at:
            override_status = "ACTIVE AT REPORT EVENT"
        else:
            override_status = "INACTIVE AT REPORT EVENT"
        node_id = _display(override.get("node_id"))
        displaced = nodes.get(node_id, {})
        value = (
            "NULL"
            if override.get("value_type") == "null"
            else override.get("value")
        )
        original_formula = displaced.get("formula")
        if not isinstance(original_formula, str) or not original_formula.strip():
            original_formula = (
                "No formula (input)" if displaced else "Unavailable"
            )
        original_value = (
            "NULL"
            if displaced and displaced.get("original_value") is None
            else displaced.get("original_value", "Unavailable")
        )
        rows.append({
            "status": override_status,
            "node_id": node_id,
            "value": value,
            "reason": override.get("reason"),
            "scope": override.get("scope"),
            "source": override.get("source"),
            "expires_at": override.get("expires_at"),
            "original_formula": original_formula,
            "original_value": original_value,
        })
    return rows, max(0, len(valid_overrides) - len(rows))


def _rows(value: Any, prefix: str = "") -> Iterable[tuple[str, Any]]:
    if isinstance(value, dict):
        for key in sorted(value):
            path = f"{prefix}.{key}" if prefix else str(key)
            yield from _rows(value[key], path)
    elif isinstance(value, list):
        if not value:
            yield prefix, ""
        elif all(not isinstance(item, (dict, list)) for item in value):
            yield prefix, ", ".join(_display(item, limit=200) for item in value)
        else:
            for index, item in enumerate(value):
                yield from _rows(item, f"{prefix}[{index}]")
    else:
        yield prefix, value


def _safe_sheet_name(module_id: str, used: set[str]) -> str:
    base = "".join(char if char not in "[]:*?/\\" else "-" for char in module_id).strip()[:31] or "Module"
    name = base
    suffix = 2
    while name in used:
        tail = f"-{suffix}"
        name = f"{base[:31 - len(tail)]}{tail}"
        suffix += 1
    used.add(name)
    return name


def _reviewed_report(payload: dict) -> dict:
    composition = payload.get("composition")
    if not isinstance(composition, dict):
        return {}
    reviewed = composition.get("reviewed_report")
    return reviewed if isinstance(reviewed, dict) else {}


def _reviewed_rows(report: dict) -> Iterable[tuple[str, str, Any]]:
    """Flatten the materialized Report Studio DSL without replaying overlays."""
    yield "Document", "Title", report.get("title")
    yield "Document", "Subtitle", report.get("subtitle")
    sections = report.get("sections") if isinstance(report.get("sections"), list) else []
    for index, section in enumerate(sections):
        if not isinstance(section, dict):
            continue
        label = _display(section.get("title")) or f"Section {index + 1}"
        kind = section.get("t")
        if kind == "profile":
            for row in section.get("rows") or []:
                if isinstance(row, list) and len(row) >= 2:
                    yield label, _display(row[0]), row[1]
        elif kind == "table":
            columns = section.get("cols") if isinstance(section.get("cols"), list) else []
            yield label, "Columns", " | ".join(_display(value) for value in columns)
            for row_index, row in enumerate(section.get("rows") or []):
                if isinstance(row, dict) and isinstance(row.get("cells"), list):
                    yield label, f"Row {row_index + 1}", " | ".join(
                        _display(value) for value in row["cells"]
                    )
        elif kind == "list":
            for item_index, item in enumerate(section.get("items") or []):
                yield label, f"Item {item_index + 1}", item
        else:
            for field in ("sub", "subhead", "body", "label", "labelBody", "note"):
                if field in section:
                    yield label, field, section.get(field)


def _style_xlsx_header(row) -> None:
    for cell in row:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.alignment = Alignment(wrap_text=True)


def render_report_xlsx(*, version_id: str, document_sha256: str, payload: dict, authority: dict) -> bytes:
    document = payload.get("document") if isinstance(payload.get("document"), dict) else {}
    reviewed = _reviewed_report(payload)
    modules = (
        [] if reviewed
        else document.get("sections") if isinstance(document.get("sections"), list) else []
    )
    model = payload.get("model") if isinstance(payload.get("model"), dict) else {}
    calculation = model.get("calculation") if isinstance(model.get("calculation"), dict) else {}
    reporting_currency, reporting_unit, reporting_scale = _model_reporting_metadata(
        model
    )
    instrument_currencies = _instrument_currencies(model)
    override_rows, omitted_override_count = _model_override_rows(
        model,
        report_event_at=authority.get("as_of"),
    )
    freshness = authority.get("freshness_evaluation") if isinstance(authority.get("freshness_evaluation"), dict) else {}
    freshness_state = _display(freshness.get("state") or authority.get("freshness") or "unknown").upper()
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover.sheet_view.showGridLines = False
    cover.freeze_panes = "A8"
    cover["A1"] = "CAOS - IMMUTABLE COMMITTEE REPORT"
    cover["A1"].font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
    cover["A1"].fill = PatternFill("solid", fgColor=_NAVY)
    cover.merge_cells("A1:D2")
    cover["A1"].alignment = Alignment(vertical="center")
    metadata = [
        ("Report version", version_id),
        ("Document SHA-256", document_sha256),
        ("Issuer", document.get("issuer_id")),
        ("Run", document.get("run_id")),
        ("As of", document.get("as_of_date")),
        ("QA status", document.get("qa_status")),
        ("Committee status", document.get("committee_status")),
        ("Prepared by", document.get("prepared_by")),
        ("Authority origin", authority.get("origin")),
        ("Authority state", authority.get("approval_state")),
        ("Freshness", freshness_state),
        ("Freshness source", freshness.get("source_kind")),
        ("Freshness reason", freshness.get("reason")),
        ("Freshness policy", freshness.get("policy_version")),
        ("Freshness observed", freshness.get("observed_at")),
        ("Freshness effective period", freshness.get("effective_period_end")),
        ("Freshness due", freshness.get("due_at")),
        ("Model engine", model.get("engine_version")),
        ("Model source fingerprint", model.get("source_fingerprint")),
        ("Model input fingerprint", model.get("input_fingerprint")),
        ("Model calculation hash", model.get("calculation_hash")),
        ("Model draft revision", model.get("draft_revision")),
        ("Model origin", (model.get("authority") or {}).get("origin") if isinstance(model.get("authority"), dict) else None),
        ("Model input origins", ", ".join((model.get("authority") or {}).get("model_input_origins") or []) if isinstance(model.get("authority"), dict) else None),
        ("Model analyst override", (model.get("authority") or {}).get("analyst_override") if isinstance(model.get("authority"), dict) else None),
        ("Model availability", calculation.get("status")),
        ("Model reporting currency", reporting_currency),
        ("Model reporting unit", reporting_unit),
        ("Model override count", len(override_rows) + omitted_override_count),
    ]
    for row_index, (label, value) in enumerate(metadata, start=4):
        cover.cell(row_index, 1, label).font = Font(name="Arial", bold=True, color=_MUTED)
        cover.cell(row_index, 2, _xlsx_text(value))
    cover.column_dimensions["A"].width = 24
    cover.column_dimensions["B"].width = 72
    cover.column_dimensions["C"].width = 18
    cover.column_dimensions["D"].width = 18
    cover.sheet_properties.pageSetUpPr.fitToPage = True
    cover.page_setup.fitToWidth = 1
    cover.page_setup.fitToHeight = 0

    if reviewed:
        reviewed_sheet = wb.create_sheet("Reviewed Report")
        reviewed_sheet.sheet_view.showGridLines = False
        reviewed_sheet.freeze_panes = "A2"
        reviewed_sheet.append(["Section", "Field", "Reviewed value"])
        _style_xlsx_header(reviewed_sheet[1])
        for section, field, value in _reviewed_rows(reviewed):
            reviewed_sheet.append([
                _xlsx_text(section), _xlsx_text(field), _xlsx_scalar(value),
            ])
        reviewed_sheet.column_dimensions["A"].width = 42
        reviewed_sheet.column_dimensions["B"].width = 38
        reviewed_sheet.column_dimensions["C"].width = 110
        for row in reviewed_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(name="Arial", size=9, color=_INK)

    summary = wb.create_sheet("Module Summary")
    summary.sheet_view.showGridLines = False
    summary.freeze_panes = "A2"
    summary.append(["Module", "Name", "Confidence", "QA status"])
    for module in modules:
        if not isinstance(module, dict):
            continue
        summary.append([
            _xlsx_text(module.get("module_id")),
            _xlsx_text(module.get("module_name")),
            _xlsx_scalar(module.get("confidence")),
            _xlsx_text(module.get("qa_status")),
        ])
    for cell in summary[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_NAVY)
    summary.column_dimensions["A"].width = 16
    summary.column_dimensions["B"].width = 42
    summary.column_dimensions["C"].width = 16
    summary.column_dimensions["D"].width = 18
    if summary.max_row > 1:
        table = Table(displayName="ModuleSummary", ref=f"A1:D{summary.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        summary.add_table(table)

    used = {sheet.title for sheet in wb.worksheets}
    thin = Side(style="thin", color="D8DCE5")
    for module_index, module in enumerate(modules, start=1):
        if not isinstance(module, dict):
            continue
        module_id = _display(module.get("module_id")) or f"Module {module_index}"
        sheet = wb.create_sheet(_safe_sheet_name(module_id, used))
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A5"
        sheet["A1"] = _xlsx_text(f"{module_id} - {_display(module.get('module_name'))}")
        sheet["A1"].font = Font(name="Arial", size=15, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor=_NAVY)
        sheet["A1"].alignment = Alignment(vertical="center", indent=1)
        sheet.merge_cells("A1:B2")
        sheet["A3"] = "Path"
        sheet["B3"] = "Frozen value"
        for cell in sheet[3]:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=_BLUE)
        data_rows = list(_rows(module.get("summary") or {})) or [("summary", "")]
        for path, value in data_rows:
            sheet.append([_xlsx_text(path), _xlsx_scalar(value)])
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(name="Arial", size=9, color=_INK)
        sheet.column_dimensions["A"].width = 46
        sheet.column_dimensions["B"].width = 90
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

    model_periods = calculation.get("periods") if isinstance(calculation.get("periods"), list) else []
    if model_periods:
        model_sheet = wb.create_sheet(_safe_sheet_name("Model", used))
        model_sheet.sheet_view.showGridLines = False
        model_sheet.freeze_panes = "A2"
        model_headers = [
            "Period key", "Label", "Kind", f"Revenue ({reporting_scale})",
            f"Adjusted EBITDA ({reporting_scale})",
            f"Cash interest ({reporting_scale})",
            f"Total debt ({reporting_scale})", f"Cash ({reporting_scale})",
            f"Net debt ({reporting_scale})", "Gross leverage",
            "Net leverage", "Interest coverage",
            f"Free cash flow ({reporting_scale})",
        ]
        model_sheet.append(model_headers)
        for period in model_periods:
            if not isinstance(period, dict):
                continue
            model_sheet.append([
                _xlsx_scalar(period.get("period_key")),
                _xlsx_scalar(period.get("label")),
                _xlsx_scalar(period.get("kind")),
                period.get("revenue"), period.get("adjusted_ebitda"),
                period.get("cash_interest"), period.get("total_debt"), period.get("cash"),
                period.get("net_debt"), period.get("gross_leverage"),
                period.get("net_leverage"), period.get("interest_coverage"),
                period.get("free_cash_flow"),
            ])
        for cell in model_sheet[1]:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=_NAVY)
            cell.alignment = Alignment(wrap_text=True)
        for column in range(1, len(model_headers) + 1):
            model_sheet.column_dimensions[model_sheet.cell(1, column).column_letter].width = (
                18 if column <= 3 else 16
            )
        for row in model_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.font = Font(name="Arial", size=9, color=_INK)
            for cell in row[3:]:
                cell.number_format = "#,##0.00;[Red](#,##0.00);-"

    if override_rows or omitted_override_count:
        override_sheet = wb.create_sheet(_safe_sheet_name("Model Overrides", used))
        override_sheet.sheet_view.showGridLines = False
        override_sheet.freeze_panes = "A2"
        override_headers = [
            "Status at report event", "Node", "Override value", "Reason", "Scope",
            "Source", "Expires at", "Displaced formula", "Displaced value",
        ]
        override_sheet.append(override_headers)
        _style_xlsx_header(override_sheet[1])
        for row in override_rows:
            override_sheet.append([
                _xlsx_text(row["status"]),
                _xlsx_text(row["node_id"]),
                _xlsx_scalar(row["value"]),
                _xlsx_text(row["reason"]),
                _xlsx_text(row["scope"]),
                _xlsx_text(row["source"]),
                _xlsx_text(row["expires_at"]),
                _xlsx_text(row["original_formula"]),
                _xlsx_scalar(row["original_value"]),
            ])
        if omitted_override_count:
            override_sheet.append([
                "TRUNCATED",
                f"{omitted_override_count} additional overrides remain in the frozen payload",
            ])
        for row in override_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(name="Arial", size=9, color=_INK)
        for column, width in enumerate([24, 42, 18, 48, 16, 32, 26, 52, 18], start=1):
            override_sheet.column_dimensions[
                override_sheet.cell(1, column).column_letter
            ].width = width

    if model_periods:
        debt_sheet = wb.create_sheet(_safe_sheet_name("Debt Schedule", used))
        debt_sheet.sheet_view.showGridLines = False
        debt_sheet.freeze_panes = "A2"
        debt_headers = [
            "Period key", "Instrument ID", "Instrument currency",
            f"Opening ({reporting_unit})", f"Closing ({reporting_unit})",
            f"Average ({reporting_unit})",
            f"Benchmark interest ({reporting_unit})",
            f"Margin interest ({reporting_unit})",
            f"Coupon interest ({reporting_unit})", f"Fees ({reporting_unit})",
            f"PIK interest ({reporting_unit})", f"Hedge effect ({reporting_unit})",
            f"FX effect ({reporting_unit})", f"Cash interest ({reporting_unit})",
            f"Debt in reporting currency ({reporting_scale})",
            f"Roll-forward residual ({reporting_unit})",
        ]
        debt_sheet.append(debt_headers)
        for period in model_periods:
            if not isinstance(period, dict):
                continue
            instruments = period.get("instruments") if isinstance(period.get("instruments"), list) else []
            for instrument in instruments:
                if not isinstance(instrument, dict):
                    continue
                debt_sheet.append([
                    _xlsx_scalar(period.get("period_key")),
                    _xlsx_scalar(instrument.get("instrument_id")),
                    _xlsx_text(instrument_currencies.get(
                        str(instrument.get("instrument_id")), "Unavailable"
                    )),
                    instrument.get("opening_balance"), instrument.get("closing_balance"),
                    instrument.get("average_balance"), instrument.get("benchmark_interest"),
                    instrument.get("margin_interest"), instrument.get("coupon_interest"),
                    instrument.get("fees"), instrument.get("pik_interest"),
                    instrument.get("hedge_effect"), instrument.get("fx_effect"),
                    instrument.get("cash_interest"), instrument.get("debt_reporting_currency"),
                    instrument.get("rollforward_residual"),
                ])
        for cell in debt_sheet[1]:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=_NAVY)
            cell.alignment = Alignment(wrap_text=True)
        for column in range(1, len(debt_headers) + 1):
            debt_sheet.column_dimensions[debt_sheet.cell(1, column).column_letter].width = (
                20 if column <= 2 else 16
            )
        for row in debt_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.font = Font(name="Arial", size=9, color=_INK)
            for cell in row[3:]:
                cell.number_format = "#,##0.00;[Red](#,##0.00);-"

    if model:
        ledger = wb.create_sheet(_safe_sheet_name("Model Gaps - Warnings", used))
        ledger.append(["Type", "Detail"])
        _style_xlsx_header(ledger[1])
        ledger.append(["Availability", _xlsx_text(calculation.get("status") or "unknown")])
        for item in calculation.get("gaps") or []:
            ledger.append(["Gap", _xlsx_text(item)])
        for item in calculation.get("warnings") or []:
            ledger.append(["Warning", _xlsx_text(item)])
        ledger.column_dimensions["A"].width = 18
        ledger.column_dimensions["B"].width = 110

    sources = wb.create_sheet("Sources - Audit")
    sources.sheet_view.showGridLines = False
    sources.freeze_panes = "A2"
    sources.append(["Source ID", "Authority origin", "As of", "Approval state", "Freshness", "Freshness policy"])
    source_ids = authority.get("source_ids") if isinstance(authority.get("source_ids"), list) else []
    for source_id in source_ids:
        sources.append([
            _xlsx_text(source_id),
            _xlsx_text(authority.get("origin")),
            _xlsx_text(authority.get("as_of")),
            _xlsx_text(authority.get("approval_state")),
            _xlsx_text(freshness_state),
            _xlsx_text(freshness.get("policy_version")),
        ])
    if model:
        sources.append([
            _xlsx_text(f"model:{_display(model.get('calculation_hash'))}"),
            "model-engine-v2",
            _xlsx_text(authority.get("as_of")),
            _xlsx_text(authority.get("approval_state")),
            _xlsx_text(freshness_state),
            _xlsx_text(model.get("engine_version")),
        ])
    for cell in sources[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_NAVY)
    for column, width in zip("ABCDEF", [54, 18, 28, 18, 14, 24]):
        sources.column_dimensions[column].width = width

    wb.properties.title = "CAOS Immutable Committee Report"
    wb.properties.subject = document_sha256
    wb.properties.creator = "CAOS Report Studio"
    wb.properties.description = f"Immutable report version {version_id}"
    output = BytesIO()
    wb.save(output)
    # Production-side structural sanity check: fail before returning malformed bytes.
    reopened = load_workbook(BytesIO(output.getvalue()), read_only=True, data_only=False)
    if "Cover" not in reopened.sheetnames or "Module Summary" not in reopened.sheetnames:
        raise ValueError("Generated workbook failed structural verification.")
    if any(
        cell.data_type == "f"
        for sheet in reopened.worksheets
        for row in sheet.iter_rows()
        for cell in row
    ):
        raise ValueError("Generated report workbook contains executable formulas.")
    reopened.close()
    return output.getvalue()


def render_report_pdf(*, version_id: str, document_sha256: str, payload: dict, authority: dict) -> bytes:
    document = payload.get("document") if isinstance(payload.get("document"), dict) else {}
    modules = document.get("sections") if isinstance(document.get("sections"), list) else []
    reviewed = _reviewed_report(payload)
    model = payload.get("model") if isinstance(payload.get("model"), dict) else {}
    calculation = model.get("calculation") if isinstance(model.get("calculation"), dict) else {}
    reporting_currency, reporting_unit, reporting_scale = _model_reporting_metadata(
        model
    )
    instrument_currencies = _instrument_currencies(model)
    override_rows, omitted_override_count = _model_override_rows(
        model,
        report_event_at=authority.get("as_of"),
    )
    freshness = authority.get("freshness_evaluation") if isinstance(authority.get("freshness_evaluation"), dict) else {}
    freshness_state = _display(freshness.get("state") or authority.get("freshness") or "unknown").upper()
    output = BytesIO()
    styles = getSampleStyleSheet()
    title = ParagraphStyle("CAOSTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=19, leading=23, textColor=colors.HexColor(f"#{_NAVY}"), alignment=TA_CENTER, spaceAfter=14)
    heading = ParagraphStyle("CAOSHeading", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=12, leading=15, textColor=colors.HexColor(f"#{_NAVY}"), spaceBefore=8, spaceAfter=6)
    body = ParagraphStyle("CAOSBody", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.5, leading=11, textColor=colors.HexColor(f"#{_INK}"))
    meta = ParagraphStyle("CAOSMeta", parent=body, fontName="Courier", fontSize=7.5, leading=10, textColor=colors.HexColor(f"#{_MUTED}"))

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#D8DCE5"))
        canvas.line(0.65 * inch, 0.52 * inch, 7.85 * inch, 0.52 * inch)
        canvas.setFont("Courier", 7)
        canvas.setFillColor(colors.HexColor(f"#{_MUTED}"))
        canvas.drawString(0.65 * inch, 0.35 * inch, f"CAOS | VERSION {version_id} | SHA {document_sha256[:16]}")
        canvas.drawRightString(7.85 * inch, 0.35 * inch, f"PAGE {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=0.65 * inch, leftMargin=0.65 * inch, topMargin=0.65 * inch, bottomMargin=0.7 * inch, title="CAOS Immutable Committee Report", author="CAOS Report Studio")
    story = [Paragraph("CAOS IMMUTABLE COMMITTEE REPORT", title)]
    metadata = [
        ["Report version", version_id], ["Document SHA-256", document_sha256],
        ["Issuer", _display(document.get("issuer_id"))], ["Run", _display(document.get("run_id"))],
        ["As of", _display(document.get("as_of_date"))], ["QA status", _display(document.get("qa_status"))],
        ["Committee status", _display(document.get("committee_status"))], ["Prepared by", _display(document.get("prepared_by"))],
        ["Authority", f"{_display(authority.get('origin'))} | {_display(authority.get('approval_state'))} | {_display(authority.get('as_of'))}"],
        ["Freshness", f"{freshness_state} | {_display(freshness.get('source_kind'))} | {_display(freshness.get('reason'))}"],
        ["Freshness policy", _display(freshness.get("policy_version"))],
        ["Model engine", _display(model.get("engine_version"))],
        ["Model calculation hash", _display(model.get("calculation_hash"))],
        ["Model availability", _display(calculation.get("status"))],
        ["Model reporting currency", reporting_currency],
        ["Model reporting unit", reporting_unit],
        ["Model override count", _display(len(override_rows) + omitted_override_count)],
        ["Model origin", _display((model.get("authority") or {}).get("origin") if isinstance(model.get("authority"), dict) else None)],
        ["Model input origins", _display((model.get("authority") or {}).get("model_input_origins") if isinstance(model.get("authority"), dict) else None)],
        ["Model analyst override", _display((model.get("authority") or {}).get("analyst_override") if isinstance(model.get("authority"), dict) else None)],
    ]
    meta_table = PdfTable([[Paragraph(f"<b>{html.escape(label)}</b>", body), Paragraph(html.escape(value), meta)] for label, value in metadata], colWidths=[1.45 * inch, 5.45 * inch])
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(f"#{_CREAM}")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D8DCE5")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E7E9EE")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.extend([meta_table, Spacer(1, 0.18 * inch)])

    def detail_table(rows: list[list], widths: list[float]) -> PdfTable:
        table = PdfTable(rows, colWidths=widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_BLUE}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8DCE5")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F8FA")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        return table

    if reviewed:
        story.append(Paragraph("REVIEWED COMPOSITION", heading))
        reviewed_rows = [[Paragraph("Section", body), Paragraph("Field", body), Paragraph("Reviewed value", body)]]
        for section, field, value in _reviewed_rows(reviewed):
            reviewed_rows.append([
                Paragraph(html.escape(section), meta),
                Paragraph(html.escape(field), meta),
                Paragraph(html.escape(_display(value)), body),
            ])
        story.append(detail_table(reviewed_rows, [1.75 * inch, 1.65 * inch, 3.5 * inch]))
    else:
        story.append(Paragraph("FROZEN MODULE REGISTER", heading))
        register_rows = [["Module", "Name", "Confidence", "QA"]]
        for module in modules:
            if isinstance(module, dict):
                register_rows.append([_display(module.get("module_id")), _display(module.get("module_name")), _display(module.get("confidence")), _display(module.get("qa_status"))])
        register = PdfTable(register_rows, colWidths=[0.95 * inch, 3.8 * inch, 0.85 * inch, 1.3 * inch], repeatRows=1)
        register.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_NAVY}")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5), ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8DCE5")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F8FA")]),
        ]))
        story.extend([register, PageBreak()])
        for index, module in enumerate(modules):
            if not isinstance(module, dict):
                continue
            story.append(Paragraph(f"{html.escape(_display(module.get('module_id')))} - {html.escape(_display(module.get('module_name')))}", heading))
            detail_rows = [[Paragraph("Path", body), Paragraph("Frozen value", body)]]
            for path, value in _rows(module.get("summary") or {}):
                detail_rows.append([Paragraph(html.escape(path), meta), Paragraph(html.escape(_display(value, limit=900)), body)])
            story.append(detail_table(detail_rows, [2.15 * inch, 4.75 * inch]))
            if index < len(modules) - 1:
                story.append(PageBreak())

    model_periods = calculation.get("periods") if isinstance(calculation.get("periods"), list) else []
    if model:
        story.extend([PageBreak(), Paragraph("MODEL ENGINE V2 - FROZEN IDENTITY", heading)])
        model_identity = [
            ["Engine version", model.get("engine_version")],
            ["Source fingerprint", model.get("source_fingerprint")],
            ["Input fingerprint", model.get("input_fingerprint")],
            ["Calculation hash", model.get("calculation_hash")],
            ["Draft revision", model.get("draft_revision")],
            ["Reporting currency", reporting_currency],
            ["Reporting unit", reporting_unit],
            ["Availability", calculation.get("status")],
            ["Gaps", " | ".join(_display(item) for item in calculation.get("gaps") or []) or "None"],
            ["Warnings", " | ".join(_display(item) for item in calculation.get("warnings") or []) or "None"],
        ]
        story.append(detail_table([
            [Paragraph("Identity", body), Paragraph("Frozen value", body)],
            *[[Paragraph(html.escape(_display(label)), meta), Paragraph(html.escape(_display(value)), body)] for label, value in model_identity],
        ], [2.15 * inch, 4.75 * inch]))
        if override_rows or omitted_override_count:
            story.append(Paragraph("MODEL ENGINE V2 - OVERRIDE LEDGER", heading))
            override_table_rows = [
                [Paragraph("Override", body), Paragraph("Frozen audit detail", body)]
            ]
            for row in override_rows:
                detail = (
                    f"Value: {_display(row['value'])} | Reason: {_display(row['reason'])} | "
                    f"Scope: {_display(row['scope'])} | Source: {_display(row['source'])} | "
                    f"Expires: {_display(row['expires_at'])} | Displaced formula: "
                    f"{_display(row['original_formula'])} | Displaced value: "
                    f"{_display(row['original_value'])}"
                )
                override_table_rows.append([
                    Paragraph(html.escape(
                        f"{row['status']} | {row['node_id']}"
                    ), meta),
                    Paragraph(html.escape(detail), body),
                ])
            if omitted_override_count:
                override_table_rows.append([
                    Paragraph("TRUNCATED", meta),
                    Paragraph(html.escape(
                        f"{omitted_override_count} additional overrides remain in the frozen payload."
                    ), body),
                ])
            story.append(detail_table(override_table_rows, [2.45 * inch, 4.45 * inch]))
        if model_periods:
            story.append(Paragraph(
                f"MODEL ENGINE V2 - CALCULATION - {html.escape(reporting_scale.upper())}",
                heading,
            ))
            calc_rows = [["Period", "Adj EBITDA", "Cash int.", "Debt", "Net debt", "Gross lev.", "Net lev.", "Coverage", "FCF"]]
            for period in model_periods:
                if isinstance(period, dict):
                    calc_rows.append([
                        _display(period.get("period_key")), _display(period.get("adjusted_ebitda")),
                        _display(period.get("cash_interest")), _display(period.get("total_debt")),
                        _display(period.get("net_debt")), _display(period.get("gross_leverage")),
                        _display(period.get("net_leverage")), _display(period.get("interest_coverage")),
                        _display(period.get("free_cash_flow")),
                    ])
            calculation_table = PdfTable(calc_rows, colWidths=[0.78 * inch] + [0.765 * inch] * 8, repeatRows=1)
            calculation_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_NAVY}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 6.2),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8DCE5")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(calculation_table)
            story.append(Paragraph(
                "MODEL ENGINE V2 - DEBT SCHEDULE - "
                f"{html.escape(reporting_unit.upper())} - REPORTING CURRENCY "
                f"{html.escape(reporting_currency)}",
                heading,
            ))
            debt_rows = [[Paragraph("Debt field", body), Paragraph("Frozen value", body)]]
            for period in model_periods:
                if not isinstance(period, dict):
                    continue
                for instrument in period.get("instruments") or []:
                    if not isinstance(instrument, dict):
                        continue
                    prefix = f"{period.get('period_key')}.{instrument.get('instrument_id')}"
                    instrument_currency = instrument_currencies.get(
                        str(instrument.get("instrument_id")), "Unavailable"
                    )
                    debt_rows.append([
                        Paragraph(html.escape(f"{prefix}.instrument_currency"), meta),
                        Paragraph(html.escape(instrument_currency), body),
                    ])
                    for field, value in instrument.items():
                        debt_rows.append([
                            Paragraph(html.escape(f"{prefix}.{field}"), meta),
                            Paragraph(html.escape(_display(value)), body),
                        ])
            story.append(detail_table(debt_rows, [3.45 * inch, 3.45 * inch]))
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return output.getvalue()


async def render_report_export(
    *,
    export_format: Literal["xlsx", "pdf"],
    version_id: str,
    document_sha256: str,
    payload: dict,
    authority: dict,
) -> bytes:
    """Render a binary report off the event loop with bounded fan-out.

    Shielding the worker task lets request cancellation propagate without
    cancelling the underlying thread. The semaphore is released only after the
    renderer actually exits, so cancelled requests cannot bypass the cap while
    their CPU-bound work is still running.
    """
    renderer = render_report_xlsx if export_format == "xlsx" else render_report_pdf
    semaphore = _export_semaphore()
    await semaphore.acquire()
    try:
        task = asyncio.create_task(asyncio.to_thread(
            renderer,
            version_id=version_id,
            document_sha256=document_sha256,
            payload=payload,
            authority=authority,
        ))
    except BaseException:
        semaphore.release()
        raise

    def _release_slot(done: asyncio.Task[bytes]) -> None:
        semaphore.release()
        # A cancelled HTTP request no longer awaits this task. Consume a
        # renderer exception to avoid an unhandled-task warning; non-cancelled
        # callers still receive the same exception from their await below.
        if not done.cancelled():
            done.exception()

    task.add_done_callback(_release_slot)
    return await asyncio.shield(task)
