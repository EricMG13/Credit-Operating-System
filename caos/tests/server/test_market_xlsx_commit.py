"""Phase 2B atomicity, idempotency, lineage, and authorization tests."""

from __future__ import annotations

import io
import json
import sqlite3
import zipfile
from pathlib import Path
from types import SimpleNamespace

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from config import get_settings
from identity import CallerIdentity, get_identity
from main import app
from routes import market_import


def _workbook(rows: list[list[object]] | None = None) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Market"
    worksheet.append([
        "FIGI", "Borrower", "Instrument", "Currency", "Price",
        "Discount Margin", "As Of",
    ])
    for row in rows or [["BBGPHASE2A", "Phase Two Co", "Phase Two TLB", "USD", 99.5, 425, "2026-07-13"]]:
        worksheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _cached_formula_workbook() -> bytes:
    content = _workbook([["BBGPHASE2FORMULA", "Formula Co", "Formula TLB", "USD", "=99+0.5", 425, "2026-07-13"]])
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(content)) as source, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                payload = payload.replace(b"<f>99+0.5</f><v />", b"<f>99+0.5</f><v>99.5</v>")
            target.writestr(item, payload)
    return output.getvalue()


@pytest.fixture
def market_enabled(monkeypatch):
    settings = SimpleNamespace(
        caos_market_xlsx_v2_enabled=True,
        caos_lineage_v2_enabled=True,
        caos_upload_concurrency=2,
        session_secret="phase2-market-test-secret-with-sufficient-length",
    )
    monkeypatch.setattr(market_import, "get_settings", lambda: settings)
    market_import._preview_sem = None
    yield settings
    market_import._preview_sem = None
    app.dependency_overrides.clear()


def _preview(
    client: TestClient,
    content: bytes,
    *,
    mapping: dict | None = None,
    issuer_mappings: dict | None = None,
) -> dict:
    response = client.post(
        "/api/rv/snapshots/import/preview",
        files={"file": ("market.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "mapping": json.dumps(mapping or {}),
            "issuer_mappings": json.dumps(issuer_mappings or {}),
        },
    )
    assert response.status_code == 200, response.text
    return response.json()


def _commit(
    client: TestClient,
    content: bytes,
    preview: dict,
    *,
    mapping: dict | None = None,
    issuer_mappings: dict | None = None,
    preview_sha256: str | None = None,
    preview_token: str | None = None,
):
    return client.post(
        "/api/rv/snapshots/import/commit",
        files={"file": ("market.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "mapping": json.dumps(mapping or {}),
            "issuer_mappings": json.dumps(issuer_mappings or {}),
            "preview_sha256": preview_sha256 or preview["workbook_sha256"],
            "preview_token": preview_token or preview["preview_token"],
            "source_label": "Bloomberg recorded test",
        },
    )


def _db_path() -> Path:
    return Path(str(__import__("database").engine.url.database))


def _identity(analyst_id: str) -> CallerIdentity:
    return CallerIdentity(
        id=analyst_id,
        email=f"{analyst_id}@example.test",
        full_name=analyst_id,
        source="profile",
    )


def test_commit_atomically_creates_owned_source_snapshot_rows_issues_and_lineage(market_enabled):
    content = _workbook([
        ["BBGPHASE2B", "Linked Market Co", "Linked TLB", "USD", 99.5, 425, "2026-07-13"],
        ["BBGPHASE2BAD", "Rejected Market Co", "Rejected TLB", "USD", 98.0, "N/A", "2026-07-13"],
    ])
    with TestClient(app) as client:
        issuer_response = client.post("/api/issuers", json={
            "name": "Linked Market Co", "figi": "BBGPHASE2B",
        })
        assert issuer_response.status_code == 201, issuer_response.text
        issuer_id = issuer_response.json()["id"]
        preview = _preview(client, content)
        assert preview["accepted_count"] == 1
        assert preview["rejected_count"] == 1
        committed = _commit(client, content, preview)
        assert committed.status_code == 200, committed.text
        body = committed.json()
        assert body["existing"] is False
        assert body["instrument_count"] == 1
        assert body["rejected_count"] == 1
        assert body["warning_count"] == 1
        assert body["document_id"] and body["source_manifest_id"]

        listed = client.get("/api/rv/snapshots").json()["snapshots"]
        assert any(row["id"] == body["snapshot_id"] for row in listed)
        context = client.post("/api/analysis/contexts", json={"name": "Imported market drive"})
        assert context.status_code == 201, context.text
        screen = client.post("/api/rv/screens", json={
            "context_id": context.json()["id"], "snapshot_id": body["snapshot_id"],
        })
        assert screen.status_code == 201, screen.text
        screen_body = screen.json()
        assert screen_body["snapshot_source_label"] == "Bloomberg recorded test"
        assert screen_body["snapshot_freshness"]["state"] == body["freshness"]["state"]
        assert screen_body["candidates"][0]["market"]["dm"] == 425.0
        assert screen_body["candidates"][0]["market"]["price"] == 99.5
        assert screen_body["candidates"][0]["evidence"]["source_document_id"] == body["document_id"]

    with sqlite3.connect(_db_path()) as connection:
        snapshot = connection.execute(
            "SELECT analyst_id, document_id, source_manifest_id FROM market_snapshots WHERE id = ?",
            (body["snapshot_id"],),
        ).fetchone()
        assert snapshot == ("local-dev", body["document_id"], body["source_manifest_id"])
        document = connection.execute(
            "SELECT issuer_id, analyst_id, storage_key FROM documents WHERE id = ?",
            (body["document_id"],),
        ).fetchone()
        assert document[0] is None and document[1] == "local-dev"
        assert (Path(get_settings().caos_storage_dir) / document[2]).is_file()
        instrument = connection.execute(
            "SELECT issuer_id, figi FROM market_instruments WHERE snapshot_id = ?",
            (body["snapshot_id"],),
        ).fetchone()
        assert instrument == (issuer_id, "BBGPHASE2B")
        issues = connection.execute(
            "SELECT severity, code FROM market_import_issues WHERE snapshot_id = ?",
            (body["snapshot_id"],),
        ).fetchall()
        assert issues == [("warning", "invalid_discount_margin")]
        lineage = connection.execute(
            "SELECT context_id, analyst_id, artifact_kind, parent_kind FROM lineage_edges WHERE artifact_id IN (?, ?)",
            (f"source_manifest:{body['source_manifest_id']}", f"market_snapshot:{body['snapshot_id']}"),
        ).fetchall()
        assert len(lineage) == 3
        assert all(row[0] is None and row[1] == "local-dev" for row in lineage)


def test_duplicate_commit_returns_existing_without_new_rows_or_file(market_enabled):
    content = _workbook([["BBGPHASE2IDEMP", "Idempotent Co", "Idempotent TLB", "USD", 99.0, 410, "2026-07-13"]])
    with TestClient(app) as client:
        preview = _preview(client, content)
        first = _commit(client, content, preview)
        assert first.status_code == 200, first.text
        second = _commit(client, content, preview)
        assert second.status_code == 200, second.text
        assert second.json()["existing"] is True
        assert second.json()["snapshot_id"] == first.json()["snapshot_id"]
        snapshot_id = first.json()["snapshot_id"]
    with sqlite3.connect(_db_path()) as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM market_snapshots WHERE id = ?", (snapshot_id,)
        ).fetchone()[0] == 1
        assert connection.execute(
            "SELECT COUNT(*) FROM market_instruments WHERE snapshot_id = ?", (snapshot_id,)
        ).fetchone()[0] == 1


def test_commit_rejects_hash_and_mapping_drift_before_writes(market_enabled):
    content = _workbook([["BBGPHASE2DRIFT", "Drift Co", "Drift TLB", "USD", 99.0, 410, "2026-07-13"]])
    app.dependency_overrides[get_identity] = lambda: _identity("drift-owner")
    with TestClient(app) as client:
        preview = _preview(client, content)
        bad_hash = _commit(client, content, preview, preview_sha256="0" * 64)
        assert bad_hash.status_code == 409
        changed_mapping = {"columns": {"price": "Discount Margin"}}
        bad_mapping = _commit(client, content, preview, mapping=changed_mapping)
        assert bad_mapping.status_code == 409
    with sqlite3.connect(_db_path()) as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM market_snapshots WHERE analyst_id = 'drift-owner'"
        ).fetchone()[0] == 0


def test_commit_repeats_antivirus_scan_instead_of_trusting_preview(market_enabled, monkeypatch):
    content = _workbook([["BBGPHASE2SCAN", "Scan Co", "Scan TLB", "USD", 99.0, 410, "2026-07-13"]])
    scans: list[bytes] = []

    async def record_scan(payload: bytes) -> None:
        scans.append(payload)

    monkeypatch.setattr(market_import.avscan, "scan", record_scan)
    app.dependency_overrides[get_identity] = lambda: _identity("scan-owner")
    with TestClient(app) as client:
        preview = _preview(client, content)
        committed = _commit(client, content, preview)
        assert committed.status_code == 200, committed.text
    assert scans == [content, content]


def test_explicit_issuer_mapping_is_token_bound_and_authorized(market_enabled):
    content = _workbook([[None, "Explicit Co", "Explicit TLB", "USD", 99.0, 410, "2026-07-13"]])
    mapping = {"columns": {"instrument_key": "Instrument"}}
    with TestClient(app) as client:
        issuer = client.post("/api/issuers", json={"name": "Explicit Mapped Co"}).json()
        explicit = {"Explicit TLB": issuer["id"]}
        preview = _preview(client, content, mapping=mapping, issuer_mappings=explicit)
        tampered = _commit(client, content, preview, mapping=mapping, issuer_mappings={})
        assert tampered.status_code == 409
        committed = _commit(client, content, preview, mapping=mapping, issuer_mappings=explicit)
        assert committed.status_code == 200, committed.text
        snapshot_id = committed.json()["snapshot_id"]
    with sqlite3.connect(_db_path()) as connection:
        assert connection.execute(
            "SELECT issuer_id FROM market_instruments WHERE snapshot_id = ?", (snapshot_id,)
        ).fetchone()[0] == issuer["id"]


def test_foreign_snapshot_is_404_and_not_listed(market_enabled):
    content = _workbook([["BBGPHASE2PRIVATE", "Private Feed Co", "Private TLB", "USD", 99.0, 410, "2026-07-13"]])
    app.dependency_overrides[get_identity] = lambda: _identity("market-owner-a")
    with TestClient(app) as client:
        preview = _preview(client, content)
        committed = _commit(client, content, preview)
        assert committed.status_code == 200, committed.text
        snapshot_id = committed.json()["snapshot_id"]

        app.dependency_overrides[get_identity] = lambda: _identity("market-owner-b")
        listed = client.get("/api/rv/snapshots")
        assert listed.status_code == 200
        assert snapshot_id not in {row["id"] for row in listed.json()["snapshots"]}
        context = client.post("/api/analysis/contexts", json={"name": "Foreign snapshot probe"})
        assert context.status_code == 201, context.text
        screen = client.post("/api/rv/screens", json={
            "context_id": context.json()["id"], "snapshot_id": snapshot_id,
        })
        assert screen.status_code == 404


def test_same_payload_is_deduplicated_only_within_analyst_scope(market_enabled):
    content = _workbook([["BBGPHASE2SCOPED", "Scoped Feed Co", "Scoped TLB", "USD", 99.0, 410, "2026-07-13"]])
    with TestClient(app) as client:
        app.dependency_overrides[get_identity] = lambda: _identity("scope-a")
        first_preview = _preview(client, content)
        first = _commit(client, content, first_preview).json()
        app.dependency_overrides[get_identity] = lambda: _identity("scope-b")
        foreign_token = _commit(client, content, first_preview)
        assert foreign_token.status_code == 409
        second_preview = _preview(client, content)
        second = _commit(client, content, second_preview).json()
    assert first["snapshot_id"] != second["snapshot_id"]
    assert first["payload_hash"] != second["payload_hash"]


def test_cached_formula_workbook_commits_reopens_and_drives_price_surface(market_enabled):
    content = _cached_formula_workbook()
    app.dependency_overrides[get_identity] = lambda: _identity("formula-owner")
    with TestClient(app) as client:
        preview = _preview(client, content)
        assert preview["formula_cell_count"] == 1
        assert preview["blocking_count"] == 0
        committed = _commit(client, content, preview)
        assert committed.status_code == 200, committed.text
        body = committed.json()
        context = client.post("/api/analysis/contexts", json={"name": "Formula market drive"}).json()
        screen = client.post("/api/rv/screens", json={
            "context_id": context["id"], "snapshot_id": body["snapshot_id"],
        })
        assert screen.status_code == 201, screen.text
        assert screen.json()["candidates"][0]["market"]["price"] == 99.5
    with sqlite3.connect(_db_path()) as connection:
        storage_key = connection.execute(
            "SELECT storage_key FROM documents WHERE id = ?", (body["document_id"],)
        ).fetchone()[0]
    raw_path = Path(get_settings().caos_storage_dir) / storage_key
    formula_view = load_workbook(raw_path, data_only=False, read_only=True)["Market"]["E2"].value
    cached_view = load_workbook(raw_path, data_only=True, read_only=True)["Market"]["E2"].value
    assert formula_view == "=99+0.5"
    assert cached_view == 99.5


def test_lineage_failure_rolls_back_database_and_unique_vault_object(market_enabled, monkeypatch):
    content = _workbook([["BBGPHASE2ROLLBACK", "Rollback Co", "Rollback TLB", "USD", 99.0, 410, "2026-07-13"]])
    vault = Path(get_settings().caos_storage_dir)
    before_files = {path for path in vault.rglob("*") if path.is_file()} if vault.exists() else set()

    async def fail_lineage(*args, **kwargs):
        raise RuntimeError("forced lineage failure")

    monkeypatch.setattr(market_import, "write_owned_artifact_lineage_edge", fail_lineage)
    app.dependency_overrides[get_identity] = lambda: _identity("rollback-owner")
    with TestClient(app, raise_server_exceptions=False) as client:
        preview = _preview(client, content)
        failed = _commit(client, content, preview)
        assert failed.status_code == 500
    after_files = {path for path in vault.rglob("*") if path.is_file()} if vault.exists() else set()
    assert after_files == before_files
    with sqlite3.connect(_db_path()) as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM market_snapshots WHERE analyst_id = 'rollback-owner'"
        ).fetchone()[0] == 0


def test_0055_roundtrips_empty_and_refuses_destructive_evidence_rollback(tmp_path):
    from test_migrations import _alembic

    db_path = tmp_path / "market-xlsx-v2.db"
    db_url = f"sqlite+aiosqlite:///{db_path}"
    assert _alembic("upgrade", "0054", db_url=db_url).returncode == 0
    for command, revision, present in (
        ("upgrade", "0055", True),
        ("downgrade", "0054", False),
        ("upgrade", "0055", True),
    ):
        result = _alembic(command, revision, db_url=db_url)
        assert result.returncode == 0, result.stderr
        with sqlite3.connect(db_path) as connection:
            tables = {row[0] for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            )}
            snapshot_columns = {
                row[1] for row in connection.execute("PRAGMA table_info(market_snapshots)")
            }
            assert ("market_import_issues" in tables) is present
            assert ("analyst_id" in snapshot_columns) is present

    with sqlite3.connect(db_path) as connection:
        connection.execute(
            """INSERT INTO market_snapshots
            (id, analyst_id, as_of, source_label, origin, method, status,
             payload_hash, import_mapping, metadata_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                "rollback-evidence", "rollback-owner", "2026-07-13T00:00:00Z",
                "Imported evidence", "live", "reported", "ready", "e" * 64,
                "{}", "{}", "2026-07-14T00:00:00Z",
            ),
        )
        connection.commit()
    refused = _alembic("downgrade", "0054", db_url=db_url)
    assert refused.returncode != 0
    assert "0055 downgrade refused" in refused.stderr
