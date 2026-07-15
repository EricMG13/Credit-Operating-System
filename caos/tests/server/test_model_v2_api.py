"""API contract tests for the owned, revision-CAS Model Engine v2 routes."""

from __future__ import annotations

import asyncio
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta, timezone
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook


_ISSUER_IDS: set[str] = set()
_CONTEXT_IDS: set[str] = set()
_RUN_IDS: set[str] = set()


def _run(coroutine):
    return asyncio.run(coroutine)


def _identity(owner: str, *, role: str = "analyst", team_id: str | None = None):
    from identity import CallerIdentity

    return CallerIdentity(
        id=owner,
        email=f"{owner}@test.local",
        full_name=owner,
        role=role,
        source="profile",
        team_id=team_id,
    )


def _as(owner: str, *, role: str = "analyst", team_id: str | None = None) -> None:
    from identity import get_identity
    from main import app

    caller = _identity(owner, role=role, team_id=team_id)
    app.dependency_overrides[get_identity] = lambda: caller


def _payload(*, revenue: float = 800.0) -> dict:
    authority = {
        "origin": "analyst",
        "method": "model-v2-api-fixture",
        "source_ids": [],
        "as_of": "2026-07-14T00:00:00Z",
    }
    return {
        "schema_version": 2,
        "reporting_currency": "USD",
        "reporting_unit": "millions",
        "periods": [{
            "period_key": "FY2026",
            "label": "FY26e",
            "kind": "forecast",
            "months": 12,
            "revenue": revenue,
            "reported_ebitda": 100.0,
            "adjustments": 10.0,
            "cash": 20.0,
            "taxes": 5.0,
            "capex": 10.0,
            "working_capital_change": -2.0,
            "other_cash_flow": 0.0,
            "authority": authority,
        }],
        "debt_instruments": [{
            "instrument_id": "tlb-1",
            "name": "First-lien term loan",
            "priority": 1,
            "seniority": "1L",
            "currency": "USD",
            "rate_type": "floating",
            "authority": authority,
            "periods": [{
                "period_key": "FY2026",
                "opening_balance": 200.0,
                "closing_balance": 190.0,
                "draws": 0.0,
                "repayments": 10.0,
                "scheduled_amortization": 0.0,
                "commitment": 220.0,
                "benchmark_rate": 0.04,
                "floor_rate": 0.05,
                "spread_rate": 0.03,
                "coupon_rate": 0.01,
                "commitment_fee_rate": 0.005,
                "pik_rate": 0.0,
                "cash_fees": 1.0,
                "hedge_effect": -0.5,
                "fx_rate": 1.0,
            }],
        }],
        "overrides": [],
        "source_ids": [],
    }


async def _seed_case(
    owner: str,
    *,
    context: bool = False,
    cp1: bool = False,
    team_id: str | None = None,
    cp1_currency: str | None = "USD",
    cp1_reporting_unit: str | None = "millions",
    cp1_reporting_unit_present: bool = True,
    reporting_profile: bool = True,
    fiscal_year_end_month: int = 12,
    fiscal_year_end_day: int = 31,
) -> dict[str, str | None]:
    from database import (
        AnalysisContextRecord,
        AsyncSessionLocal,
        Issuer,
        IssuerReportingProfile,
        ModuleOutput,
        Run,
    )

    token = uuid.uuid4().hex[:20]
    issuer_id = f"m2i-{token}"
    context_id = f"m2c-{token}" if context else None
    run_id = f"m2r-{token}" if cp1 else None
    cp1_id = f"m2o-{token}" if cp1 else None
    async with AsyncSessionLocal() as db:
        db.add(Issuer(
            id=issuer_id,
            name=f"Model V2 API {token}",
            team_id=team_id,
            created_by=owner,
        ))
        if context_id:
            db.add(AnalysisContextRecord(
                id=context_id,
                analyst_id=owner,
                name=f"Model V2 context {token}",
                issuer_ids=[issuer_id],
            ))
        if run_id:
            now = datetime.now(timezone.utc)
            if reporting_profile:
                db.add(IssuerReportingProfile(
                    issuer_id=issuer_id,
                    cadence="quarterly",
                    fiscal_year_end_month=fiscal_year_end_month,
                    fiscal_year_end_day=fiscal_year_end_day,
                    reporting_lag_days=45,
                    grace_days=7,
                    authority={"origin": "analyst"},
                    updated_by=owner,
                    updated_at=now,
                ))
            db.add(Run(
                id=run_id,
                issuer_id=issuer_id,
                analyst_id=owner,
                status="complete",
                as_of_date="2026-03-31",
                completed_at=now,
                created_at=now,
            ))
            await db.flush()
            runtime_output = {
                "normalized_financials": {
                    "revenue": {"FY2025": 900.0, "LTM_Q1_26": 950.0},
                    "adj_ebitda": {"FY2025": 95.0, "LTM_Q1_26": 100.0},
                    "net_debt_ltm": 400.0,
                    "interest_coverage_ltm": 2.0,
                },
            }
            if cp1_currency is not None:
                runtime_output["currency"] = cp1_currency
            if cp1_reporting_unit_present:
                runtime_output["reporting_unit"] = cp1_reporting_unit
            db.add(ModuleOutput(
                id=cp1_id,
                run_id=run_id,
                module_id="CP-1",
                module_name="Reported Financials",
                owned_object="normalized_financials",
                runtime_output=runtime_output,
                limitation_flags=[],
            ))
        await db.commit()
    _ISSUER_IDS.add(issuer_id)
    if context_id:
        _CONTEXT_IDS.add(context_id)
    if run_id:
        _RUN_IDS.add(run_id)
    return {
        "issuer_id": issuer_id,
        "context_id": context_id,
        "run_id": run_id,
        "cp1_id": cp1_id,
    }


async def _context_artifacts(context_id: str) -> dict:
    from database import AnalysisContextRecord, AsyncSessionLocal

    async with AsyncSessionLocal() as db:
        row = await db.get(AnalysisContextRecord, context_id)
        assert row is not None
        return dict(row.artifacts or {})


async def _tamper_saved_calculation(issuer_id: str, owner: str) -> str:
    """Simulate a stored result whose time-sensitive inputs have gone stale."""
    from sqlalchemy import select
    from database import AsyncSessionLocal, ModelDraftV2

    async with AsyncSessionLocal() as db:
        row = (await db.execute(select(ModelDraftV2).where(
            ModelDraftV2.issuer_id == issuer_id,
            ModelDraftV2.analyst_id == owner,
        ))).scalar_one()
        row.calculation_hash = "f" * 64
        calculation = dict(row.calculation)
        calculation["calculation_hash"] = row.calculation_hash
        row.calculation = calculation
        await db.commit()
        return row.calculation_hash


async def _tamper_checkpoint_payload_hash(checkpoint_id: str) -> None:
    from database import AsyncSessionLocal, ModelCheckpoint

    async with AsyncSessionLocal() as db:
        row = await db.get(ModelCheckpoint, checkpoint_id)
        assert row is not None
        row.payload_hash = "0" * 64
        await db.commit()


async def _checkpoint_authority(checkpoint_id: str) -> dict:
    from database import AsyncSessionLocal, ModelCheckpoint

    async with AsyncSessionLocal() as db:
        row = await db.get(ModelCheckpoint, checkpoint_id)
        assert row is not None
        return dict(row.authority or {})


async def _checkpoint_payload(checkpoint_id: str) -> dict:
    from database import AsyncSessionLocal, ModelCheckpoint

    async with AsyncSessionLocal() as db:
        row = await db.get(ModelCheckpoint, checkpoint_id)
        assert row is not None
        return dict(row.payload or {})


async def _add_owned_run(owner: str, issuer_id: str) -> str:
    from database import AsyncSessionLocal, Run

    run_id = f"m2r-{uuid.uuid4().hex[:20]}"
    now = datetime.now(timezone.utc)
    async with AsyncSessionLocal() as db:
        db.add(Run(
            id=run_id,
            issuer_id=issuer_id,
            analyst_id=owner,
            status="complete",
            as_of_date="2026-06-30",
            completed_at=now,
            created_at=now,
        ))
        await db.commit()
    _RUN_IDS.add(run_id)
    return run_id


async def _bind_draft_to_run(issuer_id: str, owner: str, run_id: str) -> None:
    from sqlalchemy import select
    from database import AsyncSessionLocal, ModelDraftV2

    async with AsyncSessionLocal() as db:
        draft = (await db.execute(select(ModelDraftV2).where(
            ModelDraftV2.issuer_id == issuer_id,
            ModelDraftV2.analyst_id == owner,
        ))).scalar_one()
        draft.source_run_id = run_id
        await db.commit()


async def _cleanup() -> None:
    if not _ISSUER_IDS:
        return
    from sqlalchemy import delete
    from database import (
        AnalysisContextRecord,
        AsyncSessionLocal,
        Issuer,
        LineageEdge,
        ModelCheckpoint,
        ModelDraftV2,
        ModelOverrideEvent,
        ModelWorkbookImport,
        ModuleOutput,
        Run,
    )

    async with AsyncSessionLocal() as db:
        await db.execute(delete(ModelWorkbookImport).where(
            ModelWorkbookImport.issuer_id.in_(_ISSUER_IDS)
        ))
        await db.execute(delete(ModelOverrideEvent).where(
            ModelOverrideEvent.issuer_id.in_(_ISSUER_IDS)
        ))
        await db.execute(delete(LineageEdge).where(
            LineageEdge.context_id.in_(_CONTEXT_IDS)
        ))
        await db.execute(delete(ModelCheckpoint).where(
            ModelCheckpoint.issuer_id.in_(_ISSUER_IDS)
        ))
        await db.execute(delete(ModelDraftV2).where(
            ModelDraftV2.issuer_id.in_(_ISSUER_IDS)
        ))
        await db.execute(delete(ModuleOutput).where(ModuleOutput.run_id.in_(_RUN_IDS)))
        await db.execute(delete(Run).where(Run.id.in_(_RUN_IDS)))
        await db.execute(delete(AnalysisContextRecord).where(
            AnalysisContextRecord.id.in_(_CONTEXT_IDS)
        ))
        await db.execute(delete(Issuer).where(Issuer.id.in_(_ISSUER_IDS)))
        await db.commit()


@pytest.fixture(scope="module")
def client():
    from config import get_settings
    from main import app

    settings = get_settings()
    original = (
        settings.caos_model_engine_v2_enabled,
        settings.caos_lineage_v2_enabled,
        settings.caos_tenancy_enabled,
    )
    settings.caos_model_engine_v2_enabled = True
    settings.caos_lineage_v2_enabled = True
    settings.caos_tenancy_enabled = False
    with TestClient(app) as value:
        yield value
        _run(_cleanup())
    (
        settings.caos_model_engine_v2_enabled,
        settings.caos_lineage_v2_enabled,
        settings.caos_tenancy_enabled,
    ) = original


@pytest.fixture(autouse=True)
def _reset_model_v2_contract():
    from config import get_settings
    from identity import get_identity
    from main import app

    settings = get_settings()
    settings.caos_model_engine_v2_enabled = True
    settings.caos_lineage_v2_enabled = True
    settings.caos_tenancy_enabled = False
    yield
    app.dependency_overrides.pop(get_identity, None)
    settings.caos_model_engine_v2_enabled = True
    settings.caos_lineage_v2_enabled = True
    settings.caos_tenancy_enabled = False


def _save(client, issuer_id: str, *, expected_revision: int = 0, payload: dict | None = None, context_id: str | None = None):
    return client.put(
        f"/api/models/v2/{issuer_id}",
        json={
            "expected_revision": expected_revision,
            "payload": payload or _payload(),
            "context_id": context_id,
        },
    )


def _set_override(client, issuer_id: str, expected_revision: int, *, value: float = 80.0):
    expires_at = datetime.now(timezone.utc) + timedelta(days=30)
    return client.post(
        f"/api/models/v2/{issuer_id}/overrides",
        json={
            "expected_revision": expected_revision,
            "action": "set",
            "override": {
                "node_id": "calc:FY2026:adjusted_ebitda",
                "value_type": "number",
                "value": value,
                "reason": "IC haircut",
                "scope": "draft",
                "source": "CP-6A",
                "expires_at": expires_at.isoformat(),
            },
        },
    )


def test_flag_off_returns_non_enumerable_404(client):
    from config import get_settings

    get_settings().caos_model_engine_v2_enabled = False
    response = client.get("/api/models/v2/not-visible")
    assert response.status_code == 404
    assert response.json() == {"detail": "Not Found"}


def test_calculation_preview_requires_write_role(client):
    owner = f"model-v2-calculate-viewer-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner, role="viewer")

    response = client.post(
        f"/api/models/v2/{issuer_id}/calculate",
        json={"payload": _payload()},
    )

    assert response.status_code == 403


@pytest.mark.parametrize(
    ("field_name", "invalid_value"),
    [
        ("reporting_currency", None),
        ("reporting_currency", "ZZZ"),
        ("reporting_unit", None),
        ("reporting_unit", "trillions"),
    ],
)
def test_calculation_api_never_defaults_monetary_identity(
    client,
    field_name: str,
    invalid_value: str | None,
):
    owner = f"model-v2-money-id-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner)
    payload = _payload()
    if invalid_value is None:
        payload.pop(field_name)
    else:
        payload[field_name] = invalid_value

    response = client.post(
        f"/api/models/v2/{issuer_id}/calculate",
        json={"payload": payload},
    )

    assert response.status_code == 422
    assert field_name in response.text


def test_suggested_then_saved_read_contract(client):
    owner = f"model-v2-read-{uuid.uuid4().hex[:10]}"
    seeded = _run(_seed_case(owner, cp1=True))
    issuer_id = str(seeded["issuer_id"])
    _as(owner)

    suggested = client.get(f"/api/models/v2/{issuer_id}")
    assert suggested.status_code == 200, suggested.text
    body = suggested.json()
    assert body["authority"] == "model-engine-v2"
    assert body["availability"] == "suggested"
    assert body["record"] is None
    assert body["suggested_payload"]["debt_instruments"] == []
    assert seeded["run_id"] in body["suggested_payload"]["source_ids"]
    assert body["suggested_source_run_id"] == seeded["run_id"]
    assert body["suggested_calculation"]["status"] == "partial"

    exact = client.get(
        f"/api/models/v2/{issuer_id}", params={"run_id": seeded["run_id"]}
    )
    assert exact.status_code == 200, exact.text
    assert exact.json()["suggested_source_run_id"] == seeded["run_id"]

    saved = client.put(
        f"/api/models/v2/{issuer_id}",
        json={
            "expected_revision": 0,
            "payload": body["suggested_payload"],
            "source_run_id": body["suggested_source_run_id"],
        },
    )
    assert saved.status_code == 200, saved.text
    assert saved.json()["revision"] == 1
    reread = client.get(f"/api/models/v2/{issuer_id}")
    assert reread.status_code == 200
    assert reread.json()["availability"] == "saved"
    assert reread.json()["record"]["calculation_hash"] == saved.json()["calculation_hash"]
    assert reread.json()["suggested_payload"] is None
    assert reread.json()["suggested_source_run_id"] is None

    other_run_id = _run(_add_owned_run(owner, issuer_id))
    mismatch = client.get(
        f"/api/models/v2/{issuer_id}", params={"run_id": other_run_id}
    )
    assert mismatch.status_code == 409
    assert "different source run" in mismatch.text


def test_gbp_cp1_currency_survives_api_workbook_and_checkpoint(client):
    owner = f"model-v2-gbp-{uuid.uuid4().hex[:10]}"
    seeded = _run(_seed_case(
        owner,
        context=True,
        cp1=True,
        cp1_currency="£",
        cp1_reporting_unit="£M",
    ))
    issuer_id = str(seeded["issuer_id"])
    context_id = str(seeded["context_id"])
    run_id = str(seeded["run_id"])
    _as(owner)

    suggested = client.get(f"/api/models/v2/{issuer_id}")
    assert suggested.status_code == 200, suggested.text
    suggestion = suggested.json()
    assert suggestion["suggested_payload"]["reporting_currency"] == "GBP"
    assert suggestion["suggested_payload"]["reporting_unit"] == "millions"

    saved = client.put(
        f"/api/models/v2/{issuer_id}",
        json={
            "expected_revision": 0,
            "context_id": context_id,
            "source_run_id": run_id,
            "payload": suggestion["suggested_payload"],
        },
    )
    assert saved.status_code == 200, saved.text
    assert saved.json()["payload"]["reporting_currency"] == "GBP"

    exported = client.get(f"/api/models/v2/{issuer_id}/workbook/export")
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(BytesIO(exported.content), read_only=True)
    cover = {
        key: value
        for key, value in workbook["Cover"].iter_rows(
            min_row=2, max_col=2, values_only=True
        )
        if key
    }
    assert cover["reporting_currency"] == "GBP"
    assert cover["reporting_unit"] == "millions"
    workbook.close()

    checkpoint = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints",
        json={
            "context_id": context_id,
            "label": "GBP live checkpoint",
            "expected_revision": 1,
            "calculation_hash": saved.json()["calculation_hash"],
        },
    )
    assert checkpoint.status_code == 201, checkpoint.text
    frozen = _run(_checkpoint_payload(checkpoint.json()["id"]))
    assert frozen["payload"]["reporting_currency"] == "GBP"
    assert frozen["payload"]["reporting_unit"] == "millions"


@pytest.mark.parametrize(
    ("seed_kwargs", "detail"),
    [
        ({"cp1_currency": None}, "reporting currency is missing"),
        ({"cp1_reporting_unit_present": False}, "reporting unit is missing"),
        ({"cp1_reporting_unit": None}, "reporting unit is invalid"),
        ({"reporting_profile": False}, "no canonical reported financial period"),
    ],
)
def test_cp1_binding_degrades_when_currency_or_fiscal_profile_is_unknown(
    client,
    seed_kwargs,
    detail,
):
    owner = f"model-v2-source-gap-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(
        owner,
        cp1=True,
        **seed_kwargs,
    ))["issuer_id"])
    _as(owner)

    response = client.get(f"/api/models/v2/{issuer_id}")

    assert response.status_code == 200
    assert response.json()["availability"] == "insufficient_source"
    assert detail in response.json()["detail"]


def test_atomic_integer_cas_allows_one_concurrent_winner(client):
    owner = f"model-v2-cas-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner)
    created = _save(client, issuer_id)
    assert created.status_code == 200, created.text
    assert created.json()["revision"] == 1

    barrier = threading.Barrier(2)

    def update_once(marker: str):
        payload = _payload()
        payload["ui_preferences"] = {
            "collapsed_rows": [marker],
        }
        barrier.wait(timeout=5)
        return _save(
            client,
            issuer_id,
            expected_revision=1,
            payload=payload,
        )

    with ThreadPoolExecutor(max_workers=2) as pool:
        responses = list(pool.map(update_once, ("first", "second")))
    assert sorted(response.status_code for response in responses) == [200, 409]
    winner = next(response.json() for response in responses if response.status_code == 200)
    conflict = next(response.json() for response in responses if response.status_code == 409)
    assert winner["revision"] == 2
    assert conflict["detail"]["current_revision"] == 2

    current = client.get(f"/api/models/v2/{issuer_id}").json()["record"]
    assert current["revision"] == 2
    assert (
        current["payload"]["ui_preferences"]["collapsed_rows"]
        == winner["payload"]["ui_preferences"]["collapsed_rows"]
    )
    stale_payload = _payload()
    stale_payload["ui_preferences"] = {"collapsed_rows": ["stale"]}
    stale = _save(
        client,
        issuer_id,
        expected_revision=1,
        payload=stale_payload,
    )
    assert stale.status_code == 409
    assert stale.json()["detail"]["current_revision"] == 2
    assert client.get(f"/api/models/v2/{issuer_id}").json()["record"]["revision"] == 2


def test_generic_put_cannot_bypass_cell_audit_or_forge_authority(client):
    owner = f"model-v2-put-guard-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner)
    created = _save(client, issuer_id)
    assert created.status_code == 200

    changed_base = _payload(revenue=999.0)
    raw_edit = _save(
        client, issuer_id, expected_revision=1, payload=changed_base
    )
    assert raw_edit.status_code == 422

    injected = _payload()
    injected["overrides"] = [{
        "node_id": "calc:FY2026:adjusted_ebitda",
        "value_type": "number",
        "value": 80,
        "reason": "Attempted PUT bypass",
        "scope": "draft",
        "source": "test-source",
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=30)).isoformat(),
    }]
    override_edit = _save(
        client, issuer_id, expected_revision=1, payload=injected
    )
    assert override_edit.status_code == 422

    fresh_issuer = str(_run(_seed_case(owner))["issuer_id"])
    initial_override = _save(client, fresh_issuer, payload=injected)
    assert initial_override.status_code == 422

    imported = _payload()
    for period in imported["periods"]:
        period["authority"]["origin"] = "imported"
    for instrument in imported["debt_instruments"]:
        instrument["authority"]["origin"] = "imported"
    imported_issuer = str(_run(_seed_case(owner))["issuer_id"])
    assert _save(client, imported_issuer, payload=imported).status_code == 422

    reference = _payload()
    for period in reference["periods"]:
        period["authority"]["origin"] = "reference"
    for instrument in reference["debt_instruments"]:
        instrument["authority"]["origin"] = "reference"
    reference_issuer = str(_run(_seed_case(owner))["issuer_id"])
    assert _save(client, reference_issuer, payload=reference).status_code == 422

    current = client.get(f"/api/models/v2/{issuer_id}").json()["record"]
    assert current["revision"] == 1
    assert current["payload"]["periods"][0]["revenue"] == 800
    assert client.get(f"/api/models/v2/{issuer_id}/history").json() == []


def test_live_initial_payload_must_match_exact_cp1_binder(client):
    owner = f"model-v2-live-provenance-{uuid.uuid4().hex[:10]}"
    seeded = _run(_seed_case(owner, cp1=True))
    issuer_id = str(seeded["issuer_id"])
    _as(owner)
    suggested = client.get(f"/api/models/v2/{issuer_id}").json()
    forged = suggested["suggested_payload"]
    forged["periods"][0]["revenue"] = 999_999
    response = client.put(
        f"/api/models/v2/{issuer_id}",
        json={
            "expected_revision": 0,
            "payload": forged,
            "source_run_id": suggested["suggested_source_run_id"],
        },
    )
    assert response.status_code == 422
    assert client.get(f"/api/models/v2/{issuer_id}").json()["record"] is None


def test_non_live_draft_cannot_be_relabelled_to_an_owned_run(client):
    owner = f"model-v2-non-live-run-{uuid.uuid4().hex[:10]}"
    seeded = _run(_seed_case(owner, cp1=True))
    issuer_id = str(seeded["issuer_id"])
    _as(owner)
    created = _save(client, issuer_id)
    assert created.status_code == 200, created.text

    rebound = client.put(
        f"/api/models/v2/{issuer_id}",
        json={
            "expected_revision": 1,
            "payload": created.json()["payload"],
            "source_run_id": seeded["run_id"],
        },
    )

    assert rebound.status_code == 422
    assert "cannot be relabelled" in rebound.json()["detail"]
    current = client.get(f"/api/models/v2/{issuer_id}").json()["record"]
    assert current["source_run_id"] is None
    assert current["revision"] == 1


def test_owner_and_team_isolation_return_404_without_foreign_data(client):
    owner = f"model-v2-owner-{uuid.uuid4().hex[:10]}"
    other = f"model-v2-other-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner)
    assert _save(client, issuer_id).status_code == 200
    changed = _set_override(client, issuer_id, 1)
    assert changed.status_code == 200, changed.text
    event_id = client.get(f"/api/models/v2/{issuer_id}/history").json()[0]["id"]

    _as(other)
    foreign_read = client.get(f"/api/models/v2/{issuer_id}")
    assert foreign_read.status_code == 200
    assert foreign_read.json()["record"] is None
    assert foreign_read.json()["availability"] == "unavailable"
    assert client.get(f"/api/models/v2/{issuer_id}/history").json() == []
    replay = client.post(
        f"/api/models/v2/{issuer_id}/history/{event_id}/replay",
        json={"expected_revision": 2, "mode": "undo"},
    )
    assert replay.status_code == 404

    from config import get_settings

    team_issuer = str(_run(_seed_case(owner, team_id="team-a"))["issuer_id"])
    get_settings().caos_tenancy_enabled = True
    _as(owner, team_id="team-b")
    hidden = client.get(f"/api/models/v2/{team_issuer}")
    assert hidden.status_code == 404
    assert hidden.json() == {"detail": "Issuer not found"}


def test_read_only_role_cannot_mutate_existing_draft(client):
    owner = f"model-v2-viewer-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner)
    created = _save(client, issuer_id)
    assert created.status_code == 200

    _as(owner, role="viewer")
    save = _save(client, issuer_id, expected_revision=1, payload=_payload(revenue=999.0))
    override = _set_override(client, issuer_id, 1)
    assert save.status_code == 403
    assert override.status_code == 403
    assert client.get(f"/api/models/v2/{issuer_id}").json()["record"]["revision"] == 1


def test_override_capacity_failure_is_controlled_and_atomic(client, monkeypatch):
    owner = f"model-v2-override-capacity-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner)
    created = _save(client, issuer_id)
    assert created.status_code == 200

    from routes import model_v2 as route

    def fail_capacity(*_args, **_kwargs):
        raise ValueError("Model override capacity reached.")

    monkeypatch.setattr(route, "replace_active_override", fail_capacity)
    response = _set_override(client, issuer_id, 1)

    assert response.status_code == 422
    assert response.json()["detail"] == "Model override capacity reached."
    current = client.get(f"/api/models/v2/{issuer_id}").json()["record"]
    assert current["revision"] == 1
    assert current["payload"]["overrides"] == []
    assert client.get(f"/api/models/v2/{issuer_id}/history").json() == []


def test_override_set_reset_propagates_and_audits(client):
    owner = f"model-v2-override-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner)
    created = _save(client, issuer_id)
    original_hash = created.json()["calculation_hash"]

    set_response = _set_override(client, issuer_id, 1)
    assert set_response.status_code == 200, set_response.text
    set_body = set_response.json()
    period = set_body["calculation"]["periods"][0]
    assert set_body["revision"] == 2
    assert period["adjusted_ebitda"] == 80.0
    assert period["gross_leverage"] == pytest.approx(190.0 / 80.0)
    assert period["interest_coverage"] == pytest.approx(80.0 / period["cash_interest"])
    assert set_body["payload"]["overrides"][0]["node_id"] == "calc:FY2026:adjusted_ebitda"

    set_history = client.get(f"/api/models/v2/{issuer_id}/history").json()
    assert len(set_history) == 1
    assert set_history[0]["action"] == "set"
    assert set_history[0]["original_formula"] == "reported_ebitda + adjustments"
    assert set_history[0]["original_value"] == {"value": 110.0}
    assert set_history[0]["before_value"] is None
    assert set_history[0]["after_value"]["value"] == 80.0

    reset = client.post(
        f"/api/models/v2/{issuer_id}/overrides",
        json={
            "expected_revision": 2,
            "action": "reset",
            "node_id": "calc:FY2026:adjusted_ebitda",
        },
    )
    assert reset.status_code == 200, reset.text
    reset_body = reset.json()
    assert reset_body["revision"] == 3
    assert reset_body["payload"]["overrides"] == []
    assert reset_body["calculation"]["periods"][0]["adjusted_ebitda"] == 110.0
    assert reset_body["calculation"]["periods"][0]["gross_leverage"] == pytest.approx(190.0 / 110.0)
    assert reset_body["calculation_hash"] == original_hash
    history = client.get(f"/api/models/v2/{issuer_id}/history").json()
    assert [event["action"] for event in history] == ["reset", "set"]
    assert history[0]["before_value"]["value"] == 80.0
    assert history[0]["after_value"] is None


def test_derived_override_requires_source_and_future_expiry_without_mutation(client):
    owner = f"model-v2-override-governance-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner)
    created = _save(client, issuer_id)
    assert created.status_code == 200
    base_override = {
        "node_id": "calc:FY2026:adjusted_ebitda",
        "value_type": "number",
        "value": 80.0,
        "reason": "Governance fixture",
        "scope": "draft",
    }

    missing_source = client.post(
        f"/api/models/v2/{issuer_id}/overrides",
        json={
            "expected_revision": 1,
            "action": "set",
            "override": {
                **base_override,
                "expires_at": (
                    datetime.now(timezone.utc) + timedelta(days=30)
                ).isoformat(),
            },
        },
    )
    assert missing_source.status_code == 422

    missing_expiry = client.post(
        f"/api/models/v2/{issuer_id}/overrides",
        json={
            "expected_revision": 1,
            "action": "set",
            "override": {**base_override, "source": "test-source"},
        },
    )
    assert missing_expiry.status_code == 422

    past_expiry = client.post(
        f"/api/models/v2/{issuer_id}/overrides",
        json={
            "expected_revision": 1,
            "action": "set",
            "override": {
                **base_override,
                "source": "test-source",
                "expires_at": (
                    datetime.now(timezone.utc) - timedelta(minutes=1)
                ).isoformat(),
            },
        },
    )
    assert past_expiry.status_code == 422
    assert past_expiry.json()["detail"] == (
        "Derived-cell override expiry must be in the future."
    )

    current = client.get(f"/api/models/v2/{issuer_id}").json()["record"]
    assert current["revision"] == 1
    assert current["payload"]["overrides"] == []
    assert client.get(f"/api/models/v2/{issuer_id}/history").json() == []


def test_override_undo_and_redo_are_revisioned_audit_events(client):
    owner = f"model-v2-replay-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner)
    assert _save(client, issuer_id).status_code == 200
    assert _set_override(client, issuer_id, 1).status_code == 200
    event_id = client.get(f"/api/models/v2/{issuer_id}/history").json()[0]["id"]

    undo = client.post(
        f"/api/models/v2/{issuer_id}/history/{event_id}/replay",
        json={"expected_revision": 2, "mode": "undo"},
    )
    assert undo.status_code == 200, undo.text
    assert undo.json()["revision"] == 3
    assert undo.json()["payload"]["overrides"] == []
    assert undo.json()["calculation"]["periods"][0]["adjusted_ebitda"] == 110.0

    redo = client.post(
        f"/api/models/v2/{issuer_id}/history/{event_id}/replay",
        json={"expected_revision": 3, "mode": "redo"},
    )
    assert redo.status_code == 200, redo.text
    assert redo.json()["revision"] == 4
    assert redo.json()["payload"]["overrides"][0]["value"] == 80.0
    assert redo.json()["calculation"]["periods"][0]["adjusted_ebitda"] == 80.0

    history = client.get(f"/api/models/v2/{issuer_id}/history").json()
    assert [event["action"] for event in history] == ["redo", "undo", "set"]
    assert history[0]["inverse_event_id"] == event_id
    assert history[1]["inverse_event_id"] == event_id


def test_replay_rejects_an_old_event_after_the_cell_changed_again(client):
    owner = f"model-v2-replay-stale-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner)
    assert _save(client, issuer_id).status_code == 200
    assert _set_override(client, issuer_id, 1, value=80).status_code == 200
    old_event_id = client.get(f"/api/models/v2/{issuer_id}/history").json()[0]["id"]
    assert _set_override(client, issuer_id, 2, value=70).status_code == 200

    stale = client.post(
        f"/api/models/v2/{issuer_id}/history/{old_event_id}/replay",
        json={"expected_revision": 3, "mode": "undo"},
    )

    assert stale.status_code == 409
    current = client.get(f"/api/models/v2/{issuer_id}").json()["record"]
    assert current["revision"] == 3
    assert current["payload"]["overrides"][0]["value"] == 70


def test_override_batch_is_one_revision_with_complete_audit(client):
    owner = f"model-v2-batch-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner)
    assert _save(client, issuer_id).status_code == 200

    response = client.post(
        f"/api/models/v2/{issuer_id}/overrides/batch",
        json={
            "expected_revision": 1,
            "mutations": [
                {
                    "action": "set",
                    "override": {
                        "node_id": "calc:FY2026:adjusted_ebitda",
                        "value_type": "number",
                        "value": 80.0,
                        "reason": "IC haircut",
                        "scope": "draft",
                        "source": "CP-6A",
                        "expires_at": (
                            datetime.now(timezone.utc) + timedelta(days=30)
                        ).isoformat(),
                    },
                },
                {
                    "action": "set",
                    "override": {
                        "node_id": "input:FY2026:cash",
                        "value_type": "number",
                        "value": 30.0,
                        "reason": "Updated cash bridge",
                        "scope": "draft",
                    },
                },
            ],
        },
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["revision"] == 2
    period = body["calculation"]["periods"][0]
    assert period["adjusted_ebitda"] == 80.0
    assert period["cash"] == 30.0
    assert period["net_debt"] == 160.0
    assert period["net_leverage"] == pytest.approx(2.0)

    history = client.get(f"/api/models/v2/{issuer_id}/history").json()
    assert len(history) == 2
    assert {event["node_id"] for event in history} == {
        "calc:FY2026:adjusted_ebitda",
        "input:FY2026:cash",
    }
    assert {event["revision"] for event in history} == {2}

    undo = client.post(
        f"/api/models/v2/{issuer_id}/history/{history[0]['id']}/replay",
        json={"expected_revision": 2, "mode": "undo"},
    )
    assert undo.status_code == 200, undo.text
    assert undo.json()["revision"] == 3
    assert undo.json()["payload"]["overrides"] == []
    undone_period = undo.json()["calculation"]["periods"][0]
    assert undone_period["adjusted_ebitda"] == 110.0
    assert undone_period["cash"] == 20.0

    redo = client.post(
        f"/api/models/v2/{issuer_id}/history/{history[0]['id']}/replay",
        json={"expected_revision": 3, "mode": "redo"},
    )
    assert redo.status_code == 200, redo.text
    assert redo.json()["revision"] == 4
    redone = {item["node_id"]: item for item in redo.json()["payload"]["overrides"]}
    assert redone["calc:FY2026:adjusted_ebitda"]["value"] == 80.0
    assert redone["input:FY2026:cash"]["value"] == 30.0
    redone_period = redo.json()["calculation"]["periods"][0]
    assert redone_period["adjusted_ebitda"] == 80.0
    assert redone_period["cash"] == 30.0
    replay_history = client.get(f"/api/models/v2/{issuer_id}/history").json()
    assert [event["action"] for event in replay_history] == [
        "redo", "redo", "undo", "undo", "set", "set",
    ]


def test_override_batch_rejects_duplicate_or_invalid_set_without_mutation(client):
    owner = f"model-v2-batch-invalid-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner)
    created = _save(client, issuer_id)
    assert created.status_code == 200

    override = {
        "node_id": "input:FY2026:cash",
        "value_type": "number",
        "value": 30.0,
        "reason": "Duplicate fixture",
    }
    duplicate = client.post(
        f"/api/models/v2/{issuer_id}/overrides/batch",
        json={
            "expected_revision": 1,
            "mutations": [
                {"action": "set", "override": override},
                {"action": "set", "override": {**override, "value": 40.0}},
            ],
        },
    )
    assert duplicate.status_code == 422

    invalid_second = client.post(
        f"/api/models/v2/{issuer_id}/overrides/batch",
        json={
            "expected_revision": 1,
            "mutations": [
                {"action": "set", "override": override},
                {
                    "action": "set",
                    "override": {
                        "node_id": "calc:FY2026:not_a_node",
                        "value_type": "number",
                        "value": 1.0,
                        "reason": "Invalid fixture",
                        "source": "test-source",
                        "expires_at": (
                            datetime.now(timezone.utc) + timedelta(days=30)
                        ).isoformat(),
                    },
                },
            ],
        },
    )
    assert invalid_second.status_code == 422
    current = client.get(f"/api/models/v2/{issuer_id}").json()["record"]
    assert current["revision"] == 1
    assert current["calculation_hash"] == created.json()["calculation_hash"]
    assert current["payload"]["overrides"] == []
    assert client.get(f"/api/models/v2/{issuer_id}/history").json() == []


def test_checkpoint_create_restore_and_foreign_owner_404(client):
    owner = f"model-v2-checkpoint-{uuid.uuid4().hex[:10]}"
    other = f"model-v2-checkpoint-other-{uuid.uuid4().hex[:10]}"
    seeded = _run(_seed_case(owner, context=True))
    issuer_id = str(seeded["issuer_id"])
    context_id = str(seeded["context_id"])
    _as(owner)
    created = _save(client, issuer_id, context_id=context_id)
    assert created.status_code == 200, created.text
    original_hash = created.json()["calculation_hash"]

    checkpoint = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints",
        json={
            "context_id": context_id,
            "label": "IC base case",
            "expected_revision": 1,
            "calculation_hash": original_hash,
        },
    )
    assert checkpoint.status_code == 201, checkpoint.text
    checkpoint_body = checkpoint.json()
    checkpoint_id = checkpoint_body["id"]
    assert checkpoint_body["engine_version"] == "2.0.0"
    assert checkpoint_body["calculation_hash"] == original_hash
    assert checkpoint_body["draft_revision"] == 2
    listed = client.get(f"/api/models/v2/{issuer_id}/checkpoints")
    assert listed.status_code == 200
    assert [row["id"] for row in listed.json()] == [checkpoint_id]
    artifacts = _run(_context_artifacts(context_id))
    assert artifacts["model_checkpoint_id"] == checkpoint_id
    assert any(
        ref["kind"] == "model_checkpoint"
        and ref["id"] == checkpoint_id
        and ref["version"] == checkpoint_body["payload_hash"]
        for ref in artifacts["artifact_refs"]
    )
    assert _run(_checkpoint_authority(checkpoint_id))["origin"] == "analyst"

    assert _set_override(client, issuer_id, 2).status_code == 200
    restored = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints/{checkpoint_id}/restore",
        json={"expected_revision": 3},
    )
    assert restored.status_code == 200, restored.text
    assert restored.json()["revision"] == 4
    assert restored.json()["calculation_hash"] == original_hash
    assert restored.json()["payload"]["overrides"] == []
    assert restored.json()["calculation"]["periods"][0]["adjusted_ebitda"] == 110.0
    history = client.get(f"/api/models/v2/{issuer_id}/history").json()
    assert history[0]["action"] == "restore"
    assert history[0]["source"] == checkpoint_id
    replay_restore = client.post(
        f"/api/models/v2/{issuer_id}/history/{history[0]['id']}/replay",
        json={"expected_revision": 4, "mode": "undo"},
    )
    assert replay_restore.status_code == 422

    _as(other)
    hidden = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints/{checkpoint_id}/restore",
        json={"expected_revision": 0},
    )
    assert hidden.status_code == 404
    assert hidden.json() == {"detail": "Model checkpoint not found."}


def test_empty_checkpoint_is_analyst_authority_not_live(client):
    owner = f"model-v2-empty-checkpoint-{uuid.uuid4().hex[:10]}"
    seeded = _run(_seed_case(owner, context=True))
    issuer_id = str(seeded["issuer_id"])
    context_id = str(seeded["context_id"])
    _as(owner)
    empty_payload = {
        "schema_version": 2,
        "reporting_currency": "USD",
        "reporting_unit": "millions",
        "periods": [],
        "debt_instruments": [],
        "overrides": [],
        "source_ids": [],
    }
    saved = client.put(
        f"/api/models/v2/{issuer_id}",
        json={
            "expected_revision": 0,
            "context_id": context_id,
            "payload": empty_payload,
        },
    )
    assert saved.status_code == 200, saved.text
    assert saved.json()["calculation"]["status"] == "insufficient_inputs"

    checkpoint = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints",
        json={
            "context_id": context_id,
            "label": "Empty analyst model",
            "expected_revision": 1,
            "calculation_hash": saved.json()["calculation_hash"],
        },
    )

    assert checkpoint.status_code == 201, checkpoint.text
    authority = _run(_checkpoint_authority(checkpoint.json()["id"]))
    assert authority["origin"] == "analyst"
    assert authority["model_input_origins"] == []


def test_checkpoint_reservation_serializes_a_concurrent_override(
    client, monkeypatch
):
    owner = f"model-v2-checkpoint-race-{uuid.uuid4().hex[:10]}"
    seeded = _run(_seed_case(owner, context=True))
    issuer_id = str(seeded["issuer_id"])
    context_id = str(seeded["context_id"])
    _as(owner)
    created = _save(client, issuer_id, context_id=context_id)
    assert created.status_code == 200, created.text

    from routes import model_v2 as route

    reserved = threading.Event()
    release_checkpoint = threading.Event()
    override_started = threading.Event()
    original_reserve = route._reserve_draft_for_checkpoint
    original_cas = route._cas_update

    async def hold_after_reservation(*args, **kwargs):
        result = await original_reserve(*args, **kwargs)
        reserved.set()
        released = await asyncio.to_thread(release_checkpoint.wait, 5)
        assert released
        return result

    async def observe_override_cas(*args, **kwargs):
        override_started.set()
        return await original_cas(*args, **kwargs)

    with monkeypatch.context() as patcher:
        patcher.setattr(route, "_reserve_draft_for_checkpoint", hold_after_reservation)
        patcher.setattr(route, "_cas_update", observe_override_cas)
        with ThreadPoolExecutor(max_workers=2) as pool:
            checkpoint_future = pool.submit(
                client.post,
                f"/api/models/v2/{issuer_id}/checkpoints",
                json={
                    "context_id": context_id,
                    "label": "Concurrent base case",
                    "expected_revision": 1,
                    "calculation_hash": created.json()["calculation_hash"],
                },
            )
            assert reserved.wait(timeout=5)
            override_future = pool.submit(_set_override, client, issuer_id, 1)
            assert override_started.wait(timeout=5)
            assert not override_future.done()
            release_checkpoint.set()
            checkpoint = checkpoint_future.result(timeout=10)
            override = override_future.result(timeout=10)

    assert checkpoint.status_code == 201, checkpoint.text
    assert checkpoint.json()["draft_revision"] == 2
    assert override.status_code == 409, override.text
    assert override.json()["detail"]["current_revision"] == 2
    artifacts = _run(_context_artifacts(context_id))
    assert artifacts["model_checkpoint_id"] == checkpoint.json()["id"]
    current = client.get(f"/api/models/v2/{issuer_id}").json()["record"]
    assert current["revision"] == 2
    assert current["payload"]["overrides"] == []


def test_checkpoint_cannot_be_relabelled_to_an_unrelated_owned_run(client):
    owner = f"model-v2-checkpoint-run-{uuid.uuid4().hex[:10]}"
    seeded = _run(_seed_case(owner, context=True, cp1=True))
    issuer_id = str(seeded["issuer_id"])
    context_id = str(seeded["context_id"])
    source_run_id = str(seeded["run_id"])
    _as(owner)
    suggested = client.get(f"/api/models/v2/{issuer_id}").json()
    saved = client.put(
        f"/api/models/v2/{issuer_id}",
        json={
            "expected_revision": 0,
            "payload": suggested["suggested_payload"],
            "context_id": context_id,
            "source_run_id": source_run_id,
        },
    )
    assert saved.status_code == 200, saved.text
    unrelated_run_id = _run(_add_owned_run(owner, issuer_id))

    mismatch = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints",
        json={
            "context_id": context_id,
            "label": "Wrong run",
            "issuer_run_id": unrelated_run_id,
            "expected_revision": 1,
            "calculation_hash": saved.json()["calculation_hash"],
        },
    )
    assert mismatch.status_code == 409

    exact = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints",
        json={
            "context_id": context_id,
            "label": "Exact run",
            "expected_revision": 1,
            "calculation_hash": saved.json()["calculation_hash"],
        },
    )
    assert exact.status_code == 201, exact.text
    assert exact.json()["issuer_run_id"] == source_run_id
    assert _run(_context_artifacts(context_id))["issuer_run_id"] == source_run_id


def test_restoring_runless_checkpoint_clears_a_later_live_binding(client):
    owner = f"model-v2-checkpoint-clear-{uuid.uuid4().hex[:10]}"
    seeded = _run(_seed_case(owner, context=True))
    issuer_id = str(seeded["issuer_id"])
    context_id = str(seeded["context_id"])
    _as(owner)
    saved = _save(client, issuer_id, context_id=context_id)
    assert saved.status_code == 200, saved.text
    checkpoint = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints",
        json={
            "context_id": context_id,
            "label": "Imported authority",
            "expected_revision": 1,
            "calculation_hash": saved.json()["calculation_hash"],
        },
    )
    assert checkpoint.status_code == 201, checkpoint.text
    assert checkpoint.json()["issuer_run_id"] is None
    later_run_id = _run(_add_owned_run(owner, issuer_id))
    _run(_bind_draft_to_run(issuer_id, owner, later_run_id))

    restored = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints/{checkpoint.json()['id']}/restore",
        json={"expected_revision": 2},
    )

    assert restored.status_code == 200, restored.text
    assert restored.json()["source_run_id"] is None


def test_unknown_override_target_is_422_without_revision_change(client):
    owner = f"model-v2-target-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner)
    created = _save(client, issuer_id)
    assert created.status_code == 200
    unknown = client.post(
        f"/api/models/v2/{issuer_id}/overrides",
        json={
            "expected_revision": 1,
            "action": "set",
            "override": {
                "node_id": "calc:FY2026:not_a_node",
                "value_type": "number",
                "value": 1.0,
                "reason": "Invalid target fixture",
                "source": "test-source",
                "expires_at": (
                    datetime.now(timezone.utc) + timedelta(days=30)
                ).isoformat(),
            },
        },
    )
    assert unknown.status_code == 422
    current = client.get(f"/api/models/v2/{issuer_id}").json()["record"]
    assert current["revision"] == 1
    assert current["payload"]["overrides"] == []
    assert client.get(f"/api/models/v2/{issuer_id}/history").json() == []


def test_checkpoint_rejects_stale_saved_calculation(client):
    owner = f"model-v2-checkpoint-stale-{uuid.uuid4().hex[:10]}"
    seeded = _run(_seed_case(owner, context=True))
    issuer_id = str(seeded["issuer_id"])
    context_id = str(seeded["context_id"])
    _as(owner)
    created = _save(client, issuer_id, context_id=context_id)
    assert created.status_code == 200
    stale_hash = _run(_tamper_saved_calculation(issuer_id, owner))
    read = client.get(f"/api/models/v2/{issuer_id}").json()
    assert read["requires_recalculation"] is True
    assert read["current_calculation"]["calculation_hash"] == created.json()["calculation_hash"]

    response = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints",
        json={
            "context_id": context_id,
            "label": "Must not freeze stale values",
            "expected_revision": 1,
            "calculation_hash": stale_hash,
        },
    )
    assert response.status_code == 409
    detail = response.json()["detail"]
    assert detail["message"].startswith("Model calculation is stale")
    assert detail["current_calculation_hash"] == created.json()["calculation_hash"]
    assert client.get(f"/api/models/v2/{issuer_id}/checkpoints").json() == []


def test_restore_rejects_tampered_checkpoint_envelope(client):
    owner = f"model-v2-restore-tamper-{uuid.uuid4().hex[:10]}"
    seeded = _run(_seed_case(owner, context=True))
    issuer_id = str(seeded["issuer_id"])
    context_id = str(seeded["context_id"])
    _as(owner)
    created = _save(client, issuer_id, context_id=context_id)
    checkpoint = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints",
        json={
            "context_id": context_id,
            "expected_revision": 1,
            "calculation_hash": created.json()["calculation_hash"],
        },
    )
    assert checkpoint.status_code == 201, checkpoint.text
    checkpoint_id = checkpoint.json()["id"]
    _run(_tamper_checkpoint_payload_hash(checkpoint_id))

    response = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints/{checkpoint_id}/restore",
        json={"expected_revision": 2},
    )
    assert response.status_code == 409
    current = client.get(f"/api/models/v2/{issuer_id}").json()["record"]
    assert current["revision"] == 2


def test_audit_failure_rolls_back_cas_update(client, monkeypatch):
    owner = f"model-v2-rollback-{uuid.uuid4().hex[:10]}"
    issuer_id = str(_run(_seed_case(owner))["issuer_id"])
    _as(owner)
    created = _save(client, issuer_id)
    assert created.status_code == 200
    original_hash = created.json()["calculation_hash"]

    from routes import model_v2 as route

    def fail_audit(*_args, **_kwargs):
        raise RuntimeError("forced model override audit failure")

    with monkeypatch.context() as patcher:
        patcher.setattr(route, "ModelOverrideEvent", fail_audit)
        with pytest.raises(RuntimeError, match="forced model override audit failure"):
            _set_override(client, issuer_id, 1)

    current = client.get(f"/api/models/v2/{issuer_id}").json()["record"]
    assert current["revision"] == 1
    assert current["calculation_hash"] == original_hash
    assert current["payload"]["overrides"] == []
    assert client.get(f"/api/models/v2/{issuer_id}/history").json() == []
