from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta, timezone
from io import BytesIO
from types import SimpleNamespace
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from model_engine_v2 import (
    CellOverride,
    DebtInstrument,
    DebtPeriod,
    ModelAuthority,
    ModelDraftPayload,
    ModelPeriodInput,
    calculate_model,
)


def _payload(
    run_id: str,
    *,
    derived_expires_at: datetime | None = None,
) -> ModelDraftPayload:
    authority = ModelAuthority(origin="live", method="run", source_ids=[run_id])
    return ModelDraftPayload(
        reporting_currency="GBP",
        reporting_unit="millions",
        periods=[ModelPeriodInput(
            period_key="FY2026",
            label="FY26e",
            kind="forecast",
            revenue=800,
            reported_ebitda=100,
            adjustments=10,
            cash=20,
            taxes=5,
            capex=10,
            working_capital_change=-2,
            other_cash_flow=0,
            authority=authority,
        )],
        debt_instruments=[DebtInstrument(
            instrument_id="tlb-1",
            name="First-lien term loan",
            priority=1,
            seniority="1L",
            currency="GBP",
            rate_type="floating",
            authority=authority,
            periods=[DebtPeriod(
                period_key="FY2026",
                opening_balance=200,
                closing_balance=190,
                draws=0,
                repayments=10,
                scheduled_amortization=0,
                commitment=200,
                benchmark_rate=0.05,
                spread_rate=0.03,
                commitment_fee_rate=0,
                pik_rate=0,
                cash_fees=1,
                hedge_effect=0,
                fx_rate=1,
            )],
        )],
        overrides=(
            [CellOverride(
                node_id="calc:FY2026:adjusted_ebitda",
                value_type="number",
                value=110,
                reason="Committee normalization",
                source="analyst-test",
                expires_at=derived_expires_at,
            )]
            if derived_expires_at is not None
            else []
        ),
        source_ids=[run_id],
    )


@pytest.mark.asyncio
async def test_v2_checkpoint_report_and_xlsx_preserve_exact_calculation_identity(
    seeded_db, monkeypatch
):
    from database import (
        AnalysisContextRecord,
        AsyncSessionLocal,
        ModelCheckpoint,
        Run,
        SourceManifest,
    )
    from identity import CallerIdentity, get_identity
    from analysis_contracts import ArtifactRef
    from lineage_service import write_lineage_edge
    from main import app
    from routes import reports

    analyst_id = f"model-report-{uuid4().hex}"
    issuer_id = "a71f0000-0000-0000-0000-000000000001"
    context_id = str(uuid4())
    run_id = str(uuid4())
    checkpoint_id = str(uuid4())
    manifest_id = str(uuid4())
    created_at = datetime(2026, 7, 14, 12, tzinfo=timezone.utc)
    override_expiry = created_at + timedelta(hours=1)
    draft = _payload(run_id, derived_expires_at=override_expiry)
    calculation = calculate_model(draft, evaluated_at=created_at)
    frozen = {
        "version": 2,
        "draft_id": str(uuid4()),
        "draft_revision": 3,
        "payload": draft.model_dump(mode="json"),
        "calculation": calculation.model_dump(mode="json"),
    }
    canonical = json.dumps(
        frozen, sort_keys=True, separators=(",", ":"), allow_nan=False, default=str
    )
    async with AsyncSessionLocal() as db:
        db.add_all([
                Run(
                id=run_id,
                issuer_id=issuer_id,
                analyst_id=analyst_id,
                status="complete",
                qa_status="Passed",
                committee_status="Committee Ready",
                as_of_date="2026-06-30",
                    completed_at=created_at,
                    input_manifest_ids=[manifest_id],
                    input_document_ids=[],
                    input_corpus_sha256="f" * 64,
                    input_snapshot_state="approved",
            ),
            SourceManifest(
                id=manifest_id,
                analyst_id=analyst_id,
                issuer_id=issuer_id,
                origin="live",
                method="upload",
                status="ready",
                    files=[{
                        "document_id": "report-source-document",
                        "sha256": "e" * 64,
                        "malware_scan": "clean",
                    }],
                    authority={"approval_state": "ratified"},
                created_at=created_at,
            ),
            AnalysisContextRecord(
                id=context_id,
                analyst_id=analyst_id,
                name="Exact model report",
                issuer_ids=[issuer_id],
                artifacts={
                    "model_checkpoint_id": checkpoint_id,
                    "source_manifest_id": manifest_id,
                    "issuer_run_id": run_id,
                },
                created_at=created_at,
                updated_at=created_at,
            ),
        ])
        await db.flush()
        db.add(ModelCheckpoint(
            id=checkpoint_id,
            issuer_id=issuer_id,
            analyst_id=analyst_id,
            context_id=context_id,
            issuer_run_id=run_id,
            label="Exact v2 checkpoint",
            payload_hash=hashlib.sha256(canonical.encode("utf-8")).hexdigest(),
            payload=frozen,
            authority={
                "origin": "imported",
                "model_input_origins": ["analyst", "imported"],
                "analyst_override": True,
            },
            engine_version=calculation.engine_version,
            source_fingerprint=calculation.source_fingerprint,
            input_fingerprint=calculation.input_fingerprint,
            calculation_hash=calculation.calculation_hash,
            draft_revision=3,
            created_at=created_at,
        ))
        await write_lineage_edge(
            db,
            context_id=context_id,
            analyst_id=analyst_id,
            artifact=ArtifactRef(kind="issuer_run", id=run_id),
            parent=ArtifactRef(kind="source_manifest", id=manifest_id),
            transform="run-creation",
            transform_version="2",
            enabled=True,
        )
        await db.commit()

    model_flag = {"enabled": True}
    monkeypatch.setattr(
        reports,
        "get_settings",
        lambda: SimpleNamespace(
            caos_lineage_v2_enabled=True,
            caos_model_engine_v2_enabled=model_flag["enabled"],
        ),
    )
    report_clock = {"now": created_at + timedelta(minutes=30)}
    monkeypatch.setattr(reports, "_report_event_time", lambda: report_clock["now"])
    try:
        with TestClient(app) as client:
            app.dependency_overrides[get_identity] = lambda: CallerIdentity(
                id=analyst_id,
                email="model-report@firm.test",
                full_name="Model Report Owner",
                role="viewer",
                source="profile",
            )
            denied = client.post("/api/reports/versions", json={
                "context_id": context_id,
                "run_id": run_id,
                "model_checkpoint_id": checkpoint_id,
                "payload": {"title": "Viewer must not publish"},
            })
            assert denied.status_code == 403
            denied_draft = client.put(f"/api/reports/drafts/{context_id}", json={
                "payload": {"active_id": "live-committee-pack"},
            })
            assert denied_draft.status_code == 403

            app.dependency_overrides[get_identity] = lambda: CallerIdentity(
                id=analyst_id,
                email="model-report@firm.test",
                full_name="Model Report Owner",
                role="analyst",
                source="profile",
            )
            stale_run_preview = client.post("/api/reports/versions/preview", json={
                "context_id": context_id,
                "run_id": run_id,
                "model_checkpoint_id": checkpoint_id,
                "payload": {
                    "deliverable_id": "live-committee-pack",
                    "source_run_id": "old-report-run",
                    "omit": {}, "edits": {}, "show_sources": True,
                    "hide_addbacks": False,
                },
            })
            assert stale_run_preview.status_code == 409
            assert "different run" in stale_run_preview.json()["detail"]
            opaque_old_version = client.post("/api/reports/versions/preview", json={
                "context_id": context_id,
                "run_id": run_id,
                "model_checkpoint_id": checkpoint_id,
                "payload": {
                    "deliverable_id": "live-committee-pack",
                    "source_run_id": run_id,
                    "rendered_report": {"id": "old-version", "sections": []},
                },
            })
            assert opaque_old_version.status_code == 422
            published = client.post("/api/reports/versions", json={
                "context_id": context_id,
                "run_id": run_id,
                "model_checkpoint_id": checkpoint_id,
                "payload": {
                    "deliverable_id": "live-committee-pack",
                    "source_run_id": run_id,
                    "omit": {},
                    "edits": {"s0.r3.v": "Committee Ready · analyst reviewed"},
                    "show_sources": True,
                    "hide_addbacks": False,
                },
            })
            assert published.status_code == 409
            assert "server-frozen preview" in str(published.json()["detail"])
            preview = client.post("/api/reports/versions/preview", json={
                "context_id": context_id,
                "run_id": run_id,
                "model_checkpoint_id": checkpoint_id,
                "payload": {
                    "deliverable_id": "live-committee-pack",
                    "source_run_id": run_id,
                    "omit": {},
                    "edits": {"s0.r3.v": "Committee Ready · analyst reviewed"},
                    "show_sources": True,
                    "hide_addbacks": False,
                },
            })
            assert preview.status_code == 200, preview.text
            preview_body = preview.json()
            assert preview_body["document_sha256"] == preview_body["preview_sha256"]
            assert preview_body["model_calculation_hash"] == calculation.calculation_hash
            assert preview_body["payload"]["composition"]["reviewed_report"]["sections"][0]["rows"][3][1] == "Committee Ready · analyst reviewed"
            published = client.post("/api/reports/versions", json={
                "context_id": context_id,
                "run_id": run_id,
                "model_checkpoint_id": checkpoint_id,
                "payload": preview_body["payload"]["composition"]["editorial"],
                "preview_sha256": preview_body["preview_sha256"],
            })
            assert published.status_code == 201, published.text
            version = published.json()
            assert version["model_engine_version"] == calculation.engine_version
            assert version["model_source_fingerprint"] == calculation.source_fingerprint
            assert version["model_input_fingerprint"] == calculation.input_fingerprint
            assert version["model_calculation_hash"] == calculation.calculation_hash
            assert version["model_draft_revision"] == 3
            assert version["authority"]["analyst_override"] is True
            assert version["authority"]["model_origin"] == "imported"
            assert version["authority"]["model_input_origins"] == ["analyst", "imported"]
            assert version["authority"]["model_analyst_override"] is True
            frozen_model = version["payload"]["model"]
            assert frozen_model["calculation"] == calculation.model_dump(mode="json")
            assert frozen_model["payload"]["reporting_currency"] == "GBP"
            assert frozen_model["payload"]["reporting_unit"] == "millions"
            sources_hidden_preview = client.post("/api/reports/versions/preview", json={
                "context_id": context_id,
                "run_id": run_id,
                "model_checkpoint_id": checkpoint_id,
                "payload": {
                    "deliverable_id": "live-committee-pack",
                    "source_run_id": run_id,
                    "omit": {},
                    "edits": {},
                    "show_sources": False,
                    "hide_addbacks": False,
                },
            })
            assert sources_hidden_preview.status_code == 200, sources_hidden_preview.text
            sources_hidden_body = sources_hidden_preview.json()
            assert sources_hidden_body["authority"]["analyst_override"] is True
            sources_hidden_version = client.post("/api/reports/versions", json={
                "context_id": context_id,
                "run_id": run_id,
                "model_checkpoint_id": checkpoint_id,
                "payload": sources_hidden_body["payload"]["composition"]["editorial"],
                "preview_sha256": sources_hidden_body["preview_sha256"],
            })
            assert sources_hidden_version.status_code == 201, sources_hidden_version.text
            assert sources_hidden_version.json()["authority"]["analyst_override"] is True
            summaries = client.get("/api/reports/versions", params={
                "context_id": context_id, "limit": 1,
            })
            assert summaries.status_code == 200
            assert "payload" not in summaries.json()[0]
            exact_version = client.get(f"/api/reports/versions/{version['id']}")
            assert exact_version.status_code == 200
            assert exact_version.json()["payload"]["model"]["calculation_hash"] == calculation.calculation_hash

            # The checkpoint was valid and publishable before the derived
            # replacement expired. Once it expires, neither a new preview nor
            # a previously issued frozen preview may create another version.
            report_clock["now"] = override_expiry
            expired_preview = client.post("/api/reports/versions/preview", json={
                "context_id": context_id,
                "run_id": run_id,
                "model_checkpoint_id": checkpoint_id,
                "payload": preview_body["payload"]["composition"]["editorial"],
            })
            assert expired_preview.status_code == 409
            assert "effective override" in expired_preview.json()["detail"]
            assert "new checkpoint" in expired_preview.json()["detail"]
            expired_publish = client.post("/api/reports/versions", json={
                "context_id": context_id,
                "run_id": run_id,
                "model_checkpoint_id": checkpoint_id,
                "payload": preview_body["payload"]["composition"]["editorial"],
                "preview_sha256": preview_body["preview_sha256"],
            })
            assert expired_publish.status_code == 409
            assert "effective override" in expired_publish.json()["detail"]

            # Expiry is a gate on new publications, not a mutation of an
            # already-published immutable version.
            published_after_expiry = client.get(
                f"/api/reports/versions/{version['id']}"
            )
            assert published_after_expiry.status_code == 200
            assert (
                published_after_expiry.json()["document_sha256"]
                == version["document_sha256"]
            )
            assert published_after_expiry.json()["payload"] == version["payload"]

            model_flag["enabled"] = False
            rollback_publish = client.post("/api/reports/versions", json={
                "context_id": context_id,
                "run_id": run_id,
                "model_checkpoint_id": checkpoint_id,
                "payload": preview_body["payload"]["composition"]["editorial"],
                "preview_sha256": preview_body["preview_sha256"],
            })
            assert rollback_publish.status_code == 409
            assert "cannot be republished" in rollback_publish.json()["detail"]

            exported = client.post(
                f"/api/reports/versions/{version['id']}/export?format=xlsx"
            )
            assert exported.status_code == 200, exported.text
            workbook = load_workbook(BytesIO(exported.content), read_only=True)
            row = next(workbook["Model"].iter_rows(min_row=2, values_only=True))
            headers = next(
                workbook["Model"].iter_rows(max_row=1, values_only=True)
            )
            cover = {
                key: value
                for key, value in workbook["Cover"].iter_rows(
                    min_row=4, max_col=2, values_only=True
                )
                if key
            }
            assert row[0] == "FY2026"
            assert row[4] == 110
            assert row[5] == pytest.approx(16.6)
            assert row[6] == 190
            assert row[9] == pytest.approx(190 / 110)
            assert headers[3] == "Revenue (GBP millions)"
            assert workbook["Cover"]["B24"].value == calculation.calculation_hash
            assert cover["Model reporting currency"] == "GBP"
            assert cover["Model reporting unit"] == "millions"
            debt_row = next(
                workbook["Debt Schedule"].iter_rows(
                    min_row=2, values_only=True
                )
            )
            assert debt_row[2] == "GBP"
            assert debt_row[14] == 190
            assert debt_row[15] == 0
            workbook.close()
        async with AsyncSessionLocal() as db:
            context = await db.get(AnalysisContextRecord, context_id)
            assert context is not None
            refs = [
                ref for ref in (context.artifacts or {}).get("artifact_refs", [])
                if ref.get("kind") == "model_checkpoint"
                and ref.get("id") == checkpoint_id
            ]
            assert refs == [{
                "kind": "model_checkpoint",
                "id": checkpoint_id,
                "version": hashlib.sha256(canonical.encode("utf-8")).hexdigest(),
            }]
        rebound_manifest_id = str(uuid4())
        async with AsyncSessionLocal() as db:
            db.add(SourceManifest(
                id=rebound_manifest_id,
                analyst_id=analyst_id,
                issuer_id=issuer_id,
                origin="live",
                method="upload",
                status="ready",
                files=[],
                authority={},
                created_at=created_at,
            ))
            context = await db.get(AnalysisContextRecord, context_id)
            assert context is not None
            context.artifacts = {
                **(context.artifacts or {}),
                "source_manifest_id": rebound_manifest_id,
            }
            await db.commit()
        model_flag["enabled"] = True
        with TestClient(app) as client:
            rebound = client.post("/api/reports/versions/preview", json={
                "context_id": context_id,
                "run_id": run_id,
                "model_checkpoint_id": checkpoint_id,
                "payload": {
                    "deliverable_id": "live-committee-pack",
                    "source_run_id": run_id,
                    "omit": {}, "edits": {}, "show_sources": True,
                    "hide_addbacks": False,
                },
            })
            assert rebound.status_code == 409
            assert "fully approved immutable input snapshot" in rebound.json()["detail"]
            app.dependency_overrides[get_identity] = lambda: CallerIdentity(
                id="foreign-report-analyst",
                email="foreign@firm.test",
                full_name="Foreign",
                role="analyst",
                source="profile",
            )
            assert client.get(f"/api/reports/versions/{version['id']}").status_code == 404
    finally:
        app.dependency_overrides.clear()


@pytest.mark.asyncio
async def test_v2_publication_rejects_inconsistent_checkpoint_identity(
    seeded_db, monkeypatch
):
    from database import ModelCheckpoint
    from routes import reports

    draft = _payload("run-tamper")
    calculation = calculate_model(draft)
    checkpoint = ModelCheckpoint(
        issuer_id="issuer",
        analyst_id="analyst",
        context_id="context",
        label="tampered",
        payload_hash="a" * 64,
        payload={
            "version": 2,
            "draft_revision": 1,
            "payload": draft.model_dump(mode="json"),
            "calculation": calculation.model_dump(mode="json"),
        },
        authority={},
        engine_version=calculation.engine_version,
        source_fingerprint=calculation.source_fingerprint,
        input_fingerprint=calculation.input_fingerprint,
        calculation_hash="0" * 64,
        draft_revision=1,
        created_at=datetime.now(timezone.utc),
    )
    with pytest.raises(Exception) as caught:
        reports._model_v2_checkpoint_snapshot(checkpoint)
    assert getattr(caught.value, "status_code", None) == 409


def test_report_expiry_gate_ignores_override_inactive_at_checkpoint() -> None:
    from routes import reports

    expires_at = datetime(2026, 7, 14, 12, tzinfo=timezone.utc)
    draft = _payload("inactive-override-run", derived_expires_at=expires_at)
    calculation = calculate_model(
        draft,
        evaluated_at=expires_at + timedelta(minutes=1),
    )
    adjusted_node = next(
        node
        for node in calculation.periods[0].nodes
        if node.node_id == "calc:FY2026:adjusted_ebitda"
    )
    assert adjusted_node.overridden is False

    reports._require_unexpired_effective_overrides(
        {
            "payload": draft.model_dump(mode="json"),
            "calculation": calculation.model_dump(mode="json"),
        },
        evaluated_at=expires_at + timedelta(days=1),
    )


def test_report_expiry_gate_includes_effective_input_overrides() -> None:
    from routes import reports

    expires_at = datetime(2026, 7, 14, 12, tzinfo=timezone.utc)
    draft = _payload("input-override-run")
    draft.overrides = [CellOverride(
        node_id="input:FY2026:cash",
        value_type="number",
        value=25,
        reason="Treasury cash update",
        source="analyst-test",
        expires_at=expires_at,
    )]
    calculation = calculate_model(
        draft,
        evaluated_at=expires_at - timedelta(minutes=1),
    )
    cash_node = next(
        node
        for node in calculation.periods[0].nodes
        if node.node_id == "input:FY2026:cash"
    )
    assert cash_node.overridden is True

    with pytest.raises(Exception) as expired:
        reports._require_unexpired_effective_overrides(
            {
                "payload": draft.model_dump(mode="json"),
                "calculation": calculation.model_dump(mode="json"),
            },
            evaluated_at=expires_at,
        )
    assert getattr(expired.value, "status_code", None) == 409
    assert "effective override" in str(getattr(expired.value, "detail", ""))


def test_report_composition_is_bounded_and_rejects_nonfinite_json() -> None:
    from routes import reports

    with pytest.raises(Exception) as nonfinite:
        reports._bounded_composition({"value": float("nan")})
    assert getattr(nonfinite.value, "status_code", None) == 422

    with pytest.raises(Exception) as oversized:
        reports._bounded_composition({"text": "x" * 1_000_001})
    assert getattr(oversized.value, "status_code", None) == 413

    frozen = reports._bounded_composition(
        {"model": "x" * 1_100_000},
        label="Report version",
        max_bytes=reports._MAX_REPORT_VERSION_BYTES,
    )
    assert len(frozen) > 1_000_000

    # The public 2,000 instrument-period contract expands to roughly 18.4 MB
    # once the audited node graph is frozen. Report publication must not reject
    # an otherwise valid maximum-contract checkpoint behind a smaller cap.
    maximum_contract = reports._bounded_composition(
        {"model": "x" * 18_400_000},
        label="Report version",
        max_bytes=reports._MAX_REPORT_VERSION_BYTES,
    )
    assert len(maximum_contract) > 18_000_000


def test_server_preview_paths_edit_and_omit_the_same_multi_module_dsl() -> None:
    from report_composition import ReportCompositionIntent, materialize_reviewed_report

    document = {
        "run_id": "run-1",
        "issuer_id": "issuer-1",
        "as_of_date": "2026-06-30",
        "committee_status": "Committee Ready",
        "sections": [
            {
                "module_id": "CP-1", "module_name": "Foundation",
                "qa_status": "Passed", "confidence": "High",
                "summary": {"headline": "Original CP-1"},
            },
            {
                "module_id": "CP-2", "module_name": "Business risk",
                "qa_status": "Passed", "confidence": "High",
                "summary": {"headline": "Original CP-2"},
            },
        ],
    }
    reviewed = materialize_reviewed_report(document, ReportCompositionIntent(
        source_run_id="run-1",
        omit={"1": True},
        edits={"s2.r0.c1": "Analyst-reviewed CP-2"},
    ))
    assert [section["title"] for section in reviewed["sections"]] == [
        "FROZEN ANALYSIS ENVELOPE",
        "CP-2 · Business risk",
    ]
    assert reviewed["sections"][1]["rows"][0]["cells"] == [
        "headline", "Analyst-reviewed CP-2",
    ]
