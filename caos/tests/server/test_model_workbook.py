from __future__ import annotations

import io
import zipfile
from datetime import datetime, timezone

import pytest
from fastapi import HTTPException, UploadFile
from openpyxl import Workbook, load_workbook
from pydantic import ValidationError

import model_workbook
from model_engine_v2 import (
    CellOverride,
    DebtInstrument,
    DebtPeriod,
    ModelAuthority,
    ModelDraftPayload,
    ModelPeriodInput,
    ModelUiPreferences,
)
from model_workbook import (
    SHEET_NAMES,
    ModelWorkbookError,
    parse_mapping,
    preview_workbook,
    render_model_workbook,
)


EXPORTED_AT = datetime(2026, 7, 14, 12, 0, tzinfo=timezone.utc)


def _payload() -> ModelDraftPayload:
    authority = ModelAuthority(
        origin="live",
        method="=source-method",
        source_ids=["run-1"],
        as_of=datetime(2026, 6, 30, tzinfo=timezone.utc),
    )
    return ModelDraftPayload(
        reporting_currency="USD",
        reporting_unit="millions",
        periods=[ModelPeriodInput(
            period_key="FY2026",
            label="=HYPERLINK(\"https://invalid\")",
            kind="forecast",
            months=12,
            revenue=800,
            reported_ebitda=100,
            adjustments=10,
            adjusted_ebitda=110,
            cash=20,
            total_debt=190,
            net_debt=170,
            cash_interest=18.175,
            taxes=5,
            capex=10,
            working_capital_change=-2,
            other_cash_flow=0,
            authority=authority,
        )],
        debt_instruments=[DebtInstrument(
            instrument_id="tlb-1",
            name="@malicious-name",
            priority=1,
            seniority="1L",
            currency="USD",
            rate_type="hybrid",
            maturity="2030-06-30",
            benchmark_curve="SOFR",
            amortization="1% annual",
            sources=["manifest-1"],
            authority=authority,
            periods=[DebtPeriod(
                period_key="FY2026",
                opening_balance=200,
                closing_balance=190,
                draws=5,
                repayments=15,
                scheduled_amortization=1.95,
                commitment=220,
                benchmark_rate=0.04,
                floor_rate=0.05,
                spread_rate=0.03,
                coupon_rate=0.01,
                commitment_fee_rate=0.005,
                pik_rate=0.01,
                cash_fees=1,
                hedge_effect=-0.5,
                fx_rate=1,
            )],
        )],
        overrides=[CellOverride(
            node_id="calc:FY2026:adjusted_ebitda",
            value_type="number",
            value=105,
            reason="+IC adjustment",
            source="CP-6A",
            expires_at=datetime(2027, 1, 1, tzinfo=timezone.utc),
        ), CellOverride(
            node_id="input:FY2026:revenue",
            value_type="null",
            value=None,
            scope="scenario",
            source="analyst",
        )],
        ui_preferences=ModelUiPreferences(
            show_quarters=False,
            show_assumptions=False,
            show_scenarios=False,
            warn_on_unsaved_leave=False,
            collapsed_rows=["=unsafe-row"],
        ),
        source_ids=["run-1", "run-1"],
    )


def _export(payload: ModelDraftPayload | None = None) -> bytes:
    return render_model_workbook(
        payload or _payload(),
        issuer_id="issuer-1",
        draft_revision=7,
        exported_by="=analyst",
        exported_at=EXPORTED_AT,
    )


def _replace_part(content: bytes, part: str, old: bytes, new: bytes) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(content)) as source, zipfile.ZipFile(
        output, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == part:
                assert old in payload
                payload = payload.replace(old, new, 1)
            target.writestr(item, payload)
    return output.getvalue()


def _replace_all_in_part(content: bytes, part: str, old: bytes, new: bytes) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(content)) as source, zipfile.ZipFile(
        output, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == part:
                assert old in payload
                payload = payload.replace(old, new)
            target.writestr(item, payload)
    return output.getvalue()


def _add_part(content: bytes, part: str, payload: bytes) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(content)) as source, zipfile.ZipFile(
        output, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            target.writestr(item, source.read(item.filename))
        target.writestr(part, payload)
    return output.getvalue()


def test_export_reopens_exact_six_sheets_with_stable_keys_hashes_and_safe_text() -> None:
    payload = _payload()
    content = _export(payload)
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    assert tuple(workbook.sheetnames) == SHEET_NAMES
    assert all(
        cell.data_type != "f"
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    )
    assert workbook["Assumptions"]["A2"].value == "period:FY2026"
    assert workbook["Assumptions"]["B2"].value == "FY2026"
    assert workbook["Assumptions"]["C2"].value.startswith("'=")
    assert workbook["Debt Schedule"]["A2"].value == "debt:tlb-1:FY2026"
    assert workbook["Debt Schedule"]["C2"].value.startswith("'@")
    assert workbook["Debt Schedule"]["G2"].value == "hybrid"
    assert workbook["Overrides"]["A2"].value == "calc:FY2026:adjusted_ebitda"
    assert workbook["Overrides"]["E2"].value.startswith("'+")
    workbook.close()

    preview = preview_workbook(content, filename="model.xlsx")
    assert preview.blocking_count == 0
    assert preview.warning_count == 0
    assert preview.draft_payload is not None
    assert preview.draft_payload.model_dump(mode="json") == payload.model_dump(mode="json")
    assert preview.calculation is not None
    assert preview.embedded_hashes.source_fingerprint == preview.calculation.source_fingerprint
    assert preview.embedded_hashes.input_fingerprint == preview.calculation.input_fingerprint
    assert preview.embedded_hashes.calculation_hash == preview.calculation.calculation_hash


def test_maximum_source_audit_and_ui_preferences_round_trip_exactly() -> None:
    base = _payload()
    source_ids = [f"source-{index:04d}" for index in range(8_998)] + ["s" * 240]
    instruments: list[DebtInstrument] = []
    for index in range(90):
        start = index * 100
        stop = min(start + 100, len(source_ids))
        instruments.append(base.debt_instruments[0].model_copy(update={
            "instrument_id": f"tlb-{index:03d}",
            "name": f"Term loan {index:03d}",
            "sources": source_ids[start:stop],
        }))
    raw = base.model_dump(mode="json")
    raw["debt_instruments"] = [
        instrument.model_dump(mode="json") for instrument in instruments
    ]
    raw["ui_preferences"]["collapsed_rows"] = [
        f"model-row-{index:03d}-{'x' * 20}" for index in range(100)
    ]
    raw["source_ids"] = ["run-1"]
    payload = ModelDraftPayload.model_validate(raw)

    assert len({
        *payload.source_ids,
        *payload.periods[0].authority.source_ids,
        *(source for item in payload.debt_instruments for source in item.sources),
    }) == 9_000
    over_limit = payload.model_dump(mode="json")
    over_limit["source_ids"].append("one-source-too-many")
    with pytest.raises(ValidationError, match="source audit"):
        ModelDraftPayload.model_validate(over_limit)

    preview = preview_workbook(
        _export(payload),
        filename="maximum-public-contract.xlsx",
    )

    assert preview.blocking_count == 0
    assert preview.draft_payload is not None
    assert preview.draft_payload.model_dump(mode="json") == payload.model_dump(mode="json")


def test_optional_rate_type_roundtrips_as_an_explicit_incomplete_schedule() -> None:
    payload = _payload()
    payload.debt_instruments[0].rate_type = None

    preview = preview_workbook(
        _export(payload),
        filename="missing-rate-type.xlsx",
        evaluated_at=EXPORTED_AT,
    )

    assert preview.blocking_count == 0
    assert preview.draft_payload is not None
    assert preview.draft_payload.model_dump(mode="json") == payload.model_dump(mode="json")
    assert preview.calculation is not None
    assert preview.calculation.status == "partial"
    assert any("missing debt input rate_type" in gap for gap in preview.calculation.gaps)


def test_invalid_override_expiry_blocks_without_becoming_a_permanent_override() -> None:
    workbook = load_workbook(io.BytesIO(_export()))
    workbook["Overrides"]["H2"] = "not-an-iso-timestamp"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    preview = preview_workbook(
        output.getvalue(),
        filename="invalid-override-expiry.xlsx",
    )

    assert "invalid_override_expiry" in {issue.code for issue in preview.issues}
    assert preview.blocking_count > 0
    assert preview.draft_payload is None
    assert preview.calculation is None


@pytest.mark.asyncio
async def test_model_workbook_stream_cap_fails_before_antivirus_scan(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from routes import model_workbook as workbook_routes

    scanned = False

    async def scan(_: bytes) -> None:
        nonlocal scanned
        scanned = True

    monkeypatch.setattr(workbook_routes, "MAX_MODEL_FILE_BYTES", 4)
    monkeypatch.setattr(workbook_routes.avscan, "scan", scan)
    upload = UploadFile(filename="oversized.xlsx", file=io.BytesIO(b"12345"))

    with pytest.raises(HTTPException) as caught:
        await workbook_routes._preview_bytes(
            upload,
            mapping=None,
            evaluated_at=EXPORTED_AT,
        )

    assert caught.value.status_code == 413
    assert scanned is False


def test_debt_rate_type_and_numeric_bounds_fail_closed_on_import() -> None:
    workbook = load_workbook(io.BytesIO(_export()))
    workbook["Debt Schedule"]["G2"] = "variable"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    invalid_rate = preview_workbook(output.getvalue(), filename="invalid-rate-type.xlsx")
    assert "invalid_debt_instrument" in {issue.code for issue in invalid_rate.issues}

    workbook = load_workbook(io.BytesIO(_export()))
    workbook["Debt Schedule"]["L2"] = 1e308
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    oversized = preview_workbook(output.getvalue(), filename="oversized-debt.xlsx")
    assert "invalid_debt_period" in {issue.code for issue in oversized.issues}


def test_export_refuses_to_emit_a_workbook_that_violates_import_bounds(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(model_workbook, "MAX_ROWS_PER_SHEET", 1)
    with pytest.raises(ModelWorkbookError) as exc:
        _export()
    assert exc.value.code == "workbook_export_limit"


def test_strict_cached_formula_is_audited_but_never_promoted_to_authority() -> None:
    content = _export()
    content = _replace_part(
        content,
        "xl/worksheets/sheet3.xml",
        b'<c r="G2" t="n"><v>100</v></c>',
        b'<c r="G2"><f>50+50</f><v>100</v></c>',
    )
    preview = preview_workbook(content, filename="cached-formula.xlsx")
    assert preview.blocking_count > 0
    assert preview.draft_payload is None
    assert preview.calculation is None
    assert "imported_formula_not_authoritative" in {
        issue.code for issue in preview.issues
    }
    assert len(preview.formula_audit) == 1
    assert preview.formula_audit[0].formula == "=50+50"
    assert preview.formula_audit[0].cached_value == 100
    assert preview.formula_audit[0].disposition == "comparison_only"


@pytest.mark.parametrize(
    ("formula", "code"),
    [
        ("50+[Book2.xlsx]Sheet1!A1", "external_formula"),
        ("NOW()", "volatile_formula"),
        ('RTD("malicious.prog", "", "topic")', "active_formula"),
        ("cmd|'/C calc'!A0", "active_formula"),
        ("G2", "circular_formula"),
        ("'aSsUmPtIoNs'!G2", "circular_formula"),
    ],
)
def test_unsafe_formulas_fail_closed_even_with_finite_cache(formula: str, code: str) -> None:
    content = _export()
    content = _replace_part(
        content,
        "xl/worksheets/sheet3.xml",
        b'<c r="G2" t="n"><v>100</v></c>',
        f'<c r="G2"><f>{formula}</f><v>100</v></c>'.encode(),
    )
    preview = preview_workbook(content, filename="unsafe-formula.xlsx")
    assert preview.blocking_count >= 1
    assert code in {issue.code for issue in preview.issues}
    assert code in preview.formula_audit[0].blocking_codes
    assert preview.draft_payload is None
    assert preview.calculation is None


def test_formula_without_finite_cache_is_blocking_and_never_calculated() -> None:
    content = _export()
    content = _replace_part(
        content,
        "xl/worksheets/sheet3.xml",
        b'<c r="G2" t="n"><v>100</v></c>',
        b'<c r="G2"><f>50+50</f><v></v></c>',
    )
    preview = preview_workbook(content, filename="stale-cache.xlsx")
    assert "formula_cache_required" in {issue.code for issue in preview.issues}
    assert preview.formula_audit[0].cached_value is None
    assert preview.blocking_count > 0
    assert preview.draft_payload is None
    assert preview.calculation is None


def test_strict_preview_blocks_tampered_model_tieout_and_hash() -> None:
    content = _export()
    # Model adjusted EBITDA is the fourth metric row for FY2026 (row 5).
    content = _replace_part(
        content,
        "xl/worksheets/sheet2.xml",
        b'<c r="E5" t="n"><v>105</v></c>',
        b'<c r="E5" t="n"><v>999</v></c>',
    )
    preview = preview_workbook(content, filename="tampered-model.xlsx")
    assert "model_tieout_mismatch" in {issue.code for issue in preview.issues}

    hash_tampered = _replace_part(
        _export(),
        "xl/worksheets/sheet1.xml",
        preview.embedded_hashes.calculation_hash.encode(),
        ("0" * 64).encode(),
    )
    hash_preview = preview_workbook(hash_tampered, filename="tampered-hash.xlsx")
    assert "metadata_mismatch" in {issue.code for issue in hash_preview.issues}
    assert "fingerprint_mismatch" in {issue.code for issue in hash_preview.issues}


def test_strict_preview_ties_model_audit_fields_and_sources_to_engine() -> None:
    workbook = load_workbook(io.BytesIO(_export()))
    workbook["Model"]["G5"] = "fabricated formula"
    source_sheet = workbook["Sources - Audit"]
    authority_row = next(
        row
        for row in range(2, source_sheet.max_row + 1)
        if source_sheet.cell(row, 1).value == "authority_source"
    )
    source_sheet.cell(authority_row, 3, "tampered-source")
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    preview = preview_workbook(output.getvalue(), filename="tampered-audit.xlsx")
    codes = {issue.code for issue in preview.issues}
    assert "model_tieout_mismatch" in codes
    assert "source_audit_mismatch" in codes


def test_strict_identity_metadata_is_typed_and_exposed() -> None:
    workbook = load_workbook(io.BytesIO(_export()))
    for sheet_name in ("Cover", "Sources - Audit"):
        worksheet = workbook[sheet_name]
        key_column = 1 if sheet_name == "Cover" else 2
        value_column = 2 if sheet_name == "Cover" else 3
        row = next(
            row
            for row in range(2, worksheet.max_row + 1)
            if worksheet.cell(row, key_column).value == "draft_revision"
        )
        worksheet.cell(row, value_column, 0)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    preview = preview_workbook(output.getvalue(), filename="invalid-identity.xlsx")
    assert "invalid_draft_revision" in {issue.code for issue in preview.issues}
    assert preview.identity is None

    valid = preview_workbook(_export(), filename="valid-identity.xlsx")
    assert valid.identity is not None
    assert valid.identity.issuer_id == "issuer-1"
    assert valid.identity.draft_revision == 7


def test_nominal_xlsx_with_activex_part_is_rejected_before_openpyxl() -> None:
    content = _add_part(_export(), "xl/activeX/activeX1.bin", b"active-control")
    with pytest.raises(ModelWorkbookError) as exc:
        preview_workbook(content, filename="activex.xlsx")
    assert exc.value.code == "active_or_external_content"


def test_aggregate_declared_dimensions_fail_before_blank_grid_iteration() -> None:
    content = _replace_part(
        _export(),
        "xl/worksheets/sheet1.xml",
        b'<dimension ref="A1:B14" />',
        b'<dimension ref="A1:CR6000" />',
    )
    content = _replace_part(
        content,
        "xl/worksheets/sheet2.xml",
        b'<dimension ref="A1:I13" />',
        b'<dimension ref="A1:CR6000" />',
    )
    preview = preview_workbook(content, filename="inflated-dimensions.xlsx")
    assert "workbook_dimension_limit" in {issue.code for issue in preview.issues}
    assert preview.draft_payload is None
    assert preview.formula_audit == []


def _legacy_workbook(*, duplicate_period_header: bool = False) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inputs"
    worksheet.append([
        "Period",
        "Period" if duplicate_period_header else "Display",
        "Type",
        "Adj EBITDA",
        "Cash",
        "Debt",
        "Interest",
        "Tax",
        "Capex",
        "WC",
        "Other",
    ])
    worksheet.append([
        "FY2026",
        "FY26e",
        "forecast",
        100,
        20,
        200,
        10,
        5,
        10,
        -2,
        0,
    ])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _legacy_mapping() -> dict:
    return {
        "mode": "mapped_legacy",
        "assumptions": {
            "sheet": "Inputs",
            "header_row": 1,
            "columns": {
                "period_key": "Period",
                "label": "Display",
                "kind": "Type",
                "adjusted_ebitda": "Adj EBITDA",
                "cash": "Cash",
                "total_debt": "Debt",
                "cash_interest": "Interest",
                "taxes": "Tax",
                "capex": "Capex",
                "working_capital_change": "WC",
                "other_cash_flow": "Other",
            },
        },
        "reporting_currency": "USD",
        "reporting_unit": "millions",
        "source_ids": ["legacy-document-1"],
        "authority_as_of": "2026-06-30T00:00:00Z",
    }


def _matrix_workbook(
    *,
    duplicate_account: bool = False,
    duplicate_period: bool = False,
    formula_input: bool = False,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model Matrix"
    headers = ["Account", "LTM Jun-26", "FY26E"]
    if duplicate_period:
        headers.append("FY26E")
    headers.append("FY27E")
    worksheet.append(headers)
    accounts = [
        ("Revenue", [720, 800, 800, 850] if duplicate_period else [720, 800, 850]),
        ("Adjusted EBITDA", [90, 100, 100, 110] if duplicate_period else [90, 100, 110]),
        ("Cash", [18, 20, 20, 22] if duplicate_period else [18, 20, 22]),
        ("Total Debt", [205, 200, 200, 190] if duplicate_period else [205, 200, 190]),
        ("Cash Interest", [11, 10, 10, 9] if duplicate_period else [11, 10, 9]),
        ("Taxes", [4, 5, 5, 6] if duplicate_period else [4, 5, 6]),
        ("Capex", [8, 10, 10, 11] if duplicate_period else [8, 10, 11]),
        ("Working Capital", [-1, -2, -2, -2] if duplicate_period else [-1, -2, -2]),
        ("Other Cash Flow", [0, 0, 0, 0] if duplicate_period else [0, 0, 0]),
    ]
    for label, values in accounts:
        worksheet.append([label, *values])
        if duplicate_account and label == "Cash":
            worksheet.append([label, *(value + 1 for value in values)])
    if formula_input:
        worksheet["C3"] = "=50+50"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    content = output.getvalue()
    if formula_input:
        content = _replace_part(
            content,
            "xl/worksheets/sheet1.xml",
            b"<f>50+50</f><v />",
            b"<f>50+50</f><v>100</v>",
        )
    return content


def _matrix_mapping() -> dict:
    return {
        "mode": "mapped_legacy",
        "assumptions": {
            "layout": "account_period_matrix",
            "sheet": "Model Matrix",
            "header_row": 1,
            "account_column": "Account",
            "account_rows": {
                "revenue": "Revenue",
                "adjusted_ebitda": "Adjusted EBITDA",
                "cash": "Cash",
                "total_debt": "Total Debt",
                "cash_interest": "Cash Interest",
                "taxes": "Taxes",
                "capex": "Capex",
                "working_capital_change": "Working Capital",
                "other_cash_flow": "Other Cash Flow",
            },
            "period_columns": {
                "LTM Jun-26": "LTM Jun-26",
                "FY26E": "FY26E",
                "FY27E": "FY27E",
            },
            "period_labels": {
                "LTM Jun-26": "LTM Jun-26",
                "FY26E": "FY26e",
                "FY27E": "FY27e",
            },
            "period_kinds": {
                "LTM Jun-26": "ltm",
                "FY26E": "forecast",
                "FY27E": "forecast",
            },
        },
        "debt_schedule": None,
        "overrides": None,
        "reporting_currency": "USD",
        "reporting_unit": "millions",
        "source_ids": ["matrix-workbook-fixture"],
        "authority_as_of": "2026-06-30T00:00:00Z",
    }


def _quarterly_debt_workbook() -> bytes:
    workbook = Workbook()
    assumptions = workbook.active
    assumptions.title = "Inputs"
    assumptions.append(["Period", "Display", "Type", "Adj EBITDA", "Cash"])
    assumptions.append(["Q1 26E", "Q1 26e", "forecast", 100, 20])
    assumptions.append(["YTD Q2 26E", "YTD Q2 26e", "forecast", 200, 20])
    debt = workbook.create_sheet("Debt")
    debt.append([
        "Instrument ID", "Name", "Priority", "Seniority", "Currency", "Rate Type",
        "Period", "Opening", "Draws", "Repayments", "Amortization", "Coupon",
        "Commitment Fee", "PIK", "Cash Fees", "Hedge", "FX",
    ])
    for period in ("Q1 26E", "YTD Q2 26E"):
        debt.append([
            "tlb-1", "Term Loan B", 1, "1L", "USD", "fixed", period,
            100, 0, 0, 0, 0.12, 0, 0, 0, 0, 1,
        ])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _quarterly_debt_mapping() -> dict:
    return {
        "mode": "mapped_legacy",
        "assumptions": {
            "sheet": "Inputs",
            "header_row": 1,
            "columns": {
                "period_key": "Period",
                "label": "Display",
                "kind": "Type",
                "adjusted_ebitda": "Adj EBITDA",
                "cash": "Cash",
            },
        },
        "debt_schedule": {
            "sheet": "Debt",
            "header_row": 1,
            "columns": {
                "instrument_id": "Instrument ID",
                "name": "Name",
                "priority": "Priority",
                "seniority": "Seniority",
                "currency": "Currency",
                "rate_type": "Rate Type",
                "period_key": "Period",
                "opening_balance": "Opening",
                "draws": "Draws",
                "repayments": "Repayments",
                "scheduled_amortization": "Amortization",
                "coupon_rate": "Coupon",
                "commitment_fee_rate": "Commitment Fee",
                "pik_rate": "PIK",
                "cash_fees": "Cash Fees",
                "hedge_effect": "Hedge",
                "fx_rate": "FX",
            },
        },
        "overrides": None,
        "reporting_currency": "USD",
        "reporting_unit": "millions",
        "source_ids": ["quarterly-debt-fixture"],
        "authority_as_of": "2026-06-30T00:00:00Z",
    }


def test_bounded_explicit_legacy_mapping_derives_payload_and_calculation() -> None:
    preview = preview_workbook(
        _legacy_workbook(),
        filename="legacy.xlsx",
        mapping=_legacy_mapping(),
        evaluated_at=EXPORTED_AT,
    )
    assert preview.mode == "mapped_legacy"
    assert preview.blocking_count == 0
    assert [issue.code for issue in preview.issues] == ["legacy_mapping"]
    assert preview.draft_payload is not None
    assert preview.draft_payload.periods[0].period_key == "FY2026"
    assert preview.draft_payload.periods[0].adjusted_ebitda == 100
    assert preview.draft_payload.periods[0].authority.origin == "imported"
    assert preview.calculation is not None
    assert preview.calculation.periods[0].net_leverage == pytest.approx(1.8)


def test_matrix_mapping_normalizes_period_aliases_and_builds_stable_periods() -> None:
    preview = preview_workbook(
        _matrix_workbook(),
        filename="account-period-matrix.xlsx",
        mapping=_matrix_mapping(),
        evaluated_at=EXPORTED_AT,
    )

    assert preview.blocking_count == 0
    assert preview.mapping is not None
    assert list(preview.mapping.assumptions.period_columns) == [
        "LTM-2026-06-30",
        "FY2026",
        "FY2027",
    ]
    assert preview.draft_payload is not None
    assert [period.period_key for period in preview.draft_payload.periods] == [
        "LTM-2026-06-30",
        "FY2026",
        "FY2027",
    ]
    assert preview.draft_payload.periods[1].adjusted_ebitda == 100
    assert preview.calculation is not None
    assert preview.calculation.periods[1].net_leverage == pytest.approx(1.8)


def test_long_form_quarter_and_ytd_aliases_derive_months_used_by_debt_interest() -> None:
    preview = preview_workbook(
        _quarterly_debt_workbook(),
        filename="quarterly-debt.xlsx",
        mapping=_quarterly_debt_mapping(),
        evaluated_at=EXPORTED_AT,
    )

    assert preview.blocking_count == 0
    assert preview.draft_payload is not None
    assert [period.period_key for period in preview.draft_payload.periods] == [
        "Q1-2026",
        "YTD-Q2-2026",
    ]
    assert [period.months for period in preview.draft_payload.periods] == [3, 6]
    assert preview.calculation is not None
    assert preview.calculation.periods[0].instruments[0].cash_interest == pytest.approx(3)
    assert preview.calculation.periods[1].instruments[0].cash_interest == pytest.approx(6)


def test_matrix_quarter_and_ytd_aliases_derive_bounded_period_months() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model Matrix"
    worksheet.append(["Account", "Q1 26E", "YTD Q2 26E"])
    worksheet.append(["Adjusted EBITDA", 100, 200])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    mapping = _matrix_mapping()
    mapping["assumptions"]["account_rows"] = {
        "adjusted_ebitda": "Adjusted EBITDA"
    }
    mapping["assumptions"]["period_columns"] = {
        "Q1 26E": "Q1 26E",
        "YTD Q2 26E": "YTD Q2 26E",
    }
    mapping["assumptions"]["period_labels"] = {
        "Q1 26E": "Q1 26e",
        "YTD Q2 26E": "YTD Q2 26e",
    }
    mapping["assumptions"]["period_kinds"] = {
        "Q1 26E": "forecast",
        "YTD Q2 26E": "forecast",
    }

    preview = preview_workbook(
        output.getvalue(),
        filename="quarterly-matrix.xlsx",
        mapping=mapping,
    )

    assert preview.blocking_count == 0
    assert preview.draft_payload is not None
    assert [period.months for period in preview.draft_payload.periods] == [3, 6]


def test_matrix_duplicate_account_and_period_require_reviewed_physical_selectors() -> None:
    mapping = _matrix_mapping()
    unresolved = preview_workbook(
        _matrix_workbook(duplicate_account=True, duplicate_period=True),
        filename="ambiguous-matrix.xlsx",
        mapping=mapping,
    )

    assert unresolved.blocking_count > 0
    assert {(item.field, item.selector) for item in unresolved.ambiguities} == {
        ("cash", "row"),
        ("FY2026", "column"),
    }
    assert unresolved.draft_payload is None

    mapping["assumptions"]["account_row_indices"] = {"cash": 4}
    mapping["assumptions"]["period_column_indices"] = {"FY26E": 3}
    resolved = preview_workbook(
        _matrix_workbook(duplicate_account=True, duplicate_period=True),
        filename="reviewed-matrix.xlsx",
        mapping=mapping,
    )

    assert resolved.blocking_count == 0
    assert resolved.ambiguities == []
    assert resolved.draft_payload is not None
    assert resolved.draft_payload.periods[1].cash == 20


def test_matrix_formula_is_audited_but_never_promoted_to_model_authority() -> None:
    preview = preview_workbook(
        _matrix_workbook(formula_input=True),
        filename="matrix-formula.xlsx",
        mapping=_matrix_mapping(),
    )

    assert "matrix_formula_not_authoritative" in {issue.code for issue in preview.issues}
    assert preview.draft_payload is None
    assert preview.calculation is None
    entry = next(item for item in preview.formula_audit if item.cell == "C3")
    assert entry.cached_value == 100
    assert entry.disposition == "comparison_only"


def test_legacy_formula_is_audited_but_never_promoted_to_model_authority() -> None:
    content = _replace_part(
        _legacy_workbook(),
        "xl/worksheets/sheet1.xml",
        b'<c r="D2" t="n"><v>100</v></c>',
        b'<c r="D2"><f>50+50</f><v>100</v></c>',
    )
    preview = preview_workbook(
        content,
        filename="legacy-formula.xlsx",
        mapping=_legacy_mapping(),
        evaluated_at=EXPORTED_AT,
    )
    assert preview.blocking_count > 0
    assert "imported_formula_not_authoritative" in {
        issue.code for issue in preview.issues
    }
    assert preview.draft_payload is None
    assert preview.calculation is None
    assert preview.formula_audit[0].cached_value == 100
    assert preview.formula_audit[0].disposition == "comparison_only"


def test_deep_formula_chain_is_bounded_without_python_recursion_failure() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inputs"
    worksheet.append([
        "Period", "Display", "Type", "Adj EBITDA", "Cash", "Debt",
        "Interest", "Tax", "Capex", "WC", "Other", "Formula Audit",
    ])
    worksheet.append([
        "FY2026", "FY26e", "forecast", 100, 20, 200, 10, 5, 10, -2, 0,
        "=L3",
    ])
    for row_number in range(3, 1_102):
        worksheet.cell(row_number, 12, f"=L{row_number + 1}")
    worksheet.cell(1_102, 12, "=1")
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    content = _replace_all_in_part(
        output.getvalue(),
        "xl/worksheets/sheet1.xml",
        b"<v />",
        b"<v>1</v>",
    )

    preview = preview_workbook(
        content,
        filename="deep-formulas.xlsx",
        mapping=_legacy_mapping(),
        evaluated_at=EXPORTED_AT,
    )
    assert preview.blocking_count == 0
    assert len(preview.formula_audit) == 1_101
    assert preview.draft_payload is not None


def test_formula_reference_fanout_has_an_aggregate_dos_bound() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inputs"
    worksheet.append([
        "Period", "Display", "Type", "Adj EBITDA", "Cash", "Debt",
        "Interest", "Tax", "Capex", "WC", "Other", "Formula Audit",
    ])
    formula = "=" + "+".join(["A1"] * 1_001)
    worksheet.append([
        "FY2026", "FY26e", "forecast", 100, 20, 200, 10, 5, 10, -2, 0,
        formula,
    ])
    for row_number in range(3, 53):
        worksheet.cell(row_number, 12, formula)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    content = _replace_all_in_part(
        output.getvalue(),
        "xl/worksheets/sheet1.xml",
        b"<v />",
        b"<v>1</v>",
    )

    preview = preview_workbook(
        content,
        filename="formula-fanout.xlsx",
        mapping=_legacy_mapping(),
    )
    assert "formula_dependency_limit" in {issue.code for issue in preview.issues}
    assert preview.draft_payload is None


def test_legacy_duplicate_header_is_an_explicit_blocking_ambiguity() -> None:
    mapping = _legacy_mapping()
    mapping["assumptions"]["columns"]["label"] = "Period"
    preview = preview_workbook(
        _legacy_workbook(duplicate_period_header=True),
        filename="ambiguous.xlsx",
        mapping=mapping,
    )
    assert preview.blocking_count > 0
    assert {ambiguity.field for ambiguity in preview.ambiguities} == {"period_key", "label"}
    assert "ambiguous_mapping" in {issue.code for issue in preview.issues}


def test_reviewed_one_based_column_selectors_resolve_duplicate_headers() -> None:
    mapping = _legacy_mapping()
    mapping["assumptions"]["columns"]["label"] = "Period"
    mapping["assumptions"]["column_indices"] = {
        "period_key": 1,
        "label": 2,
    }

    preview = preview_workbook(
        _legacy_workbook(duplicate_period_header=True),
        filename="reviewed-ambiguity.xlsx",
        mapping=mapping,
    )

    assert preview.blocking_count == 0
    assert preview.ambiguities == []
    assert preview.draft_payload is not None
    assert preview.draft_payload.periods[0].period_key == "FY2026"
    assert preview.draft_payload.periods[0].label == "FY26e"


def test_mapping_and_file_boundary_are_extra_forbid_and_xlsx_only() -> None:
    mapping = _legacy_mapping()
    mapping["surprise"] = True
    with pytest.raises(ModelWorkbookError, match="extra"):
        parse_mapping(mapping)
    with pytest.raises(ModelWorkbookError) as exc:
        preview_workbook(_legacy_workbook(), filename="legacy.xls", mapping=_legacy_mapping())
    assert exc.value.code == "xlsx_required"


@pytest.mark.parametrize(
    ("field_name", "invalid_value"),
    [
        ("reporting_currency", None),
        ("reporting_currency", ""),
        ("reporting_currency", "ZZZ"),
        ("reporting_unit", None),
        ("reporting_unit", ""),
        ("reporting_unit", "trillions"),
    ],
)
def test_legacy_mapping_requires_supported_monetary_identity(
    field_name: str,
    invalid_value: str | None,
) -> None:
    mapping = _legacy_mapping()
    if invalid_value is None:
        mapping.pop(field_name)
    else:
        mapping[field_name] = invalid_value

    with pytest.raises(ModelWorkbookError, match=field_name) as invalid:
        parse_mapping(mapping)
    assert invalid.value.code == "invalid_mapping"
