"""Secure API-boundary tests for Model Engine v2 workbook import/export."""

from __future__ import annotations

import asyncio
import base64
import hashlib
import io
import json
import sqlite3
import uuid
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, func, select


_ISSUER_IDS: set[str] = set()
_OWNERS: set[str] = set()
_CONTEXT_IDS: set[str] = set()


def _run(coroutine):
    return asyncio.run(coroutine)


def _identity(owner: str, *, role: str = "analyst", team_id: str | None = None):
    from identity import CallerIdentity

    return CallerIdentity(
        id=owner,
        email=f"{owner}@test.local",
        full_name=owner,
        role=role,
        source="profile",
        team_id=team_id,
    )


def _as(owner: str, *, role: str = "analyst", team_id: str | None = None) -> None:
    from identity import get_identity
    from main import app

    caller = _identity(owner, role=role, team_id=team_id)
    app.dependency_overrides[get_identity] = lambda: caller
    _OWNERS.add(owner)


def _payload(
    *,
    revenue: float = 800.0,
    origin: str = "analyst",
    include_override: bool = False,
) -> dict:
    authority = {
        "origin": origin,
        "method": "model-workbook-api-fixture",
        "source_ids": ["fixture-source"],
        "as_of": "2026-06-30T00:00:00Z",
    }
    return {
        "schema_version": 2,
        "reporting_currency": "USD",
        "reporting_unit": "millions",
        "periods": [
            {
                "period_key": "FY2026",
                "label": "FY26e",
                "kind": "forecast",
                "months": 12,
                "revenue": revenue,
                "reported_ebitda": 100.0,
                "adjustments": 10.0,
                "cash": 20.0,
                "taxes": 5.0,
                "capex": 10.0,
                "working_capital_change": -2.0,
                "other_cash_flow": 0.0,
                "authority": authority,
            }
        ],
        "debt_instruments": [
            {
                "instrument_id": "tlb-1",
                "name": "First-lien term loan",
                "priority": 1,
                "seniority": "1L",
                "currency": "USD",
                "rate_type": "floating",
                "maturity": "2030-06-30",
                "benchmark_curve": "SOFR",
                "sources": ["fixture-source"],
                "authority": authority,
                "periods": [
                    {
                        "period_key": "FY2026",
                        "opening_balance": 200.0,
                        "closing_balance": 190.0,
                        "draws": 0.0,
                        "repayments": 10.0,
                        "scheduled_amortization": 0.0,
                        "commitment": 220.0,
                        "benchmark_rate": 0.04,
                        "floor_rate": 0.05,
                        "spread_rate": 0.03,
                        "coupon_rate": 0.01,
                        "commitment_fee_rate": 0.005,
                        "pik_rate": 0.0,
                        "cash_fees": 1.0,
                        "hedge_effect": -0.5,
                        "fx_rate": 1.0,
                    }
                ],
            }
        ],
        "overrides": [
            {
                "node_id": "calc:FY2026:adjusted_ebitda",
                "value_type": "number",
                "value": 105.0,
                "reason": "Workbook API fixture",
                "scope": "draft",
                "source": "fixture-source",
                "expires_at": "2099-01-01T00:00:00+00:00",
            }
        ] if include_override else [],
        "source_ids": ["fixture-source"],
    }


async def _seed_issuer(owner: str, *, team_id: str | None = None) -> str:
    from database import AsyncSessionLocal, Issuer

    token = uuid.uuid4().hex[:20]
    issuer_id = f"mwi-{token}"
    async with AsyncSessionLocal() as db:
        db.add(
            Issuer(
                id=issuer_id,
                name=f"Model workbook API {token}",
                team_id=team_id,
                created_by=owner,
            )
        )
        await db.commit()
    _ISSUER_IDS.add(issuer_id)
    _OWNERS.add(owner)
    return issuer_id


async def _seed_context(owner: str, issuer_id: str) -> str:
    from database import AnalysisContextRecord, AsyncSessionLocal

    context_id = f"mwc-{uuid.uuid4().hex[:20]}"
    async with AsyncSessionLocal() as db:
        db.add(AnalysisContextRecord(
            id=context_id,
            analyst_id=owner,
            name="Model workbook restore lineage",
            issuer_ids=[issuer_id],
        ))
        await db.commit()
    _CONTEXT_IDS.add(context_id)
    return context_id


async def _seed_stale_draft(
    owner: str,
    issuer_id: str,
    *,
    payload_value: dict,
    evaluated_at: datetime,
) -> None:
    from database import AsyncSessionLocal, ModelDraftV2
    from model_engine_v2 import ModelDraftPayload, calculate_model

    payload = ModelDraftPayload.model_validate(payload_value)
    calculation = calculate_model(payload, evaluated_at=evaluated_at)
    async with AsyncSessionLocal() as db:
        db.add(
            ModelDraftV2(
                id=f"mwd-{uuid.uuid4().hex[:20]}",
                issuer_id=issuer_id,
                analyst_id=owner,
                payload=payload.model_dump(mode="json"),
                calculation=calculation.model_dump(mode="json"),
                source_fingerprint=calculation.source_fingerprint,
                input_fingerprint=calculation.input_fingerprint,
                engine_version=calculation.engine_version,
                calculation_hash=calculation.calculation_hash,
                revision=1,
                created_at=evaluated_at,
                updated_at=evaluated_at,
            )
        )
        await db.commit()


def _save(
    client: TestClient,
    issuer_id: str,
    *,
    expected_revision: int = 0,
    payload: dict | None = None,
    context_id: str | None = None,
):
    return client.put(
        f"/api/models/v2/{issuer_id}",
        json={
            "expected_revision": expected_revision,
            "payload": payload or _payload(),
            "context_id": context_id,
            "source_run_id": None,
        },
    )


def _set_override(
    client: TestClient,
    issuer_id: str,
    *,
    expected_revision: int,
    expires_at: datetime | None = None,
):
    override = {
        "node_id": "calc:FY2026:adjusted_ebitda",
        "value_type": "number",
        "value": 105.0,
        "reason": "Workbook API fixture",
        "scope": "draft",
        "source": "fixture-source",
    }
    if expires_at is not None:
        override["expires_at"] = expires_at.isoformat()
    return client.post(
        f"/api/models/v2/{issuer_id}/overrides",
        json={
            "expected_revision": expected_revision,
            "action": "set",
            "override": override,
        },
    )


def _export(client: TestClient, issuer_id: str) -> bytes:
    response = client.get(f"/api/models/v2/{issuer_id}/workbook/export")
    assert response.status_code == 200, response.text
    return response.content


def _preview(
    client: TestClient,
    issuer_id: str,
    content: bytes,
    *,
    expected_revision: int,
    filename: str = "model.xlsx",
    mapping: dict | None = None,
):
    return client.post(
        f"/api/models/v2/{issuer_id}/workbook/import/preview",
        files={
            "file": (
                filename,
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "mapping": "" if mapping is None else json.dumps(mapping),
            "expected_revision": str(expected_revision),
        },
    )


def _commit(
    client: TestClient,
    issuer_id: str,
    content: bytes,
    preview: dict,
    *,
    expected_revision: int,
    filename: str = "model.xlsx",
    mapping: dict | None = None,
    preview_sha256: str | None = None,
    preview_token: str | None = None,
):
    return client.post(
        f"/api/models/v2/{issuer_id}/workbook/import/commit",
        files={
            "file": (
                filename,
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "mapping": "" if mapping is None else json.dumps(mapping),
            "expected_revision": str(expected_revision),
            "preview_sha256": preview_sha256 or preview["workbook_sha256"],
            "preview_token": preview_token
            if preview_token is not None
            else preview.get("preview_token", ""),
        },
    )


async def _counts(owner: str, issuer_id: str) -> dict[str, int]:
    from database import (
        AsyncSessionLocal,
        Document,
        LineageEdge,
        ModelDraftV2,
        ModelOverrideEvent,
        ModelWorkbookImport,
        SourceManifest,
    )

    async with AsyncSessionLocal() as db:

        async def count(model, *conditions) -> int:
            return int(
                (
                    await db.execute(
                        select(func.count()).select_from(model).where(*conditions)
                    )
                ).scalar_one()
            )

        return {
            "documents": await count(
                Document, Document.analyst_id == owner, Document.issuer_id == issuer_id
            ),
            "manifests": await count(
                SourceManifest,
                SourceManifest.analyst_id == owner,
                SourceManifest.issuer_id == issuer_id,
            ),
            "imports": await count(
                ModelWorkbookImport,
                ModelWorkbookImport.analyst_id == owner,
                ModelWorkbookImport.issuer_id == issuer_id,
            ),
            "drafts": await count(
                ModelDraftV2,
                ModelDraftV2.analyst_id == owner,
                ModelDraftV2.issuer_id == issuer_id,
            ),
            "events": await count(
                ModelOverrideEvent,
                ModelOverrideEvent.analyst_id == owner,
                ModelOverrideEvent.issuer_id == issuer_id,
            ),
            "lineage": await count(LineageEdge, LineageEdge.analyst_id == owner),
        }


def _vault_files() -> set[Path]:
    from config import get_settings

    root = Path(get_settings().caos_storage_dir) / "models"
    return (
        {path for path in root.rglob("*") if path.is_file()} if root.exists() else set()
    )


async def _artifact_rows(import_id: str) -> dict:
    from database import (
        AsyncSessionLocal,
        Document,
        LineageEdge,
        ModelDraftV2,
        ModelWorkbookImport,
        SourceManifest,
    )

    async with AsyncSessionLocal() as db:
        imported = await db.get(ModelWorkbookImport, import_id)
        assert imported is not None
        document = await db.get(Document, imported.document_id)
        manifest = await db.get(SourceManifest, imported.source_manifest_id)
        draft = await db.get(ModelDraftV2, imported.draft_id)
        lineage = (
            await db.execute(
                select(LineageEdge).where(
                    LineageEdge.analyst_id == imported.analyst_id,
                    LineageEdge.artifact_id
                    == f"source_manifest:{imported.source_manifest_id}",
                    LineageEdge.parent_id == f"document:{imported.document_id}",
                )
            )
        ).scalar_one_or_none()
        assert document is not None and manifest is not None and draft is not None
        return {
            "import": imported,
            "document": document,
            "manifest": manifest,
            "draft": draft,
            "lineage": lineage,
        }


async def _checkpoint_parent_ids(owner: str, checkpoint_id: str) -> set[str]:
    from database import AsyncSessionLocal, LineageEdge

    async with AsyncSessionLocal() as db:
        return set(
            (
                await db.execute(
                    select(LineageEdge.parent_id).where(
                        LineageEdge.analyst_id == owner,
                        LineageEdge.artifact_id
                        == f"model_checkpoint:{checkpoint_id}",
                    )
                )
            ).scalars()
        )


async def _cleanup() -> None:
    if not _ISSUER_IDS:
        return
    from database import (
        AnalysisContextRecord,
        AsyncSessionLocal,
        Document,
        Issuer,
        LineageEdge,
        ModelCheckpoint,
        ModelDraftV2,
        ModelOverrideEvent,
        ModelWorkbookImport,
        SourceManifest,
    )
    import model_storage

    async with AsyncSessionLocal() as db:
        storage_keys = list(
            (
                await db.execute(
                    select(Document.storage_key).where(
                        Document.issuer_id.in_(_ISSUER_IDS),
                        Document.analyst_id.in_(_OWNERS),
                        Document.doc_type == "ModelWorkbook",
                    )
                )
            ).scalars()
        )
        await db.execute(
            delete(ModelWorkbookImport).where(
                ModelWorkbookImport.issuer_id.in_(_ISSUER_IDS)
            )
        )
        await db.execute(
            delete(ModelOverrideEvent).where(
                ModelOverrideEvent.issuer_id.in_(_ISSUER_IDS)
            )
        )
        await db.execute(delete(LineageEdge).where(LineageEdge.analyst_id.in_(_OWNERS)))
        if _CONTEXT_IDS:
            await db.execute(
                delete(ModelCheckpoint).where(
                    ModelCheckpoint.context_id.in_(_CONTEXT_IDS)
                )
            )
        await db.execute(
            delete(ModelDraftV2).where(ModelDraftV2.issuer_id.in_(_ISSUER_IDS))
        )
        if _CONTEXT_IDS:
            await db.execute(
                delete(AnalysisContextRecord).where(
                    AnalysisContextRecord.id.in_(_CONTEXT_IDS)
                )
            )
        await db.execute(
            delete(SourceManifest).where(SourceManifest.issuer_id.in_(_ISSUER_IDS))
        )
        await db.execute(delete(Document).where(Document.issuer_id.in_(_ISSUER_IDS)))
        await db.execute(delete(Issuer).where(Issuer.id.in_(_ISSUER_IDS)))
        await db.commit()
    for key in storage_keys:
        model_storage.remove_uncommitted(key)


@pytest.fixture(scope="module")
def client():
    from config import get_settings
    from main import app

    settings = get_settings()
    original = (
        settings.caos_model_engine_v2_enabled,
        settings.caos_lineage_v2_enabled,
        settings.caos_tenancy_enabled,
    )
    settings.caos_model_engine_v2_enabled = True
    settings.caos_lineage_v2_enabled = True
    settings.caos_tenancy_enabled = False
    with TestClient(app) as value:
        yield value
        _run(_cleanup())
    (
        settings.caos_model_engine_v2_enabled,
        settings.caos_lineage_v2_enabled,
        settings.caos_tenancy_enabled,
    ) = original


@pytest.fixture(autouse=True)
def _reset_contract():
    from config import get_settings
    from identity import get_identity
    from main import app
    from routes import model_workbook as route

    settings = get_settings()
    settings.caos_model_engine_v2_enabled = True
    settings.caos_lineage_v2_enabled = True
    settings.caos_tenancy_enabled = False
    route._preview_sem = None
    yield
    app.dependency_overrides.pop(get_identity, None)
    settings.caos_model_engine_v2_enabled = True
    settings.caos_lineage_v2_enabled = True
    settings.caos_tenancy_enabled = False
    route._preview_sem = None


def _tamper_metadata(content: bytes, key: str, value: object) -> bytes:
    workbook = load_workbook(io.BytesIO(content))
    for sheet_name, key_column, value_column in (
        ("Cover", 1, 2),
        ("Sources - Audit", 2, 3),
    ):
        worksheet = workbook[sheet_name]
        row_number = next(
            row
            for row in range(2, worksheet.max_row + 1)
            if worksheet.cell(row, key_column).value == key
        )
        worksheet.cell(row_number, value_column, value)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _formula_without_cache(content: bytes) -> bytes:
    output = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(content)) as source,
        zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target,
    ):
        for item in source.infolist():
            payload = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet3.xml":
                original = b'<c r="G2" t="n"><v>100</v></c>'
                assert original in payload
                payload = payload.replace(
                    original,
                    b'<c r="G2"><f>50+50</f><v></v></c>',
                    1,
                )
            target.writestr(item, payload)
    return output.getvalue()


def _legacy_workbook(*, duplicate_period_header: bool = False) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inputs"
    worksheet.append(
        [
            "Period",
            "Period" if duplicate_period_header else "Display",
            "Type",
            "Adj EBITDA",
            "Cash",
            "Debt",
            "Interest",
            "Tax",
            "Capex",
            "WC",
            "Other",
        ]
    )
    worksheet.append(
        [
            "FY2026",
            "FY26e",
            "forecast",
            100,
            20,
            200,
            10,
            5,
            10,
            -2,
            0,
        ]
    )
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _legacy_mapping() -> dict:
    return {
        "mode": "mapped_legacy",
        "assumptions": {
            "sheet": "Inputs",
            "header_row": 1,
            "columns": {
                "period_key": "Period",
                "label": "Display",
                "kind": "Type",
                "adjusted_ebitda": "Adj EBITDA",
                "cash": "Cash",
                "total_debt": "Debt",
                "cash_interest": "Interest",
                "taxes": "Tax",
                "capex": "Capex",
                "working_capital_change": "WC",
                "other_cash_flow": "Other",
            },
        },
        "reporting_currency": "USD",
        "reporting_unit": "millions",
        "source_ids": ["legacy-workbook-fixture"],
        "authority_as_of": "2026-06-30T00:00:00Z",
    }


def _matrix_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Model Matrix"
    worksheet.append(["Account", "LTM Jun-26", "FY26E"])
    worksheet.append(["Adjusted EBITDA", 90, 100])
    worksheet.append(["Cash", 18, 20])
    worksheet.append(["Total Debt", 205, 200])
    worksheet.append(["Cash Interest", 11, 10])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _matrix_mapping() -> dict:
    return {
        "mode": "mapped_legacy",
        "assumptions": {
            "layout": "account_period_matrix",
            "sheet": "Model Matrix",
            "header_row": 1,
            "account_column": "Account",
            "account_rows": {
                "adjusted_ebitda": "Adjusted EBITDA",
                "cash": "Cash",
                "total_debt": "Total Debt",
                "cash_interest": "Cash Interest",
            },
            "period_columns": {
                "LTM Jun-26": "LTM Jun-26",
                "FY26E": "FY26E",
            },
            "period_labels": {
                "LTM Jun-26": "LTM Jun-26",
                "FY26E": "FY26e",
            },
            "period_kinds": {
                "LTM Jun-26": "ltm",
                "FY26E": "forecast",
            },
        },
        "debt_schedule": None,
        "overrides": None,
        "reporting_currency": "USD",
        "reporting_unit": "millions",
        "source_ids": ["matrix-workbook-api-fixture"],
        "authority_as_of": "2026-06-30T00:00:00Z",
    }


def test_flag_off_runs_before_body_validation_and_export_has_exact_six_sheets(client):
    from config import get_settings

    get_settings().caos_model_engine_v2_enabled = False
    disabled = client.post(
        "/api/models/v2/not-enumerable/workbook/import/preview",
        content=b"not even multipart",
        headers={"content-type": "application/octet-stream"},
    )
    assert disabled.status_code == 404
    assert disabled.json() == {"detail": "Not Found"}

    get_settings().caos_model_engine_v2_enabled = True
    owner = f"mwi-export-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    saved = _save(client, issuer_id)
    assert saved.status_code == 200, saved.text

    response = client.get(f"/api/models/v2/{issuer_id}/workbook/export")
    assert response.status_code == 200, response.text
    assert response.headers["cache-control"] == "private, no-store"
    assert response.headers["x-caos-model-revision"] == "1"
    workbook = load_workbook(
        io.BytesIO(response.content), read_only=True, data_only=False
    )
    assert tuple(workbook.sheetnames) == (
        "Cover",
        "Model",
        "Assumptions",
        "Debt Schedule",
        "Overrides",
        "Sources - Audit",
    )
    workbook.close()


def test_preview_is_read_only_and_av_corrupt_and_xls_fail_closed(client, monkeypatch):
    from routes import model_workbook as route

    owner = f"mwi-preview-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    assert _save(client, issuer_id).status_code == 200
    content = _export(client, issuer_id)
    before = _run(_counts(owner, issuer_id))
    before_files = _vault_files()

    preview = _preview(client, issuer_id, content, expected_revision=1)
    assert preview.status_code == 200, preview.text
    assert preview.json()["preview_token"]
    assert _run(_counts(owner, issuer_id)) == before
    assert _vault_files() == before_files

    async def infected(_content: bytes) -> None:
        raise HTTPException(422, "fixture malware")

    with monkeypatch.context() as patcher:
        patcher.setattr(route.avscan, "scan", infected)
        rejected = _preview(client, issuer_id, content, expected_revision=1)
    assert rejected.status_code == 422
    assert rejected.json() == {"detail": "fixture malware"}

    corrupt = _preview(client, issuer_id, b"not-an-xlsx", expected_revision=1)
    assert corrupt.status_code == 422
    legacy_binary = _preview(
        client,
        issuer_id,
        content,
        expected_revision=1,
        filename="model.xls",
    )
    assert legacy_binary.status_code == 422
    assert legacy_binary.json()["detail"]["code"] == "xlsx_required"

    scans: list[bytes] = []

    async def record_scan(payload: bytes) -> None:
        scans.append(payload)

    with monkeypatch.context() as patcher:
        patcher.setattr(route.avscan, "scan", record_scan)
        hostile_name = _preview(
            client,
            issuer_id,
            content,
            expected_revision=1,
            filename=f"{'x' * 1_000}.xlsx",
        )
    assert hostile_name.status_code == 413
    assert scans == []
    assert _run(_counts(owner, issuer_id)) == before
    assert _vault_files() == before_files


@pytest.mark.parametrize("capacity", ["authority", "payload"])
def test_preview_blocks_when_import_provenance_would_exceed_source_capacity(
    client, capacity
):
    owner = f"mwi-source-capacity-{capacity}-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    payload = _payload()
    if capacity == "authority":
        source_ids = [f"authority-source-{index}" for index in range(100)]
        payload["periods"][0]["authority"]["source_ids"] = source_ids
        payload["debt_instruments"][0]["authority"]["source_ids"] = source_ids
    else:
        payload["source_ids"] = [
            f"payload-source-{index}" for index in range(500)
        ]
    saved = _save(client, issuer_id, payload=payload)
    assert saved.status_code == 200, saved.text
    content = _export(client, issuer_id)
    before = _run(_counts(owner, issuer_id))

    response = _preview(client, issuer_id, content, expected_revision=1)

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["blocking_count"] == 1
    assert body["preview_token"] is None
    assert body["draft_payload"] is None
    assert body["calculation"] is None
    assert {item["code"] for item in body["issues"]} == {
        "import_authority_capacity"
    }
    assert _run(_counts(owner, issuer_id)) == before


def test_export_refuses_expired_override_calculation_until_resaved(client):
    owner = f"mwi-stale-export-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    now = datetime.now(timezone.utc)
    payload = _payload(include_override=True)
    payload["overrides"][0]["expires_at"] = (now - timedelta(minutes=1)).isoformat()
    _run(
        _seed_stale_draft(
            owner,
            issuer_id,
            payload_value=payload,
            evaluated_at=now - timedelta(minutes=2),
        )
    )

    stale = client.get(f"/api/models/v2/{issuer_id}/workbook/export")
    assert stale.status_code == 409
    assert stale.json()["detail"]["message"] == (
        "Model calculation is stale; recalculate and save before exporting."
    )
    assert stale.json()["detail"]["current_revision"] == 1

    resaved = _save(
        client,
        issuer_id,
        expected_revision=1,
        payload=payload,
    )
    assert resaved.status_code == 200, resaved.text
    assert resaved.json()["revision"] == 2
    exported = client.get(f"/api/models/v2/{issuer_id}/workbook/export")
    assert exported.status_code == 200, exported.text


def test_signed_preview_binds_owner_issuer_mapping_revision_hash_and_identity(
    client, monkeypatch
):
    from routes import model_workbook as route

    owner = f"mwi-token-{uuid.uuid4().hex[:10]}"
    other = f"mwi-token-other-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    other_issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    assert _save(client, issuer_id).status_code == 200
    content = _export(client, issuer_id)
    preview_response = _preview(client, issuer_id, content, expected_revision=1)
    assert preview_response.status_code == 200, preview_response.text
    preview = preview_response.json()
    token = preview["preview_token"]
    assert token

    tampered_token = f"{token[:-1]}{'0' if token[-1] != '0' else '1'}"
    assert (
        _commit(
            client,
            issuer_id,
            content,
            preview,
            expected_revision=1,
            preview_token=tampered_token,
        ).status_code
        == 409
    )

    _as(other)
    assert (
        _commit(client, issuer_id, content, preview, expected_revision=1).status_code
        == 409
    )
    _as(owner)

    assert (
        _commit(
            client, other_issuer_id, content, preview, expected_revision=1
        ).status_code
        == 409
    )
    assert (
        _commit(
            client,
            issuer_id,
            content,
            preview,
            expected_revision=2,
        ).status_code
        == 409
    )
    assert (
        _commit(
            client,
            issuer_id,
            content,
            preview,
            expected_revision=1,
            preview_sha256="0" * 64,
        ).status_code
        == 409
    )
    assert (
        _commit(
            client,
            issuer_id,
            content,
            preview,
            expected_revision=1,
            mapping=_legacy_mapping(),
        ).status_code
        == 409
    )

    original_preview_workbook = route.preview_workbook

    def drift_identity(*args, **kwargs):
        reparsed = original_preview_workbook(*args, **kwargs)
        assert reparsed.identity is not None
        return reparsed.model_copy(
            update={
                "identity": reparsed.identity.model_copy(
                    update={"exported_by": "different-exporter"}
                )
            }
        )

    with monkeypatch.context() as patcher:
        patcher.setattr(route, "preview_workbook", drift_identity)
        drifted = _commit(client, issuer_id, content, preview, expected_revision=1)
    assert drifted.status_code == 409
    assert drifted.json() == {
        "detail": "Workbook identity differs from the preview; preview again."
    }

    wrong_issuer = _tamper_metadata(content, "issuer_id", other_issuer_id)
    identity_preview = _preview(client, issuer_id, wrong_issuer, expected_revision=1)
    assert identity_preview.status_code == 200, identity_preview.text
    identity_body = identity_preview.json()
    assert identity_body["preview_token"] is None
    assert "workbook_issuer_mismatch" in {
        issue["code"] for issue in identity_body["issues"]
    }

    wrong_revision = _tamper_metadata(content, "draft_revision", 2)
    revision_preview = _preview(client, issuer_id, wrong_revision, expected_revision=1)
    assert revision_preview.status_code == 200, revision_preview.text
    revision_body = revision_preview.json()
    assert revision_body["preview_token"] is None
    assert "workbook_revision_mismatch" in {
        issue["code"] for issue in revision_body["issues"]
    }
    assert _run(_counts(owner, issuer_id))["imports"] == 0


def test_new_draft_accepts_positive_strict_source_revision_and_commits_revision_one(
    client,
):
    from model_engine_v2 import ModelDraftPayload
    from model_workbook import render_model_workbook

    owner = f"mwi-new-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    content = render_model_workbook(
        ModelDraftPayload.model_validate(_payload()),
        issuer_id=issuer_id,
        draft_revision=7,
        exported_by="source-analyst",
        exported_at=datetime(2026, 7, 14, 12, 0, tzinfo=timezone.utc),
    )
    preview_response = _preview(client, issuer_id, content, expected_revision=0)
    assert preview_response.status_code == 200, preview_response.text
    preview = preview_response.json()
    assert preview["identity"]["draft_revision"] == 7
    assert preview["preview_token"]
    committed = _commit(client, issuer_id, content, preview, expected_revision=0)
    assert committed.status_code == 200, committed.text
    assert committed.json()["committed_revision"] == 1
    assert committed.json()["record"]["revision"] == 1


def test_preview_token_expires_with_active_override_and_is_rechecked_after_parse(
    client, monkeypatch
):
    from routes import model_workbook as route

    owner = f"mwi-expiry-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    expires_at = datetime.now(timezone.utc) + timedelta(minutes=5)
    assert _save(client, issuer_id).status_code == 200
    overridden = _set_override(
        client,
        issuer_id,
        expected_revision=1,
        expires_at=expires_at,
    )
    assert overridden.status_code == 200, overridden.text
    content = _export(client, issuer_id)
    preview_response = _preview(client, issuer_id, content, expected_revision=2)
    assert preview_response.status_code == 200, preview_response.text
    preview = preview_response.json()
    token = preview["preview_token"]
    assert token
    encoded = token.rsplit(".", 1)[0]
    claim = json.loads(base64.urlsafe_b64decode(encoded + "=" * (-len(encoded) % 4)))
    assert claim["exp"] <= int(expires_at.timestamp())
    assert claim["exp"] < int(datetime.now(timezone.utc).timestamp()) + 15 * 60

    before = _run(_counts(owner, issuer_id))
    before_files = _vault_files()
    with monkeypatch.context() as patcher:
        patcher.setattr(route.time, "time", lambda: claim["exp"] + 1)
        with pytest.raises(HTTPException) as expired:
            route._require_token_live(claim)
    assert expired.value.status_code == 409

    rechecks: list[dict] = []

    def expire_after_parse(token_claim: dict) -> None:
        rechecks.append(token_claim)
        raise HTTPException(
            409,
            "Preview token expired while the workbook was revalidated; preview again.",
        )

    with monkeypatch.context() as patcher:
        patcher.setattr(route, "_require_token_live", expire_after_parse)
        expired_during_reparse = _commit(
            client, issuer_id, content, preview, expected_revision=2
        )
    assert expired_during_reparse.status_code == 409
    assert expired_during_reparse.json() == {
        "detail": "Preview token expired while the workbook was revalidated; preview again."
    }
    assert rechecks == [claim]
    assert _run(_counts(owner, issuer_id)) == before
    assert _vault_files() == before_files


def test_commit_revalidates_exact_calculation_and_atomically_creates_artifacts(
    client, monkeypatch
):
    from config import get_settings
    from routes import model_workbook as route

    owner = f"mwi-commit-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    assert _save(client, issuer_id).status_code == 200
    content = _export(client, issuer_id)
    scans: list[bytes] = []

    async def record_scan(payload: bytes) -> None:
        scans.append(payload)

    monkeypatch.setattr(route.avscan, "scan", record_scan)
    preview_response = _preview(client, issuer_id, content, expected_revision=1)
    assert preview_response.status_code == 200, preview_response.text
    preview = preview_response.json()
    assert preview["calculation"] is not None
    preview_hash = preview["calculation"]["calculation_hash"]
    equivalent_response = _preview(client, issuer_id, content, expected_revision=1)
    assert equivalent_response.status_code == 200, equivalent_response.text
    equivalent = equivalent_response.json()
    assert equivalent["calculation"]["calculation_hash"] == preview_hash
    assert equivalent["draft_payload"] == preview["draft_payload"]

    committed = _commit(client, issuer_id, content, preview, expected_revision=1)
    assert committed.status_code == 200, committed.text
    body = committed.json()
    assert scans == [content, content, content]
    assert body["existing"] is False
    assert body["calculation_hash"] == preview_hash
    assert body["record"]["calculation_hash"] == preview_hash
    assert body["record"]["revision"] == 2
    assert body["record"]["source_run_id"] is None
    for period in body["record"]["payload"]["periods"]:
        assert period["authority"]["origin"] == "imported"
        assert period["authority"]["method"] == "model-workbook-import"
        assert body["document_id"] in period["authority"]["source_ids"]
        assert body["source_manifest_id"] in period["authority"]["source_ids"]
    for instrument in body["record"]["payload"]["debt_instruments"]:
        assert instrument["authority"]["origin"] == "imported"
        assert instrument["authority"]["method"] == "model-workbook-import"

    rows = _run(_artifact_rows(body["import_id"]))
    imported = rows["import"]
    document = rows["document"]
    manifest = rows["manifest"]
    assert rows["lineage"] is not None
    assert imported.document_id == body["document_id"]
    assert imported.source_manifest_id == body["source_manifest_id"]
    assert imported.committed_revision == 2
    assert imported.calculation_hash == preview_hash
    assert imported.mapping == {}
    assert document.analyst_id == owner
    assert document.issuer_id == issuer_id
    assert document.doc_type == "ModelWorkbook"
    assert manifest.analyst_id == owner
    assert manifest.issuer_id == issuer_id
    assert manifest.authority["calculation_hash"] == preview_hash
    assert manifest.authority["version_id"] == body["import_id"]
    stored = Path(get_settings().caos_storage_dir) / document.storage_key
    assert stored.read_bytes() == content
    assert hashlib.sha256(stored.read_bytes()).hexdigest() == body["workbook_sha256"]

    counts = _run(_counts(owner, issuer_id))
    files = _vault_files()
    replay = _commit(client, issuer_id, content, equivalent, expected_revision=1)
    assert replay.status_code == 200, replay.text
    replay_body = replay.json()
    assert replay_body["existing"] is True
    assert replay_body["import_id"] == body["import_id"]
    assert replay_body["document_id"] == body["document_id"]
    assert replay_body["source_manifest_id"] == body["source_manifest_id"]
    assert (
        replay_body["calculation_hash"] == equivalent["calculation"]["calculation_hash"]
    )
    assert scans == [content, content, content, content]
    assert _run(_counts(owner, issuer_id)) == counts
    assert _vault_files() == files


def test_ambiguous_commit_retains_raw_workbook_referenced_by_durable_rows(
    client, monkeypatch
):
    from routes import model_workbook as route

    owner = f"mwi-ambiguous-commit-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    assert _save(client, issuer_id).status_code == 200
    content = _export(client, issuer_id)
    preview_response = _preview(client, issuer_id, content, expected_revision=1)
    assert preview_response.status_code == 200, preview_response.text
    before = _run(_counts(owner, issuer_id))
    before_files = _vault_files()
    original_commit = route.AsyncSession.commit

    async def commit_then_lose_ack(session):
        await original_commit(session)
        raise TimeoutError("simulated lost database commit acknowledgement")

    with monkeypatch.context() as patcher:
        patcher.setattr(route.AsyncSession, "commit", commit_then_lose_ack)
        with pytest.raises(
            TimeoutError, match="simulated lost database commit acknowledgement"
        ):
            _commit(
                client,
                issuer_id,
                content,
                preview_response.json(),
                expected_revision=1,
            )

    after = _run(_counts(owner, issuer_id))
    after_files = _vault_files()
    assert after["documents"] == before["documents"] + 1
    assert after["manifests"] == before["manifests"] + 1
    assert after["imports"] == before["imports"] + 1
    assert len(after_files) == len(before_files) + 1
    stored = next(path for path in after_files if path not in before_files)
    assert stored.read_bytes() == content


def test_restoring_pre_import_checkpoint_drops_superseded_import_lineage(client):
    owner = f"mwi-restore-lineage-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    context_id = _run(_seed_context(owner, issuer_id))
    _as(owner)
    saved = _save(client, issuer_id, context_id=context_id)
    assert saved.status_code == 200, saved.text

    original_checkpoint = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints",
        json={
            "context_id": context_id,
            "label": "Before workbook import",
            "expected_revision": 1,
            "calculation_hash": saved.json()["calculation_hash"],
        },
    )
    assert original_checkpoint.status_code == 201, original_checkpoint.text

    content = _export(client, issuer_id)
    preview_response = _preview(client, issuer_id, content, expected_revision=2)
    assert preview_response.status_code == 200, preview_response.text
    imported = _commit(
        client,
        issuer_id,
        content,
        preview_response.json(),
        expected_revision=2,
    )
    assert imported.status_code == 200, imported.text
    imported_body = imported.json()

    restored = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints/"
        f"{original_checkpoint.json()['id']}/restore",
        json={"expected_revision": 3},
    )
    assert restored.status_code == 200, restored.text
    assert restored.json()["revision"] == 4
    assert imported_body["document_id"] not in restored.json()["payload"]["source_ids"]
    assert (
        imported_body["source_manifest_id"]
        not in restored.json()["payload"]["source_ids"]
    )

    post_restore = client.post(
        f"/api/models/v2/{issuer_id}/checkpoints",
        json={
            "context_id": context_id,
            "label": "After restoring pre-import model",
            "expected_revision": 4,
            "calculation_hash": restored.json()["calculation_hash"],
        },
    )
    assert post_restore.status_code == 201, post_restore.text
    parents = _run(_checkpoint_parent_ids(owner, post_restore.json()["id"]))
    assert f"source_manifest:{imported_body['source_manifest_id']}" not in parents
    assert f"document:{imported_body['document_id']}" not in parents


def test_commit_denies_read_only_before_rescan_or_write(client, monkeypatch):
    from routes import model_workbook as route

    owner = f"mwi-viewer-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    assert _save(client, issuer_id).status_code == 200
    content = _export(client, issuer_id)
    preview_response = _preview(client, issuer_id, content, expected_revision=1)
    assert preview_response.status_code == 200, preview_response.text
    preview = preview_response.json()
    before = _run(_counts(owner, issuer_id))
    before_files = _vault_files()
    scans: list[bytes] = []

    async def record_scan(payload: bytes) -> None:
        scans.append(payload)

    monkeypatch.setattr(route.avscan, "scan", record_scan)
    _as(owner, role="viewer")
    denied_preview = _preview(
        client, issuer_id, content, expected_revision=1
    )
    assert denied_preview.status_code == 403
    denied = _commit(client, issuer_id, content, preview, expected_revision=1)
    assert denied.status_code == 403
    assert scans == []
    assert _run(_counts(owner, issuer_id)) == before
    assert _vault_files() == before_files


def test_cas_conflict_has_no_partial_rows_or_storage(client):
    owner = f"mwi-cas-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    assert _save(client, issuer_id).status_code == 200
    content = _export(client, issuer_id)
    preview_response = _preview(client, issuer_id, content, expected_revision=1)
    assert preview_response.status_code == 200, preview_response.text
    preview = preview_response.json()
    advanced_payload = _payload()
    advanced_payload["ui_preferences"] = {"collapsed_rows": ["cas-advanced"]}
    advanced = _save(
        client,
        issuer_id,
        expected_revision=1,
        payload=advanced_payload,
    )
    assert advanced.status_code == 200, advanced.text
    assert advanced.json()["revision"] == 2
    before = _run(_counts(owner, issuer_id))
    before_files = _vault_files()

    conflicted = _commit(client, issuer_id, content, preview, expected_revision=1)
    assert conflicted.status_code == 409
    assert conflicted.json()["detail"]["current_revision"] == 2
    assert _run(_counts(owner, issuer_id)) == before
    assert _vault_files() == before_files
    current = client.get(f"/api/models/v2/{issuer_id}").json()["record"]
    assert current["revision"] == 2
    assert current["payload"]["ui_preferences"]["collapsed_rows"] == [
        "cas-advanced"
    ]


def test_lineage_failure_rolls_back_cas_rows_and_unique_storage(client, monkeypatch):
    from routes import model_workbook as route

    owner = f"mwi-rollback-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    assert _save(client, issuer_id).status_code == 200
    content = _export(client, issuer_id)
    preview_response = _preview(client, issuer_id, content, expected_revision=1)
    assert preview_response.status_code == 200, preview_response.text
    preview = preview_response.json()
    before = _run(_counts(owner, issuer_id))
    before_files = _vault_files()
    original = client.get(f"/api/models/v2/{issuer_id}").json()["record"]

    async def fail_lineage(*_args, **_kwargs):
        raise RuntimeError("forced model workbook lineage failure")

    monkeypatch.setattr(route, "write_owned_artifact_lineage_edge", fail_lineage)
    with pytest.raises(RuntimeError, match="forced model workbook lineage failure"):
        _commit(client, issuer_id, content, preview, expected_revision=1)
    assert _run(_counts(owner, issuer_id)) == before
    assert _vault_files() == before_files
    current = client.get(f"/api/models/v2/{issuer_id}").json()["record"]
    assert current["revision"] == original["revision"] == 1
    assert current["calculation_hash"] == original["calculation_hash"]


def test_formula_and_mapping_ambiguity_are_blocking_and_not_committable(client):
    owner = f"mwi-blockers-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    assert _save(client, issuer_id).status_code == 200
    content = _formula_without_cache(_export(client, issuer_id))
    formula_response = _preview(client, issuer_id, content, expected_revision=1)
    assert formula_response.status_code == 200, formula_response.text
    formula_preview = formula_response.json()
    assert formula_preview["blocking_count"] > 0
    assert formula_preview["preview_token"] is None
    assert "formula_cache_required" in {
        issue["code"] for issue in formula_preview["issues"]
    }
    assert (
        _commit(
            client, issuer_id, content, formula_preview, expected_revision=1
        ).status_code
        == 422
    )

    legacy = _legacy_workbook(duplicate_period_header=True)
    mapping = _legacy_mapping()
    mapping["assumptions"]["columns"]["label"] = "Period"
    ambiguity_response = _preview(
        client,
        issuer_id,
        legacy,
        expected_revision=1,
        mapping=mapping,
    )
    assert ambiguity_response.status_code == 200, ambiguity_response.text
    ambiguity = ambiguity_response.json()
    assert ambiguity["blocking_count"] > 0
    assert ambiguity["ambiguities"]
    assert ambiguity["preview_token"] is None
    assert "ambiguous_mapping" in {issue["code"] for issue in ambiguity["issues"]}
    assert _run(_counts(owner, issuer_id))["imports"] == 0


def test_same_legacy_workbook_replays_once_but_can_be_reimported_at_later_revision(
    client,
):
    owner = f"mwi-idempotent-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    content = _legacy_workbook()
    mapping = _legacy_mapping()
    first_preview_response = _preview(
        client,
        issuer_id,
        content,
        expected_revision=0,
        mapping=mapping,
    )
    assert first_preview_response.status_code == 200, first_preview_response.text
    first_preview = first_preview_response.json()
    first = _commit(
        client,
        issuer_id,
        content,
        first_preview,
        expected_revision=0,
        mapping=mapping,
    )
    assert first.status_code == 200, first.text
    assert first.json()["record"]["revision"] == 1
    replay = _commit(
        client,
        issuer_id,
        content,
        first_preview,
        expected_revision=0,
        mapping=mapping,
    )
    assert replay.status_code == 200, replay.text
    assert replay.json()["existing"] is True
    assert replay.json()["import_id"] == first.json()["import_id"]

    advanced_payload = client.get(
        f"/api/models/v2/{issuer_id}"
    ).json()["record"]["payload"]
    advanced_payload["ui_preferences"]["collapsed_rows"] = [
        "post-import-advanced"
    ]
    advanced = _save(
        client,
        issuer_id,
        expected_revision=1,
        payload=advanced_payload,
    )
    assert advanced.status_code == 200, advanced.text
    assert advanced.json()["revision"] == 2
    superseded_replay = _commit(
        client,
        issuer_id,
        content,
        first_preview,
        expected_revision=0,
        mapping=mapping,
    )
    assert superseded_replay.status_code == 409
    assert "superseded" in superseded_replay.json()["detail"]
    second_preview_response = _preview(
        client,
        issuer_id,
        content,
        expected_revision=2,
        mapping=mapping,
    )
    assert second_preview_response.status_code == 200, second_preview_response.text
    second_preview = second_preview_response.json()
    second = _commit(
        client,
        issuer_id,
        content,
        second_preview,
        expected_revision=2,
        mapping=mapping,
    )
    assert second.status_code == 200, second.text
    assert second.json()["existing"] is False
    assert second.json()["import_id"] != first.json()["import_id"]
    assert second.json()["record"]["revision"] == 3
    assert _run(_counts(owner, issuer_id))["imports"] == 2


def test_matrix_preview_commit_identity_binds_normalized_mapping_exactly(client):
    owner = f"mwi-matrix-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    content = _matrix_workbook()
    mapping = _matrix_mapping()

    preview_response = _preview(
        client,
        issuer_id,
        content,
        expected_revision=0,
        mapping=mapping,
    )
    assert preview_response.status_code == 200, preview_response.text
    preview = preview_response.json()
    assert preview["blocking_count"] == 0
    assert preview["preview_token"]
    assert list(preview["mapping"]["assumptions"]["period_columns"]) == [
        "LTM-2026-06-30",
        "FY2026",
    ]

    changed_mapping = json.loads(json.dumps(mapping))
    changed_mapping["assumptions"]["period_labels"]["FY26E"] = "FY 2026 edited"
    mismatch = _commit(
        client,
        issuer_id,
        content,
        preview,
        expected_revision=0,
        mapping=changed_mapping,
    )
    assert mismatch.status_code == 409

    committed = _commit(
        client,
        issuer_id,
        content,
        preview,
        expected_revision=0,
        mapping=mapping,
    )
    assert committed.status_code == 200, committed.text
    body = committed.json()
    assert body["record"]["revision"] == 1
    assert [
        period["period_key"] for period in body["record"]["payload"]["periods"]
    ] == ["LTM-2026-06-30", "FY2026"]


def test_owner_and_team_isolation_fail_non_enumerably_before_file_scan(
    client, monkeypatch
):
    from config import get_settings
    from routes import model_workbook as route

    owner = f"mwi-owner-{uuid.uuid4().hex[:10]}"
    other = f"mwi-other-{uuid.uuid4().hex[:10]}"
    issuer_id = _run(_seed_issuer(owner))
    _as(owner)
    assert _save(client, issuer_id).status_code == 200
    content = _export(client, issuer_id)
    preview_response = _preview(client, issuer_id, content, expected_revision=1)
    assert preview_response.status_code == 200, preview_response.text
    preview = preview_response.json()

    _as(other)
    assert client.get(f"/api/models/v2/{issuer_id}/workbook/export").status_code == 404
    foreign_commit = _commit(client, issuer_id, content, preview, expected_revision=1)
    assert foreign_commit.status_code == 409

    team_issuer_id = _run(_seed_issuer(owner, team_id="team-a"))
    get_settings().caos_tenancy_enabled = True
    scans: list[bytes] = []

    async def record_scan(payload: bytes) -> None:
        scans.append(payload)

    monkeypatch.setattr(route.avscan, "scan", record_scan)
    _as(other, team_id="team-b")
    hidden = _preview(
        client,
        team_issuer_id,
        content,
        expected_revision=0,
    )
    assert hidden.status_code == 404
    assert hidden.json() == {"detail": "Issuer not found"}
    assert scans == []


def test_0057_backfills_scoped_revision_fingerprints_and_refuses_evidence_downgrade(
    tmp_path,
):
    from test_migrations import _alembic

    db_path = tmp_path / "model-workbook-import-idempotency.db"
    db_url = f"sqlite+aiosqlite:///{db_path}"
    upgraded = _alembic("upgrade", "0056", db_url=db_url)
    assert upgraded.returncode == 0, upgraded.stderr
    mapping = {"mode": "mapped_legacy", "source_ids": ["fixture"]}
    with sqlite3.connect(db_path) as connection:
        for import_id in ("import-a", "import-b"):
            connection.execute(
                """INSERT INTO model_workbook_imports
                (id, analyst_id, issuer_id, draft_id, document_id,
                 source_manifest_id, workbook_sha256, mapping, issues,
                 committed_revision, calculation_hash, committed_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    import_id,
                    "migration-owner",
                    "migration-issuer",
                    f"draft-{import_id}",
                    f"document-{import_id}",
                    f"manifest-{import_id}",
                    "a" * 64,
                    json.dumps(mapping),
                    "[]",
                    3,
                    "b" * 64,
                    "2026-07-14T00:00:00Z",
                ),
            )
        connection.commit()
    migration = _alembic("upgrade", "0057", db_url=db_url)
    assert migration.returncode == 0, migration.stderr
    canonical = json.dumps(
        {
            "analyst_id": "migration-owner",
            "issuer_id": "migration-issuer",
            "workbook_sha256": "a" * 64,
            "mapping": mapping,
            "expected_revision": 2,
        },
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    )
    expected = hashlib.sha256(canonical.encode("utf-8")).hexdigest()
    with sqlite3.connect(db_path) as connection:
        columns = {
            row[1]: row[3]
            for row in connection.execute("PRAGMA table_info(model_workbook_imports)")
        }
        fingerprints = connection.execute(
            "SELECT id, import_fingerprint FROM model_workbook_imports ORDER BY id"
        ).fetchall()
        indexes = {
            row[1]: row[2]
            for row in connection.execute("PRAGMA index_list(model_workbook_imports)")
        }
    assert columns["import_fingerprint"] == 1
    assert fingerprints[0] == ("import-a", expected)
    assert fingerprints[1][1] != expected
    assert len({row[1] for row in fingerprints}) == 2
    assert indexes["uq_model_workbook_imports_fingerprint"] == 1
    refused = _alembic("downgrade", "0056", db_url=db_url)
    assert refused.returncode != 0
    assert "0057 downgrade refused" in refused.stderr
