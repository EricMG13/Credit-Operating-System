"""Phase 7 synthetic release journey across canonical artifact identities.

This is publication-plumbing evidence, not a gold credit view. The uploaded
text comes only from fixtures labelled SYNTHETIC TEST DOCUMENT, and the test
promotes the completed run to Committee Ready only after validating the real
module outputs. That promotion exercises report gating without representing a
CP-5 investment conclusion.
"""

from __future__ import annotations

import asyncio
import io
import json
from pathlib import Path
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from conftest import wait_for_run


REFERENCE_ISSUER_ID = "a71f0000-0000-0000-0000-000000000001"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _pdf() -> bytes:
    return b"%PDF-1.4\n1 0 obj\n<<>>\nendobj\ntrailer\n<<>>\n%%EOF"


def _market_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Market"
    sheet.append([
        "FIGI", "Borrower", "Instrument", "Currency", "Price",
        "Discount Margin", "As Of",
    ])
    sheet.append([
        "BBGPHASE7", "Atlas Forge Industries", "Phase 7 TLB", "USD",
        99.5, 425, "2026-07-14",
    ])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _identity(owner: str):
    from identity import CallerIdentity

    return CallerIdentity(
        id=owner,
        email=f"{owner}@test.local",
        full_name="Phase 7 Release Fixture",
        role="analyst",
        source="profile",
    )


def test_phase7_source_market_run_modules_model_report_xlsx_identity(
    monkeypatch,
):
    import ingest
    from config import get_settings
    from identity import get_identity
    from main import app
    import routes.ingestion as ingestion_route

    async def clean_scan(_content: bytes):
        return "clean"

    monkeypatch.setattr(ingestion_route.avscan, "scan", clean_scan)

    owner = f"phase7-{uuid4().hex}"
    settings = get_settings()
    for flag in (
        "caos_lineage_v2_enabled",
        "caos_market_xlsx_v2_enabled",
        "caos_model_engine_v2_enabled",
        "caos_cp_4d_enabled",
        "caos_cp_2g_enabled",
    ):
        monkeypatch.setattr(settings, flag, True)

    fixture_dir = Path(__file__).parents[1] / "fixtures" / "modules"
    source_text = "\n\n".join([
        (fixture_dir / "cp-4d-synthetic-structure.txt").read_text(),
        (fixture_dir / "cp-2g-synthetic-disclosure.txt").read_text(),
    ])
    assert source_text.count("SYNTHETIC TEST DOCUMENT") >= 2
    monkeypatch.setattr(
        ingest,
        "extract_pdf_text",
        lambda content, filename="phase7.pdf": (source_text, False),
    )
    app.dependency_overrides[get_identity] = lambda: _identity(owner)

    try:
        with TestClient(app) as client:
            context_response = client.post("/api/analysis/contexts", json={
                "name": "Phase 7 synthetic release journey",
                "issuer_ids": [REFERENCE_ISSUER_ID],
            })
            assert context_response.status_code == 201, context_response.text
            context_id = context_response.json()["id"]

            upload = client.post(
                "/api/ingestion/upload/document",
                data={
                    "issuer_id": REFERENCE_ISSUER_ID,
                    "context_id": context_id,
                    "run_mode": "legal",
                    "source_kind": "legal_document",
                    "fiscal_period": "FY2025",
                },
                files={"file": ("phase7-synthetic.pdf", _pdf(), "application/pdf")},
            )
            assert upload.status_code == 200, upload.text
            uploaded = upload.json()

            market_bytes = _market_workbook()
            issuer_mappings = json.dumps({"BBGPHASE7": REFERENCE_ISSUER_ID})
            preview = client.post(
                "/api/rv/snapshots/import/preview",
                files={"file": ("phase7-market.xlsx", market_bytes, XLSX_MIME)},
                data={"mapping": "{}", "issuer_mappings": issuer_mappings},
            )
            assert preview.status_code == 200, preview.text
            preview_body = preview.json()
            assert preview_body["blocking_count"] == 0
            commit = client.post(
                "/api/rv/snapshots/import/commit",
                files={"file": ("phase7-market.xlsx", market_bytes, XLSX_MIME)},
                data={
                    "mapping": "{}",
                    "issuer_mappings": issuer_mappings,
                    "preview_sha256": preview_body["workbook_sha256"],
                    "preview_token": preview_body["preview_token"],
                    "source_label": "Phase 7 synthetic Bloomberg-style fixture",
                },
            )
            assert commit.status_code == 200, commit.text
            market = commit.json()

            screen = client.post("/api/rv/screens", json={
                "context_id": context_id,
                "snapshot_id": market["snapshot_id"],
            })
            assert screen.status_code == 201, screen.text
            screen_body = screen.json()
            assert screen_body["candidates"][0]["market"]["price"] == 99.5
            assert screen_body["candidates"][0]["market"]["dm"] == 425.0

            created = client.post("/api/runs", json={
                "issuer_id": REFERENCE_ISSUER_ID,
                "context_id": context_id,
                "as_of_date": "2026-07-14",
            })
            assert created.status_code == 201, created.text
            run_id = created.json()["id"]
            finished = wait_for_run(client, run_id, timeout_s=60.0)
            assert finished["status"] == "complete"

            modules_response = client.get(f"/api/runs/{run_id}/modules")
            assert modules_response.status_code == 200, modules_response.text
            modules = {
                row["module_id"]: row for row in modules_response.json()
            }
            for module_id, schema_version in (
                ("CP-4D", "cp-4d.v1"),
                ("CP-2G", "cp-2g.v1"),
            ):
                assert module_id in modules
                runtime = modules[module_id]["runtime_output"]
                assert runtime["schema_version"] == schema_version
                assert len(runtime["prompt_bundle_fingerprint"]) == 64
                assert runtime["prompt_bundle_files"]
                assert runtime["module_status"] in {
                    "Blocked", "Completed with Limitations", "Not Applicable",
                }
                for claim in modules[module_id]["claims"]:
                    assert claim["evidence"]

            context = client.get(f"/api/analysis/contexts/{context_id}").json()
            typed_refs = {
                (ref["kind"], ref["id"])
                for ref in context["artifacts"]["artifact_refs"]
            }
            assert ("document", uploaded["document_id"]) in typed_refs
            assert ("source_manifest", uploaded["source_manifest_id"]) in typed_refs
            assert ("market_snapshot", market["snapshot_id"]) in typed_refs
            assert ("issuer_run", run_id) in typed_refs
            lineage = client.get(
                f"/api/analysis/contexts/{context_id}/lineage"
            ).json()["edges"]
            run_parents = {
                (edge["parent"]["kind"], edge["parent"]["id"])
                for edge in lineage if edge["artifact"]["id"] == run_id
            }
            assert ("document", uploaded["document_id"]) in run_parents
            assert ("source_manifest", uploaded["source_manifest_id"]) in run_parents
            assert ("market_snapshot", market["snapshot_id"]) in run_parents

            notifications = client.get("/api/notifications").json()["items"]
            run_events = [
                event for event in notifications if event["subject_id"] == run_id
            ]
            assert len(run_events) == 1
            assert f"run={run_id}" in run_events[0]["href"]

            # Publication-plumbing only: this is not a CP-5 conclusion.
            asyncio.run(_promote_synthetic_run_for_publication(run_id))

            model_read = client.get(
                f"/api/models/v2/{REFERENCE_ISSUER_ID}",
                params={"run_id": run_id},
            )
            assert model_read.status_code == 200, model_read.text
            suggestion = model_read.json()
            assert suggestion["availability"] == "suggested"
            assert suggestion["suggested_source_run_id"] == run_id
            save = client.put(f"/api/models/v2/{REFERENCE_ISSUER_ID}", json={
                "expected_revision": 0,
                "context_id": context_id,
                "source_run_id": run_id,
                "payload": suggestion["suggested_payload"],
            })
            assert save.status_code == 200, save.text
            saved = save.json()
            assert saved["source_run_id"] == run_id

            checkpoint = client.post(
                f"/api/models/v2/{REFERENCE_ISSUER_ID}/checkpoints",
                json={
                    "context_id": context_id,
                    "label": "Phase 7 synthetic release checkpoint",
                    "issuer_run_id": run_id,
                    "expected_revision": saved["revision"],
                    "calculation_hash": saved["calculation_hash"],
                },
            )
            assert checkpoint.status_code == 201, checkpoint.text
            checkpoint_body = checkpoint.json()

            editorial = {
                "deliverable_id": "live-committee-pack",
                "source_run_id": run_id,
                "omit": {},
                "edits": {},
                "show_sources": True,
                "hide_addbacks": False,
            }
            report_preview = client.post("/api/reports/versions/preview", json={
                "context_id": context_id,
                "run_id": run_id,
                "model_checkpoint_id": checkpoint_body["id"],
                "payload": editorial,
            })
            assert report_preview.status_code == 200, report_preview.text
            frozen = report_preview.json()
            assert frozen["model_calculation_hash"] == saved["calculation_hash"]
            published = client.post("/api/reports/versions", json={
                "context_id": context_id,
                "run_id": run_id,
                "model_checkpoint_id": checkpoint_body["id"],
                "payload": frozen["payload"]["composition"]["editorial"],
                "preview_sha256": frozen["preview_sha256"],
            })
            assert published.status_code == 201, published.text
            report = published.json()
            assert report["run_id"] == run_id
            assert report["model_checkpoint_id"] == checkpoint_body["id"]
            assert report["model_calculation_hash"] == saved["calculation_hash"]

            exported = client.post(
                f"/api/reports/versions/{report['id']}/export?format=xlsx"
            )
            assert exported.status_code == 200, exported.text
            workbook = load_workbook(io.BytesIO(exported.content), read_only=True)
            assert workbook["Cover"]["B24"].value == saved["calculation_hash"]
            first_model_row = next(
                workbook["Model"].iter_rows(min_row=2, values_only=True)
            )
            assert first_model_row[0]
            assert first_model_row[3] is not None
            workbook.close()

            final_lineage = client.get(
                f"/api/analysis/contexts/{context_id}/lineage"
            ).json()["edges"]
            report_parents = {
                (edge["parent"]["kind"], edge["parent"]["id"])
                for edge in final_lineage if edge["artifact"]["id"] == report["id"]
            }
            assert report_parents == {
                ("issuer_run", run_id),
                ("model_checkpoint", checkpoint_body["id"]),
                ("source_manifest", uploaded["source_manifest_id"]),
            }
    finally:
        app.dependency_overrides.pop(get_identity, None)


async def _promote_synthetic_run_for_publication(run_id: str) -> None:
    """Test-only report gate setup after real module execution is verified."""
    from datetime import datetime, timezone

    from database import (
        AsyncSessionLocal,
        IssuerReportingProfile,
        ModuleOutput,
        Run,
    )
    from sqlalchemy import select

    async with AsyncSessionLocal() as db:
        run = await db.get(Run, run_id)
        assert run is not None and run.status == "complete"
        run.qa_status = "Passed"
        run.committee_status = "Committee Ready"
        profile = await db.get(IssuerReportingProfile, run.issuer_id)
        if profile is None:
            db.add(IssuerReportingProfile(
                issuer_id=run.issuer_id,
                cadence="quarterly",
                fiscal_year_end_month=12,
                fiscal_year_end_day=31,
                reporting_lag_days=45,
                grace_days=7,
                authority={"origin": "synthetic-test-fixture"},
                updated_by=run.analyst_id,
                updated_at=datetime.now(timezone.utc),
            ))
        else:
            profile.fiscal_year_end_month = 12
            profile.fiscal_year_end_day = 31
        cp1 = (await db.execute(select(ModuleOutput).where(
            ModuleOutput.run_id == run_id,
            ModuleOutput.module_id == "CP-1",
        ))).scalar_one()
        cp1_output = dict(cp1.runtime_output or {})
        cp1_output.setdefault("currency", "USD")
        cp1_output.setdefault("reporting_unit", "millions")
        # The legal/ESG fixture intentionally contains no financial statement.
        # Supply a labelled synthetic CP-1 bridge only after the real run and
        # specialized-module assertions, so Model v2/report plumbing can be
        # exercised without treating invented figures as analytical output.
        cp1_output["normalized_financials"] = {
            "revenue": {"FY2025": 900.0, "LTM_Q1_26": 950.0},
            "adj_ebitda": {"FY2025": 95.0, "LTM_Q1_26": 100.0},
            "net_debt_ltm": 400.0,
            "interest_coverage_ltm": 2.0,
        }
        cp1.runtime_output = cp1_output
        await db.commit()
