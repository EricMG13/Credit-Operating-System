from io import BytesIO
import hashlib
import json
from datetime import datetime, timezone
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from pypdf import PdfReader
from fastapi.testclient import TestClient

from report_exports import render_report_pdf, render_report_xlsx
from model_engine_v2 import (
    CellOverride,
    DebtInstrument,
    DebtPeriod,
    ModelAuthority,
    ModelDraftPayload,
    ModelPeriodInput,
    calculate_model,
)


PAYLOAD = {
    "document": {
        "run_id": "run-1",
        "issuer_id": "issuer-1",
        "as_of_date": "2026-06-30",
        "qa_status": "Passed",
        "committee_status": "Committee Ready",
        "prepared_by": "analyst-1",
        "sections": [
            {
                "module_id": "CP-1",
                "module_name": "Credit foundation",
                "confidence": 0.91,
                "qa_status": "Passed",
                "summary": {
                    "normalized_financials": {
                        "revenue": {"FY2025": 1_250.0},
                        "net_leverage_adj_ltm": 5.2,
                    },
                    "limitations": ["No covenant EBITDA bridge"],
                },
            },
            {
                "module_id": "CP-4C",
                "module_name": "Capacity",
                "confidence": 0.78,
                "qa_status": "Restricted",
                "summary": {"headroom": 85.0, "unit": "$M"},
            },
        ],
    },
}
AUTHORITY = {
    "origin": "live",
    "as_of": "2026-07-13T12:00:00Z",
    "approval_state": "published",
    "source_ids": ["manifest-1", "run-1", "checkpoint-1"],
    "freshness": "due",
    "freshness_evaluation": {
        "state": "due",
        "source_kind": "derived_artifact",
        "observed_at": "2026-06-01T00:00:00Z",
        "effective_period_end": None,
        "expected_next_at": None,
        "due_at": "2026-07-15T00:00:00Z",
        "age_days": 42,
        "reason": "derived_artifact_refresh_due",
        "policy_version": "caos-freshness-v1",
    },
}


def _payload_hash(payload: dict) -> str:
    canonical = json.dumps(
        payload, sort_keys=True, separators=(",", ":"), allow_nan=False, default=str
    )
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _v2_report_payload() -> dict:
    authority = ModelAuthority(origin="live", method="test", source_ids=["run-1"])
    draft = ModelDraftPayload(
        reporting_currency="GBP",
        reporting_unit="millions",
        periods=[ModelPeriodInput(
            period_key="FY2026", label="FY26e", kind="forecast", revenue=800,
            reported_ebitda=100, adjustments=10, cash=20, taxes=5, capex=10,
            working_capital_change=-2, other_cash_flow=0, authority=authority,
        )],
        debt_instruments=[DebtInstrument(
            instrument_id="tlb-1", name="First-lien term loan", priority=1,
            seniority="1L", currency="GBP", rate_type="floating",
            authority=authority,
            periods=[DebtPeriod(
                period_key="FY2026", opening_balance=200, closing_balance=190,
                draws=0, repayments=10, scheduled_amortization=0,
                commitment=200, benchmark_rate=0.05, spread_rate=0.03,
                commitment_fee_rate=0, pik_rate=0, cash_fees=1,
                hedge_effect=0, fx_rate=1,
            )],
        )],
        overrides=[
            CellOverride(
                node_id="calc:FY2026:adjusted_ebitda",
                value_type="number",
                value=110,
                reason="IC EBITDA bridge",
                scope="draft",
                source="Analyst bridge",
                expires_at=datetime(2026, 12, 31, tzinfo=timezone.utc),
            ),
            CellOverride(
                node_id="input:FY2026:cash",
                value_type="number",
                value=20,
                reason="Expired cash check",
                scope="draft",
                source="Prior review",
                expires_at=datetime(2026, 7, 13, tzinfo=timezone.utc),
            ),
        ],
        source_ids=["run-1"],
    )
    calculation = calculate_model(
        draft, evaluated_at=datetime(2026, 7, 14, tzinfo=timezone.utc)
    )
    return {
        **PAYLOAD,
        "model": {
            "engine_version": calculation.engine_version,
            "source_fingerprint": calculation.source_fingerprint,
            "input_fingerprint": calculation.input_fingerprint,
            "calculation_hash": calculation.calculation_hash,
            "draft_revision": 3,
            "authority": {
                "origin": "imported",
                "model_input_origins": ["imported", "analyst"],
                "analyst_override": True,
            },
            "payload": draft.model_dump(mode="json"),
            "calculation": calculation.model_dump(mode="json"),
        },
    }


def test_xlsx_export_reopens_with_typed_values_and_audit_sheets():
    content = render_report_xlsx(
        version_id="version-1",
        document_sha256="a" * 64,
        payload=PAYLOAD,
        authority=AUTHORITY,
    )
    assert content.startswith(b"PK")
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    assert workbook.sheetnames[:3] == ["Cover", "Module Summary", "CP-1"]
    assert "Sources - Audit" in workbook.sheetnames
    assert workbook["Cover"]["B4"].value == "version-1"
    assert workbook["Cover"]["B14"].value == "DUE"
    assert workbook["Cover"]["B17"].value == "caos-freshness-v1"
    assert workbook["Module Summary"]["A2"].value == "CP-1"
    cp1_values = list(workbook["CP-1"].iter_rows(min_row=4, values_only=True))
    leverage = next(value for path, value in cp1_values if path == "normalized_financials.net_leverage_adj_ltm")
    assert leverage == 5.2
    assert isinstance(leverage, float)
    assert workbook["Sources - Audit"]["A2"].value == "manifest-1"
    assert workbook["Sources - Audit"]["E2"].value == "DUE"
    workbook.close()


def test_xlsx_export_preserves_exact_model_engine_result_and_debt_components():
    payload = _v2_report_payload()
    content = render_report_xlsx(
        version_id="version-model-v2",
        document_sha256="f" * 64,
        payload=payload,
        authority=AUTHORITY,
    )
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    assert "Model" in workbook.sheetnames
    assert "Debt Schedule" in workbook.sheetnames
    assert "Model Overrides" in workbook.sheetnames
    model_rows = list(workbook["Model"].iter_rows(min_row=2, values_only=True))
    model_headers = next(workbook["Model"].iter_rows(max_row=1, values_only=True))
    cover = {
        key: value
        for key, value in workbook["Cover"].iter_rows(
            min_row=4, max_col=2, values_only=True
        )
        if key
    }
    assert model_rows[0][0] == "FY2026"
    assert model_rows[0][4] == 110
    assert model_rows[0][5] == 16.6
    assert model_rows[0][6] == 190
    assert model_rows[0][9] == pytest.approx(190 / 110)
    assert model_headers[3] == "Revenue (GBP millions)"
    debt_rows = list(workbook["Debt Schedule"].iter_rows(min_row=2, values_only=True))
    assert debt_rows[0][1] == "tlb-1"
    assert debt_rows[0][2] == "GBP"
    assert debt_rows[0][5] == 195
    assert debt_rows[0][6] == pytest.approx(9.75)
    assert debt_rows[0][7] == pytest.approx(5.85)
    assert cover["Model reporting currency"] == "GBP"
    assert cover["Model reporting unit"] == "millions"
    override_rows = list(
        workbook["Model Overrides"].iter_rows(min_row=2, values_only=True)
    )
    assert override_rows[0] == (
        "ACTIVE AT REPORT EVENT",
        "calc:FY2026:adjusted_ebitda",
        110,
        "IC EBITDA bridge",
        "draft",
        "Analyst bridge",
        "2026-12-31T00:00:00Z",
        "reported_ebitda + adjustments",
        110,
    )
    assert override_rows[1][0] == "INACTIVE AT REPORT EVENT"
    assert override_rows[1][1] == "input:FY2026:cash"
    assert override_rows[1][7] == "No formula (input)"
    assert override_rows[1][8] == 20
    assert workbook["Cover"]["B24"].value == payload["model"]["calculation_hash"]
    assert any(
        row[0] == f"model:{payload['model']['calculation_hash']}"
        for row in workbook["Sources - Audit"].iter_rows(min_row=2, values_only=True)
    )
    workbook.close()


def test_xlsx_debt_schedule_does_not_depend_on_an_override_ledger():
    payload = _v2_report_payload()
    payload["model"]["payload"]["overrides"] = []

    workbook = load_workbook(BytesIO(render_report_xlsx(
        version_id="version-model-no-overrides",
        document_sha256="6" * 64,
        payload=payload,
        authority=AUTHORITY,
    )), read_only=True)

    assert "Model" in workbook.sheetnames
    assert "Debt Schedule" in workbook.sheetnames
    assert "Model Overrides" not in workbook.sheetnames
    workbook.close()


def test_xlsx_export_neutralizes_formula_injection_in_every_text_surface():
    hostile_payload = {
        "document": {
            **PAYLOAD["document"],
            "issuer_id": '=HYPERLINK("https://example.invalid","issuer")',
            "sections": [{
                "module_id": "+SUM(1,1)",
                "module_name": "@malicious-name",
                "confidence": "-1+1",
                "qa_status": "=CMD()",
                "summary": {"=hostile.path": "@hostile-value"},
            }],
        },
    }
    hostile_authority = {
        **AUTHORITY,
        "source_ids": ["-1+cmd"],
        "origin": "=origin",
    }
    content = render_report_xlsx(
        version_id="version-hostile",
        document_sha256="9" * 64,
        payload=hostile_payload,
        authority=hostile_authority,
    )
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    assert not any(
        cell.data_type == "f"
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    assert workbook["Cover"]["B6"].value.startswith("'=")
    assert workbook["Module Summary"]["A2"].value.startswith("'+")
    assert workbook["Module Summary"]["B2"].value.startswith("'@")
    hostile_sheet = next(
        sheet for sheet in workbook.worksheets if sheet.title not in {
            "Cover", "Module Summary", "Sources - Audit"
        }
    )
    assert hostile_sheet["A4"].value.startswith("'=")
    assert hostile_sheet["B4"].value.startswith("'@")
    assert workbook["Sources - Audit"]["A2"].value.startswith("'-")
    workbook.close()


def test_pdf_export_reopens_and_carries_version_authority_and_modules():
    content = render_report_pdf(
        version_id="version-1",
        document_sha256="b" * 64,
        payload=PAYLOAD,
        authority=AUTHORITY,
    )
    assert content.startswith(b"%PDF-")
    reader = PdfReader(BytesIO(content))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "CAOS IMMUTABLE COMMITTEE REPORT" in text
    assert "version-1" in text
    assert "CP-1" in text
    assert "Committee Ready" in text
    assert "DUE" in text
    assert "caos-freshness-v1" in text
    assert len(reader.pages) >= 3


def test_reviewed_composition_and_exact_model_round_trip_in_pdf_and_xlsx():
    payload = _v2_report_payload()
    payload["composition"] = {
        "reviewed_report": {
            "id": "live-committee-pack",
            "title": "Reviewed IC memo",
            "subtitle": "issuer-1 · run run-1",
            "file": "issuer-1-memo",
            "icon": "document",
            "srcs": [],
            "sections": [{
                "t": "text",
                "title": "CP-1 · Reviewed conclusion",
                "body": "Analyst-edited downside conclusion retained.",
            }],
        },
        "editorial": {
            "source_run_id": "run-1",
            "edits": {"s1.body": "Analyst-edited downside conclusion retained."},
            "omit": {"2": True},
        },
    }
    xlsx_content = render_report_xlsx(
        version_id="version-reviewed",
        document_sha256="8" * 64,
        payload=payload,
        authority=AUTHORITY,
    )
    workbook = load_workbook(BytesIO(xlsx_content), read_only=True, data_only=False)
    reviewed_rows = list(workbook["Reviewed Report"].iter_rows(min_row=2, values_only=True))
    assert any(row[2] == "Analyst-edited downside conclusion retained." for row in reviewed_rows)
    assert not any("Capacity" in str(value) for row in reviewed_rows for value in row)
    assert workbook["Cover"]["B24"].value == payload["model"]["calculation_hash"]
    assert workbook["Cover"]["B26"].value == "imported"
    assert workbook["Model"]["E2"].value == 110
    assert workbook["Debt Schedule"]["F2"].value == 195
    workbook.close()

    pdf_content = render_report_pdf(
        version_id="version-reviewed",
        document_sha256="8" * 64,
        payload=payload,
        authority=AUTHORITY,
    )
    reader = PdfReader(BytesIO(pdf_content))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    compact_text = " ".join(text.split())
    assert "Analyst-edited downside conclusion retained." in text
    assert "Capacity" not in text
    assert payload["model"]["calculation_hash"] in text
    assert "FY2026" in text
    assert "110" in text
    assert "190" in text
    assert "tlb-1" in text
    assert "195" in text
    assert "9.75" in text
    assert "Model reporting currency" in compact_text
    assert "GBP" in text
    assert "MODEL ENGINE V2 - CALCULATION - GBP MILLIONS" in compact_text
    assert (
        "MODEL ENGINE V2 - DEBT SCHEDULE - MILLIONS - REPORTING CURRENCY GBP"
        in compact_text
    )
    assert "MODEL ENGINE V2 - OVERRIDE LEDGER" in compact_text
    assert "ACTIVE AT REPORT EVENT | calc:FY2026:adjusted_ebitda" in compact_text
    assert "INACTIVE AT REPORT EVENT | input:FY2026:cash" in compact_text
    assert "Displaced formula: reported_ebitda + adjustments" in compact_text


def test_partial_model_exports_prominent_availability_gaps_and_warnings():
    payload = _v2_report_payload()
    payload["model"]["calculation"] = {
        **payload["model"]["calculation"],
        "status": "partial",
        "gaps": ["FY2026: total debt is required"],
        "warnings": ["Manual input requires review"],
    }
    workbook = load_workbook(BytesIO(render_report_xlsx(
        version_id="version-partial",
        document_sha256="7" * 64,
        payload=payload,
        authority=AUTHORITY,
    )), read_only=True)
    ledger = list(workbook["Model Gaps - Warnings"].iter_rows(values_only=True))
    assert ("Availability", "partial") in ledger
    assert ("Gap", "FY2026: total debt is required") in ledger
    assert ("Warning", "Manual input requires review") in ledger
    workbook.close()
    pdf = PdfReader(BytesIO(render_report_pdf(
        version_id="version-partial",
        document_sha256="7" * 64,
        payload=payload,
        authority=AUTHORITY,
    )))
    text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    assert "partial" in text
    assert "FY2026: total debt is required" in text
    assert "Manual input requires review" in text


def test_insufficient_model_without_periods_keeps_xlsx_gap_ledger():
    payload = {
        **PAYLOAD,
        "model": {
            "engine_version": "2.0.0",
            "source_fingerprint": "source",
            "input_fingerprint": "input",
            "calculation_hash": "insufficient",
            "draft_revision": 1,
            "authority": {
                "origin": "live",
                "model_input_origins": ["live"],
                "analyst_override": False,
            },
            "payload": {
                "reporting_currency": "GBP",
                "reporting_unit": "millions",
                "debt_instruments": [],
            },
            "calculation": {
                "status": "insufficient_inputs",
                "gaps": ["No forecast periods are available"],
                "warnings": [],
                "periods": [],
            },
        },
    }
    workbook = load_workbook(BytesIO(render_report_xlsx(
        version_id="report-insufficient",
        document_sha256="c" * 64,
        payload=payload,
        authority=AUTHORITY,
    )), read_only=True)
    assert "Model" not in workbook.sheetnames
    assert "Model Gaps - Warnings" in workbook.sheetnames
    rows = list(workbook["Model Gaps - Warnings"].iter_rows(values_only=True))
    cover = {
        key: value
        for key, value in workbook["Cover"].iter_rows(
            min_row=4, max_col=2, values_only=True
        )
        if key
    }
    assert ("Availability", "insufficient_inputs") in rows
    assert ("Gap", "No forecast periods are available") in rows
    assert cover["Model reporting currency"] == "GBP"
    assert cover["Model reporting unit"] == "millions"
    workbook.close()
    pdf = PdfReader(BytesIO(render_report_pdf(
        version_id="report-insufficient",
        document_sha256="c" * 64,
        payload=payload,
        authority=AUTHORITY,
    )))
    text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    compact_text = " ".join(text.split())
    assert "Model reporting currency" in compact_text
    assert "Model reporting unit" in compact_text
    assert "GBP" in text
    assert "millions" in text


@pytest.mark.parametrize("state", ["due", "stale", "unknown"])
def test_binary_exports_render_each_non_current_evaluated_state(state):
    authority = {
        **AUTHORITY,
        "freshness": state,
        "freshness_evaluation": {
            **AUTHORITY["freshness_evaluation"],
            "state": state,
            "reason": f"fixture_{state}",
        },
    }
    xlsx = load_workbook(BytesIO(render_report_xlsx(
        version_id="version-state", document_sha256="e" * 64,
        payload=PAYLOAD, authority=authority,
    )), read_only=True)
    assert xlsx["Cover"]["B14"].value == state.upper()
    xlsx.close()
    pdf = PdfReader(BytesIO(render_report_pdf(
        version_id="version-state", document_sha256="e" * 64,
        payload=PAYLOAD, authority=authority,
    )))
    text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    assert state.upper() in text


@pytest.mark.asyncio
async def test_version_export_endpoint_returns_owned_binary_files_and_headers(seeded_db):
    from database import AnalysisContextRecord, AsyncSessionLocal, ModelCheckpoint, ReportVersion, Run
    from identity import CallerIdentity, get_identity
    from main import app

    issuer_id = "a71f0000-0000-0000-0000-000000000001"
    suffix = uuid4().hex
    context_id = f"report-export-context-{suffix}"
    run_id = f"report-export-run-{suffix}"
    checkpoint_id = f"report-export-checkpoint-{suffix}"
    version_id = f"report-export-version-{suffix}"
    payload_hash = _payload_hash(PAYLOAD)
    async with AsyncSessionLocal() as db:
        db.add_all([
            AnalysisContextRecord(
                id=context_id, analyst_id="report-export-owner",
                name="Export verification", issuer_ids=[issuer_id], artifacts={},
            ),
            Run(
                id=run_id, issuer_id=issuer_id, analyst_id="report-export-owner",
                status="complete", qa_status="Passed", committee_status="Committee Ready",
                completed_at=datetime.now(timezone.utc),
            ),
        ])
        await db.flush()
        db.add(ModelCheckpoint(
            id=checkpoint_id, issuer_id=issuer_id, analyst_id="report-export-owner",
            context_id=context_id, issuer_run_id=run_id,
            label="Export snapshot", payload_hash="d" * 64, payload={}, authority={},
        ))
        await db.flush()
        db.add(ReportVersion(
            id=version_id,
            context_id=context_id,
            analyst_id="report-export-owner",
            run_id=run_id,
            model_checkpoint_id=checkpoint_id,
            thesis_version_id=None,
            status="published",
            payload=PAYLOAD,
            document_sha256=payload_hash,
            authority=AUTHORITY,
            created_at=datetime.now(timezone.utc),
        ))
        await db.commit()
    app.dependency_overrides[get_identity] = lambda: CallerIdentity(
        id="report-export-owner",
        email="owner@firm.test",
        full_name="Report Owner",
        role="analyst",
        source="profile",
    )
    try:
        with TestClient(app) as client:
            xlsx = client.post(f"/api/reports/versions/{version_id}/export?format=xlsx")
            assert xlsx.status_code == 200, xlsx.text
            assert xlsx.headers["content-type"].startswith("application/vnd.openxmlformats")
            assert xlsx.headers["x-caos-document-sha256"] == payload_hash
            exported_workbook = load_workbook(BytesIO(xlsx.content), read_only=True)
            assert exported_workbook["Cover"]["B4"].value == version_id
            # No context lineage was seeded, so export must correct the legacy
            # stored state to UNKNOWN instead of fabricating CURRENT/DUE.
            assert exported_workbook["Cover"]["B14"].value == "UNKNOWN"
            exported_workbook.close()
            pdf = client.post(f"/api/reports/versions/{version_id}/export?format=pdf")
            assert pdf.status_code == 200, pdf.text
            assert pdf.headers["content-type"].startswith("application/pdf")
            pdf_reader = PdfReader(BytesIO(pdf.content))
            assert len(pdf_reader.pages) >= 3
            assert "UNKNOWN" in "\n".join(page.extract_text() or "" for page in pdf_reader.pages)

            app.dependency_overrides[get_identity] = lambda: CallerIdentity(
                id="report-export-foreign",
                email="foreign@firm.test",
                full_name="Foreign Analyst",
                role="analyst",
                source="profile",
            )
            assert client.post(f"/api/reports/versions/{version_id}/export?format=xlsx").status_code == 404
        from database import ReportVersion

        async with AsyncSessionLocal() as db:
            tampered = await db.get(ReportVersion, version_id)
            assert tampered is not None
            tampered.payload = {**PAYLOAD, "tampered": True}
            await db.commit()
        app.dependency_overrides[get_identity] = lambda: CallerIdentity(
            id="report-export-owner",
            email="owner@firm.test",
            full_name="Report Owner",
            role="analyst",
            source="profile",
        )
        with TestClient(app) as client:
            assert client.get(f"/api/reports/versions/{version_id}").status_code == 409
            assert client.post(f"/api/reports/versions/{version_id}/export?format=pdf").status_code == 409
            assert client.post(f"/api/reports/versions/{version_id}/export?format=xlsx").status_code == 409
    finally:
        app.dependency_overrides.clear()
